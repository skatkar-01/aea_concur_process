"""
src/concur_writer.py
────────────────────
Builds a formatted multi-sheet Excel report from extracted Concur expense data.
Writes to outputs/ folder with cardholder name and month as filename.

Uses consistent styling with enterprise theme matching the attached excel_writer.py:
  - Employee Report: summary & financial details
  - Transactions: all extracted expense transactions
  - Receipts: all extracted receipts/invoices
  - Reconciliation: transaction↔receipt matching with confidence scores
  - Summary: aggregate comments and reconciliation overview

Excel theme:
  - Navy header (#1C2B3A) with white text
  - Column headers (#2C3E50) with white text
  - Alternating row backgrounds for readability
  - Labeled key-value sections
  - Color-coded reconciliation status (matched=green, unmatched=red)
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.logging_config import get_logger

logger = get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Theme constants
# ─────────────────────────────────────────────────────────────────────────────

NAVY         = "1C2B3A"
COL_HDR_BG   = "2C3E50"
WHITE        = "FFFFFF"
ALT_ROW      = "F5F6F7"
LABEL_BG     = "ECEEF0"
BORDER_CLR   = "D0D5DC"
BODY_COLOR   = "2C2C2C"
MATCHED_BG   = "EBF5EB"      # Light green for matched
UNMATCHED_BG = "FDECEA"      # Light red for unmatched
WARN_BG      = "FDFAE8"      # Light yellow for medium/low confidence

_bdr    = Side(style="thin", color=BORDER_CLR)
BORDER  = Border(left=_bdr, right=_bdr, top=_bdr, bottom=_bdr)


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    """Create a solid fill with the given hex color."""
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(size: int = 10) -> Font:
    """Header font: Calibri, bold, white."""
    return Font(name="Calibri", bold=True, color=WHITE, size=size)


def _body_font(bold: bool = False, size: int = 10) -> Font:
    """Body font: Calibri, dark gray."""
    return Font(name="Calibri", bold=bold, color=BODY_COLOR, size=size)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    """Create alignment with specified horizontal, vertical, and wrap settings."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _autofit(ws, mn: int = 10, mx: int = 50) -> None:
    """Fit column widths to content, respecting min/max bounds."""
    for col in ws.columns:
        best = mn
        col_letter = None
        for cell in col:
            try:
                ltr = cell.column_letter
                if col_letter is None:
                    col_letter = ltr
                if cell.row != 1 and cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, mx))
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = best


# ─────────────────────────────────────────────────────────────────────────────
# Layout helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_banner(ws, title: str) -> None:
    """Add navy banner with title at the top of the sheet."""
    last = get_column_letter(20)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 28


def _section_header(ws, row: int, col_start: int, col_end: int, label: str) -> None:
    """Add a section header row with label spanning multiple columns."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.font      = _hdr_font(10)
    c.fill      = _fill(COL_HDR_BG)
    c.alignment = _align("left", "center")
    c.border    = BORDER


def _col_headers(ws, row: int, labels: list[str], col_start: int = 1) -> None:
    """Add column header row with title-cased labels."""
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i,
                    value=lbl.replace("_", " ").title())
        c.font      = _hdr_font()
        c.fill      = _fill(COL_HDR_BG)
        c.alignment = _align("center", "center")
        c.border    = BORDER


def _data_row(ws, row: int, values: list, alt: bool = False,
              row_fills: dict | None = None, wrap: bool = False,
              max_height: int | None = None, col_start: int = 1) -> None:
    """
    Write a data row. Handles alternating backgrounds, custom fills, text wrapping.
    
    Args:
        ws:        Worksheet
        row:       Row number
        values:    List of cell values
        alt:       Use alternating row color
        row_fills: Dict mapping column index to custom fill color
        wrap:      Enable text wrapping
        max_height: Set row height
        col_start: Starting column index (1-based)
    """
    bg = ALT_ROW if alt else WHITE
    for i, val in enumerate(values):
        col = col_start + i
        v   = val if val not in (None, "null", "NULL", "") else "—"
        if isinstance(v, (list, dict)):
            v = json.dumps(v, ensure_ascii=False)
        c   = ws.cell(row=row, column=col, value=v)
        c.fill      = _fill(row_fills.get(col, bg) if row_fills else bg)
        c.font      = _body_font()
        c.alignment = _align("left", "top" if wrap else "center", wrap=wrap)
        c.border    = BORDER
    if max_height:
        ws.row_dimensions[row].height = max_height


def _kv(ws, start_row: int, label: str, value, alt: bool = False) -> int:
    """
    Write a key-value pair in two columns (columns B & C).
    Multi-line values each get their own row. Returns number of rows consumed.
    
    Args:
        ws:        Worksheet
        start_row: Starting row number
        label:     Key label (displayed in column B, first row only)
        value:     Value (displayed in column C, may span multiple rows)
        alt:       Use alternating background color
        
    Returns:
        Number of rows consumed
    """
    bg  = ALT_ROW if alt else WHITE
    raw = value if value not in (None, "null", "NULL", "") else "—"
    lines = [l.strip() for l in str(raw).replace("\\n", "\n").split("\n") if l.strip()] or ["—"]

    for offset, line in enumerate(lines):
        row = start_row + offset
        # Label cell (column B)
        lc = ws.cell(row=row, column=2, value=label if offset == 0 else "")
        lc.font      = _body_font(bold=(offset == 0))
        lc.fill      = _fill(LABEL_BG)
        lc.alignment = _align("left", "center")
        lc.border    = BORDER

        # Value cell (column C)
        vc = ws.cell(row=row, column=3, value=line)
        vc.font      = _body_font()
        vc.fill      = _fill(bg)
        vc.alignment = _align("left", "center")
        vc.border    = BORDER

    return len(lines)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_employee_report(wb: Workbook, concur_record) -> None:
    """
    Build Employee Report sheet with:
      - Report Information (employee name, ID, dates, status)
      - Financial Summary (amounts, approvals, payments)
      - Reconciliation Overview (transaction/receipt counts)
      - Approval Log (timeline of approvals)
    """
    er      = concur_record.employee_report
    apr_log = concur_record.approval_log

    ws = wb.create_sheet("Employee Report", 0)
    _sheet_banner(ws, "Employee Report")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 44

    r = 3

    # ── Report Information ─────────────────────────────────────────────────────
    _section_header(ws, r, 2, 3, "Report Information")
    r += 1
    for i, (lbl, val) in enumerate([
        ("Employee Name",   er.employee_name),
        ("Employee ID",     er.employee_id),
        ("Report ID",       er.report_id),
        ("Report Date",     er.report_date),
        ("Approval Status", er.approval_status),
        ("Payment Status",  er.payment_status),
        ("Currency",        er.currency),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # ── Financial Summary ──────────────────────────────────────────────────────
    _section_header(ws, r, 2, 3, "Financial Summary")
    r += 1
    for i, (lbl, val) in enumerate([
        ("Report Total",             er.report_total),
        ("Total Amount Claimed",     er.total_amount_claimed),
        ("Amount Approved",          er.amount_approved),
        ("Personal Expenses",        er.personal_expenses),
        ("Amount Due Employee",      er.amount_due_employee),
        ("Amount Due Company Card",  er.amount_due_company_card),
        ("Total Paid By Company",    er.total_paid_by_company),
        ("Amount Due From Employee", er.amount_due_from_employee),
        ("Total Paid By Employee",   er.total_paid_by_employee),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # ── Reconciliation Overview ────────────────────────────────────────────────
    _section_header(ws, r, 2, 3, "Reconciliation Overview")
    r += 1
    for i, (lbl, val) in enumerate([
        ("Total Transactions",  len(concur_record.transactions)),
        ("Total Receipts",      len(concur_record.receipts)),
        ("Matched",             concur_record.matched_count),
        ("Unmatched",           concur_record.unmatched_count),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # ── Approval Log ───────────────────────────────────────────────────────────
    APR_COLS     = ["date", "approver_name", "status", "note"]
    APR_COL_LBLS = ["Date", "Approver Name", "Status", "Note"]
    APR_START    = 2

    _section_header(ws, r, APR_START, APR_START + len(APR_COLS) - 1, "Approval Log")
    r += 1
    _col_headers(ws, r, APR_COL_LBLS, col_start=APR_START)
    r += 1

    for i, entry in enumerate(apr_log):
        _data_row(ws, r, [getattr(entry, k, None) for k in APR_COLS],
                  alt=(i % 2 == 0), col_start=APR_START)
        r += 1

    _autofit(ws)
    ws.column_dimensions["A"].width = 2


def _build_transactions(wb: Workbook, concur_record) -> None:
    """
    Build Transactions sheet with all extracted expense transactions.
    Columns: transaction_id, date, type, amount, vendor, payment method, cost center, etc.
    """
    txns = concur_record.transactions
    ws = wb.create_sheet("Transactions")
    _sheet_banner(ws, f"Transactions  ({len(txns)} records)")
    
    cols = [
        "transaction_id", "transaction_date", "expense_type", "business_purpose",
        "vendor_description", "payment_type", "amount", "cost_center",
        "project", "attendees", "comments"
    ]
    _col_headers(ws, 3, cols)
    
    for i, txn in enumerate(txns, 1):
        _data_row(ws, 3 + i, [getattr(txn, k, None) for k in cols], alt=(i % 2 == 0))
    
    _autofit(ws)
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_receipts(wb: Workbook, concur_record) -> None:
    """
    Build Receipts sheet with all extracted receipts/invoices.
    Columns: receipt_id, order_id, date, vendor, amount, details, detailed summary.
    """
    rcps = concur_record.receipts
    ws = wb.create_sheet("Receipts")
    _sheet_banner(ws, f"Receipts  ({len(rcps)} records)")
    
    cols = ["receipt_id", "order_id", "date", "vendor", "amount", "line_items", "details"]
    _col_headers(ws, 3, cols)
    
    for i, rcp in enumerate(rcps, 1):
        _data_row(ws, 3 + i, [getattr(rcp, k, None) for k in cols],
                  alt=(i % 2 == 0), wrap=True, max_height=80)
    
    _autofit(ws, mx=35)
    ws.column_dimensions[get_column_letter(6)].width = 58
    ws.column_dimensions[get_column_letter(7)].width = 80
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_reconciliation(wb: Workbook, concur_record) -> None:
    """
    Build Reconciliation sheet showing transaction↔receipt matching.
    Color-coded by status (matched=green, unmatched=red) and confidence.
    """
    recon = concur_record.reconciliation
    matched_ct = concur_record.matched_count
    unmatched_ct = concur_record.unmatched_count
    
    ws = wb.create_sheet("Reconciliation")
    _sheet_banner(ws, 
        f"Reconciliation  —  Matched: {matched_ct}   Unmatched: {unmatched_ct}")
    
    cols = ["transaction_id", "receipt_id", "match_status", "confidence", "comment"]
    _col_headers(ws, 3, cols)
    
    # Color mapping: (status, confidence) → background color
    status_map = {
        ("matched", "high"):   MATCHED_BG,
        ("matched", "medium"): WARN_BG,
        ("matched", "low"):    WARN_BG,
    }
    
    for i, rec in enumerate(recon, 1):
        ms = (rec.match_status or "").lower()
        cf = (rec.confidence or "").lower()
        bg = status_map.get((ms, cf), UNMATCHED_BG if ms == "unmatched" else MATCHED_BG)
        
        _data_row(ws, 3 + i, [rec.transaction_id, rec.receipt_id, rec.match_status, rec.confidence, rec.comment],
                  row_fills={1: bg, 2: bg, 3: bg, 4: bg})
    
    _autofit(ws)
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_summary(wb: Workbook, concur_record) -> None:
    """
    Build Summary sheet with:
      - Reconciliation comments and approval comments
      - Key metrics summary
      - Processing metadata
    """
    ws = wb.create_sheet("Summary", len(wb.sheetnames))
    _sheet_banner(ws, "Extraction & Reconciliation Summary")
    
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22
    
    r = 3
    
    # ── Comments ───────────────────────────────────────────────────────────────
    summary = concur_record.report_summary
    _section_header(ws, r, 2, 3, "Comments & Notes")
    r += 1
    
    for i, (lbl, val) in enumerate([
        ("Reconciliation Comment", summary.reconciliation_comment),
        ("Approval Comment",       summary.approval_comment),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))
    
    r += 1
    
    # ── Metrics Summary ────────────────────────────────────────────────────────
    _section_header(ws, r, 2, 3, "Metrics Summary")
    r += 1
    
    for i, (lbl, val) in enumerate([
        ("Total Transactions",      len(concur_record.transactions)),
        ("Total Receipts",          len(concur_record.receipts)),
        ("Matched Entries",         concur_record.matched_count),
        ("Unmatched Entries",       concur_record.unmatched_count),
        ("High-Confidence Matches", sum(1 for e in concur_record.reconciliation 
                                        if e.is_matched and e.is_high_confidence)),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))
    
    r += 1
    
    # ── Processing Info ───────────────────────────────────────────────────────
    _section_header(ws, r, 2, 3, "Processing Info")
    r += 1
    
    for i, (lbl, val) in enumerate([
        ("Extraction Timestamp", datetime.now().isoformat()),
        ("Employee",             concur_record.employee_report.employee_name),
        ("Report ID",            concur_record.employee_report.report_id),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))
    
    _autofit(ws)
    ws.column_dimensions["A"].width = 2


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def write_concur_excel(concur_record, output_base: Path, month_info, pdf_filename: str) -> Path:
    """
    Extract and format ConcurRecord into a multi-sheet Excel workbook.
    
    Creates sheets:
      1. Employee Report    - Summary, financials, reconciliation overview
      2. Transactions       - All extracted expense transactions
      3. Receipts           - All extracted receipts/invoices
      4. Reconciliation     - Transaction↔receipt matching with confidence
      5. Summary            - Comments, metrics, and processing info
    
    Output path: {output_base}/concur/{MM}-{YY}/{pdf_filename}.xlsx
    
    Args:
        concur_record: ConcurRecord object (from extract_concur_record)
        output_base: Base output folder (e.g., "outputs")
        month_info: MonthInfo object with year and month
        pdf_filename: Original PDF filename (without extension)
    
    Returns:
        Path to the created Excel file
        
    Raises:
        ValueError: If concur_record is missing critical fields
    """
    # Build output path: outputs/concur/MM-YY/pdfname.xlsx
    mm = f"{month_info.month:02d}"  # Zero-padded month (03, 12, etc)
    yy = f"{month_info.year % 100:02d}"  # Last 2 digits of year (26 for 2026)
    
    # Remove .pdf extension if present, then add .xlsx
    base_name = pdf_filename.replace(".pdf", "").replace(".PDF", "")
    filename = f"{base_name}.xlsx"
    
    output_folder = output_base / "concur" / f"{mm}-{yy}"
    output_path = output_folder / filename
    
    # Ensure output folder exists
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet
    
    _build_employee_report(wb, concur_record)
    _build_transactions(wb, concur_record)
    _build_receipts(wb, concur_record)
    _build_reconciliation(wb, concur_record)
    _build_summary(wb, concur_record)
    
    # Save
    wb.save(output_path)
    logger.info(
        "concur_excel_saved",
        path=str(output_path),
        employee=concur_record.employee_report.employee_name,
        txn_count=len(concur_record.transactions),
        rcp_count=len(concur_record.receipts),
    )
    
    return output_path
