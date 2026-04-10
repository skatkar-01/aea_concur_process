"""
src/writer.py
──────────────
Converts a validated Statement model into a formatted XLSX workbook.
No I/O decisions here — the caller passes in the output path.
All styling constants are module-level; business logic is in functions.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import Statement, Transaction
from utils.logging_config import get_logger
from utils.metrics import METRICS, timed

logger = get_logger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY        = "1C2B3A"
COL_HDR_BG  = "2C3E50"
WHITE       = "FFFFFF"
ALT_ROW     = "F5F6F7"
TOTAL_BG    = "DDE3EA"
BORDER_CLR  = "D0D5DC"
BODY_COLOR  = "2C2C2C"

# ── Border definitions ────────────────────────────────────────────────────────
_thin  = Side(style="thin",   color=BORDER_CLR)
_med   = Side(style="medium", color="8A9BB0")
BORDER     = Border(left=_thin, right=_thin, top=_thin,  bottom=_thin)
TOT_BORDER = Border(left=_thin, right=_thin, top=_med,   bottom=_med)
SEC_BORDER = Border(left=_thin, right=_thin, top=_med,   bottom=_thin)

# ── Column config ─────────────────────────────────────────────────────────────
COLUMNS = [
    "last_name", "first_name", "card_number", "process_date",
    "merchant_name", "transaction_desc",
    "current_opening", "charges", "credits", "current_closing",
]
AMOUNT_COLS = {7, 8, 9, 10}   # 1-based; right-aligned
N_COLS      = len(COLUMNS)
LAST_COL    = get_column_letter(N_COLS)


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(size: int = 10) -> Font:
    return Font(name="Calibri", bold=True, color=WHITE, size=size)


def _body_font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=BODY_COLOR, size=size)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Amount formatter ──────────────────────────────────────────────────────────

def _fmt(value: Optional[float]) -> Optional[str]:
    if value is None:
        return None
    prefix = "-$" if value < 0 else "$"
    return f"{prefix}{abs(value):,.2f}"


# ── Cell writer ───────────────────────────────────────────────────────────────

def _put(
    ws,
    row: int,
    col: int,
    value: object,
    *,
    bg: str = WHITE,
    bold: bool = False,
    border: Border = BORDER,
    align_h: str = "left",
) -> None:
    cell = ws.cell(row=row, column=col,
                   value=value if value not in (None, "null", "") else "—")
    cell.fill      = _fill(bg)
    cell.font      = _body_font(bold=bold)
    cell.alignment = _align(align_h, "center")
    cell.border    = border


# ── Auto-fit columns ──────────────────────────────────────────────────────────

def _autofit(ws, min_w: int = 10, max_w: int = 55) -> None:
    for col in ws.columns:
        best = min_w
        col_letter = None
        for cell in col:
            try:
                if col_letter is None:
                    col_letter = cell.column_letter
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, max_w))
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = best


# ── Banner rows ───────────────────────────────────────────────────────────────

def _write_banner(ws, statement: Statement, r: int) -> int:
    """Write the top two metadata rows. Returns next available row."""
    n_ch   = statement.total_cardholders
    n_txns = statement.total_transactions

    company    = statement.company_name   or ""
    stmt_type  = statement.statement_type or ""
    period     = statement.period         or ""
    banner_txt = f"{company}  —  {stmt_type}  |  Period: {period}".strip(" —|")

    ws.merge_cells(f"A{r}:{LAST_COL}{r}")
    c = ws.cell(row=r, column=1, value=banner_txt)
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[r].height = 28
    r += 1

    ws.merge_cells(f"A{r}:{LAST_COL}{r}")
    c2 = ws.cell(row=r, column=1,
                 value=f"{n_ch} cardholder(s)   |   {n_txns} transaction(s) total")
    c2.font      = Font(name="Calibri", color="6B7280", size=9)
    c2.fill      = _fill("F9FAFB")
    c2.alignment = _align("left", "center")
    r += 1

    return r


def _write_global_header(ws, r: int) -> int:
    """Write the frozen column-header row. Returns next available row."""
    for col, label in enumerate(COLUMNS, 1):
        c = ws.cell(row=r, column=col, value=label.replace("_", " ").title())
        c.font      = _hdr_font()
        c.fill      = _fill(COL_HDR_BG)
        c.alignment = _align("center" if col in AMOUNT_COLS else "left", "center")
        c.border    = BORDER
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    return r + 1


# ── Per-cardholder section ────────────────────────────────────────────────────

def _write_cardholder_header(ws, ch, r: int) -> int:
    first = (ch.first_name or "").strip()
    last  = (ch.last_name  or "").strip()
    card  = (ch.card_number or "").strip()
    n_txn = len(ch.transactions)
    label = (
        f"  {first} {last}".strip()
        + (f"   |   Card: {card}" if card else "")
        + f"   |   {n_txn} transaction(s)"
    )
    ws.merge_cells(f"A{r}:{LAST_COL}{r}")
    dc = ws.cell(row=r, column=1, value=label)
    dc.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    dc.fill      = _fill(COL_HDR_BG)
    dc.alignment = _align("left", "center")
    dc.border    = SEC_BORDER
    ws.row_dimensions[r].height = 20
    return r + 1


def _write_transaction(ws, txn: Transaction, ch, row: int, idx: int) -> None:
    bg = ALT_ROW if idx % 2 == 0 else WHITE
    vals = [
        txn.last_name    or ch.last_name,
        txn.first_name   or ch.first_name,
        txn.card_number  or ch.card_number,
        txn.process_date,
        txn.merchant_name,
        txn.transaction_desc,
        _fmt(txn.current_opening),
        _fmt(txn.charges),
        _fmt(txn.credits),
        _fmt(txn.current_closing),
    ]
    for col, val in enumerate(vals, 1):
        _put(ws, row, col, val, bg=bg,
             align_h="right" if col in AMOUNT_COLS else "left")


def _write_total_row(ws, ch, r: int) -> int:
    tot = ch.total_row
    tot_vals = [
        (ch.last_name or "") + " Total",
        None, None, None, None, None,
        _fmt(tot.current_opening) if tot else None,
        _fmt(tot.charges)         if tot else None,
        _fmt(tot.credits)         if tot else None,
        _fmt(tot.current_closing) if tot else None,
    ]
    for col, val in enumerate(tot_vals, 1):
        _put(ws, r, col, val,
             bg=TOTAL_BG, bold=True, border=TOT_BORDER,
             align_h="right" if col in AMOUNT_COLS else "left")
    return r + 1


def _write_spacer(ws, r: int) -> int:
    for col in range(1, N_COLS + 1):
        ws.cell(row=r, column=col).fill = _fill(WHITE)
    ws.row_dimensions[r].height = 8
    return r + 1


# ── Public API ────────────────────────────────────────────────────────────────

def write_xlsx(statement: Statement, output_path: Path) -> Path:
    """
    Render a Statement to a formatted XLSX workbook and save it.

    Args:
        statement:   Validated Statement model.
        output_path: Destination file path (created / overwritten).

    Returns:
        The resolved output_path.

    Raises:
        OSError: If the file cannot be written.
    """
    log = logger.bind(output=output_path.name)
    log.info("xlsx_write_start",
             cardholders=statement.total_cardholders,
             transactions=statement.total_transactions)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    with timed(METRICS.xlsx_write_duration):
        r = 1
        r = _write_banner(ws, statement, r)
        r = _write_global_header(ws, r)

        for ch_idx, ch in enumerate(statement.cardholders):
            r = _write_cardholder_header(ws, ch, r)

            for i, txn in enumerate(ch.transactions):
                _write_transaction(ws, txn, ch, r, i)
                r += 1

            r = _write_total_row(ws, ch, r)

            if ch_idx < len(statement.cardholders) - 1:
                r = _write_spacer(ws, r)

        _autofit(ws)
        wb.save(output_path)

    log.info("xlsx_write_complete", path=str(output_path))
    METRICS.files_processed.inc()
    return output_path.resolve()
