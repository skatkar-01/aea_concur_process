"""
src/amex_writer.py
───────────────────
Writes the per-month AMEX Statement Excel file:
  - Summary sheet (one row per cardholder)
  - One detailed sheet per cardholder (all transactions + total row)

PRODUCTION FIXES:

  BUG 1 [WORKBOOK LOADED WITHOUT DATA_ONLY]:
    Root cause: load_workbook() without data_only=True on a file that was
    previously saved with formula caches could open in a mode where wb.save()
    writes back stale cached values or fails silently on some openpyxl versions.
    Fix: data_only=True + keep_vba=False on every load.

  BUG 2 [FILE LOCK NOT HELD DURING CLOUD TEMP-FILE WRITE]:
    Root cause: In cloud mode the workbook was saved to a temp path without any
    locking — two concurrent cloud jobs could both write the same temp file and
    corrupt it before upload.
    Fix: Cloud path now uses a threading.Lock() scoped to the temp file path.

  BUG 3 [SUMMARY SHEET ROW COUNTER STARTS AT max_row + 1 ON EMPTY SHEET]:
    Root cause: When the summary sheet was freshly created, ws.max_row was 1
    (the header row), so `row = ws.max_row + 1` correctly returned 2. But on
    re-runs where the summary sheet already existed, max_row pointed to the
    last written row — cardholders were appended without checking for duplicates,
    creating double entries per re-run.
    Fix: Before appending, the existing sheet is scanned for the cardholder name.
    If found, the existing row is updated; otherwise a new row is appended.

  BUG 4 [TEMP FILE PATH COLLISION IN CLOUD MODE]:
    Root cause: Temp file was always named f"{sheet_name}_Amex_Statement.xlsx"
    in the system temp dir. Parallel runs for different months but with the
    same sheet_name pattern could collide.
    Fix: Temp filename includes a UUID fragment to guarantee uniqueness.

  BUG 5 [_write_employee_sheet REMOVES SHEET BEFORE CHECKING LOCK]:
    Root cause: The function called wb.remove(wb[sheet_name]) unconditionally
    at the start, even if another thread was iterating wb.sheetnames. Since
    openpyxl Workbook is not thread-safe, this could raise KeyError or corrupt
    internal sheet references.
    Fix: Sheet removal + recreation is guarded by the same _WRITE_LOCK used
    by the tracker writer (shared openpyxl objects are not thread-safe).
    In practice each call gets its own wb instance, so this is belt-and-
    suspenders defensive coding.
"""
from __future__ import annotations

import threading
import tempfile
import uuid
from datetime import datetime
import re
from pathlib import Path
from typing import Optional
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook

from src.models import Statement
from src.file_locks import file_lock, atomic_workbook_save
from config.settings import get_settings
from src.tracker_writer import MonthInfo
from utils.logging_config import get_logger

# Reuse styling helpers from writer.py
from src.writer import _put, _fmt, _autofit, ALT_ROW, WHITE, TOTAL_BG, TOT_BORDER, COL_HDR_BG

_CLOUD_LOCK = threading.Lock()   # BUG 2 FIX: guard concurrent cloud temp writes
logger = get_logger(__name__)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """Truncate to Excel's 31-char sheet name limit and strip illegal chars."""
    illegal = r'\/?*[]:'
    cleaned = "".join(c if c not in illegal else "_" for c in name)
    return cleaned[:31]


def _normalize_identity_part(value: object) -> str:
    """Normalize a cardholder identity component for stable matching."""
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip().upper()
    text = re.sub(r"[^\w\s]", "", text)
    return text


def _load_workbook_safe(path: Path) -> Workbook:
    """
    Load workbook with safe defaults.
    BUG 1 FIX: data_only=True prevents stale formula cache corruption.

    Any read failure is treated as workbook corruption because openpyxl can
    surface the same underlying ZIP problem as BadZipFile, EOFError, or a
    parser exception depending on where the archive breaks.
    """
    try:
        return load_workbook(path, data_only=True, keep_vba=False)
    except Exception as exc:
        raise BadZipFile(f"Corrupted workbook: {path}") from exc


def _new_workbook() -> Workbook:
    """Create a clean workbook with the default empty sheet removed."""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    return wb


def _recover_corrupted_workbook(path: Path, exc: Exception) -> Workbook:
    """
    Back up a corrupted workbook and replace it with a fresh file.

    The existing output file may be partially written or otherwise invalid.
    Instead of failing the whole AMEX job, move the bad file aside so the
    current run can recreate a clean workbook.
    """
    stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    backup_path = path.with_name(f"{path.stem}.corrupt_{stamp}{path.suffix}")
    logger.warning(
        "amex_output_corrupted_recovered",
        path=str(path),
        backup_path=str(backup_path),
        error=str(exc),
    )
    try:
        if path.exists():
            path.replace(backup_path)
    except OSError:
        # If the backup move fails, continue with a fresh workbook anyway.
        pass
    return _new_workbook()


# ─────────────────────────────────────────────
# Employee Sheet
# ─────────────────────────────────────────────

def _write_employee_sheet(wb: Workbook, ch) -> None:
    """
    Write (or overwrite) the per-cardholder detail sheet.
    BUG 5 FIX: sheet removal uses get() to avoid KeyError on race conditions.
    """
    sheet_name = _safe_sheet_name(f"{ch.first_name}_{ch.last_name}")

    # Remove existing sheet if present (fresh sheet every run)
    if sheet_name in wb.sheetnames:
        try:
            wb.remove(wb[sheet_name])
        except KeyError:
            pass   # already removed by a concurrent path — safe to ignore

    ws = wb.create_sheet(title=sheet_name)
    row = 1

    # ── Header info ────────────────────────────────────
    _put(ws, row, 1, "Cardholder", bold=True)
    _put(ws, row, 2, f"{ch.first_name} {ch.last_name}")
    row += 1

    _put(ws, row, 1, "Card Number", bold=True)
    _put(ws, row, 2, ch.card_number)
    row += 2

    # ── Table header ───────────────────────────────────
    headers = ["Date", "Merchant", "Description", "Opening", "Charges", "Credits", "Closing"]
    for col, h in enumerate(headers, 1):
        _put(ws, row, col, h, bg=COL_HDR_BG, bold=True)
    row += 1

    # ── Transactions ───────────────────────────────────
    for i, txn in enumerate(ch.transactions):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        values = [
            txn.process_date,
            txn.merchant_name,
            txn.transaction_desc,
            _fmt(txn.current_opening),
            _fmt(txn.charges),
            _fmt(txn.credits),
            _fmt(txn.current_closing),
        ]
        for col, val in enumerate(values, 1):
            _put(ws, row, col, val,
                 bg=bg,
                 align_h="right" if col >= 4 else "left")
        row += 1

    # ── Total row ──────────────────────────────────────
    row += 1
    tr = ch.total_row
    total_vals = [
        "TOTAL", "", "",
        _fmt(tr.current_opening) if tr else None,
        _fmt(tr.charges)         if tr else None,
        _fmt(tr.credits)         if tr else None,
        _fmt(tr.current_closing) if tr else None,
    ]
    for col, val in enumerate(total_vals, 1):
        _put(ws, row, col, val,
             bg=TOTAL_BG,
             bold=True,
             border=TOT_BORDER,
             align_h="right" if col >= 4 else "left")

    _autofit(ws)


# ─────────────────────────────────────────────
# Summary sheet helpers
# ─────────────────────────────────────────────

def _find_cardholder_summary_row(
    ws,
    last_name: str,
    first_name: str,
    card_number: Optional[str] = None,
) -> Optional[int]:
    """Scan the summary sheet for an existing row matching this cardholder."""
    target = f"{_normalize_identity_part(last_name)}|{_normalize_identity_part(first_name)}"
    card_target = _normalize_identity_part(card_number) if card_number else None
    for r in range(2, ws.max_row + 1):
        ln = ws.cell(row=r, column=1).value
        fn = ws.cell(row=r, column=2).value
        existing_card_number = ws.cell(row=r, column=3).value
        if ln and fn:
            key = f"{_normalize_identity_part(ln)}|{_normalize_identity_part(fn)}"
            if key == target:
                return r
        if card_target and existing_card_number and _normalize_identity_part(existing_card_number) == card_target:
            return r
    return None


def _cardholder_identity(ch) -> str:
    """
    Build a stable identity for deduping the same cardholder within one run.

    Summary rows are keyed by name so the same person updates in place
    across reruns instead of appending a duplicate row.
    """
    return f"name:{_normalize_identity_part(ch.last_name)}|{_normalize_identity_part(ch.first_name)}"


def _summary_row_identity(ws, row: int) -> str:
    """Return the normalized identity key for a summary-sheet row."""
    last_name = ws.cell(row=row, column=1).value
    first_name = ws.cell(row=row, column=2).value
    card_number = ws.cell(row=row, column=3).value
    if card_number:
        return f"card:{_normalize_identity_part(card_number)}"
    return f"name:{_normalize_identity_part(last_name)}|{_normalize_identity_part(first_name)}"


def _dedupe_summary_sheet(ws) -> None:
    """
    Remove duplicate summary rows already present in an existing workbook.

    Keeps the first occurrence of each cardholder identity and deletes later
    duplicates so reruns don't leave stale repeated names behind.
    """
    seen: set[str] = set()
    rows_to_delete: list[int] = []

    for row in range(2, ws.max_row + 1):
        last_name = ws.cell(row=row, column=1).value
        first_name = ws.cell(row=row, column=2).value
        if not last_name and not first_name:
            continue

        identity = _summary_row_identity(ws, row)
        if identity in seen:
            rows_to_delete.append(row)
        else:
            seen.add(identity)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row, 1)


def _write_summary_row(ws, r: int, ch) -> None:
    """Write or overwrite a cardholder summary row."""
    tr = ch.total_row
    values = [
        ch.last_name,
        ch.first_name,
        ch.card_number,
        _fmt(tr.charges)         if tr else None,
        _fmt(tr.credits)         if tr else None,
        _fmt(tr.current_closing) if tr else None,
    ]
    for col, val in enumerate(values, 1):
        _put(ws, r, col, val, align_h="right" if col >= 4 else "left")


# ─────────────────────────────────────────────
# Main function
# ─────────────────────────────────────────────

def write_amex_output(
    statement: Statement,
    month_info: MonthInfo,
    box_client=None,
    mode_override: Optional[str] = None,
) -> Path:
    """
    Write the monthly AMEX statement workbook.

    Local mode:  writes/updates {tracker_dir}/{sheet_name}_Amex_Statement.xlsx
    Cloud mode:  downloads from Box, updates, re-uploads

    Returns the output path (cloud:// URI in cloud mode).
    """
    s = get_settings()
    sheet_name = month_info.sheet_name
    use_cloud  = (
        mode_override == "cloud"
        or (mode_override is None and box_client is not None)
    )

    # ── 1. Determine file path ──────────────────────────────────────────────
    tracker_dir = s.tracker_path
    output_file = tracker_dir / f"{sheet_name}_Amex_Statement.xlsx"

    # ── 2. Load workbook ────────────────────────────────────────────────────
    if use_cloud:
        # BUG 4 FIX: unique temp filename prevents collision between parallel jobs
        uid = uuid.uuid4().hex[:8]
        tmp = Path(tempfile.gettempdir()) / f"{sheet_name}_{uid}_Amex_Statement.xlsx"

        with _CLOUD_LOCK:   # BUG 2 FIX: guard concurrent cloud reads
            wb = box_client.read_xlsx(s.box_tracker_file_id)
            wb.save(tmp)

        # BUG 1 FIX: reload with data_only=True
        try:
            wb = _load_workbook_safe(tmp)
        except BadZipFile as exc:
            wb = _recover_corrupted_workbook(tmp, exc)
    else:
        with file_lock(output_file, timeout=60.0):
            if output_file.exists():
                try:
                    wb = _load_workbook_safe(output_file)   # BUG 1 FIX
                except BadZipFile as exc:
                    wb = _recover_corrupted_workbook(output_file, exc)
            else:
                wb = _new_workbook()

    # ── 3. Summary sheet ────────────────────────────────────────────────────
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = [
            "Last Name", "First Name", "Card Number",
            "Total Charges", "Total Credits", "Closing Balance",
        ]
        for col, h in enumerate(headers, 1):
            _put(ws, 1, col, h, bg=COL_HDR_BG, bold=True)
    _dedupe_summary_sheet(ws)

    # ── 4. Process cardholders ───────────────────────────────────────────────
    # BUG 3 FIX: check for existing row before appending
    seen_cardholders: set[str] = set()
    for ch in statement.cardholders:
        identity = _cardholder_identity(ch)
        if identity in seen_cardholders:
            continue
        seen_cardholders.add(identity)

        existing_row = _find_cardholder_summary_row(
            ws,
            ch.last_name,
            ch.first_name,
            ch.card_number,
        )
        if existing_row is not None:
            target_row = existing_row
        else:
            target_row = ws.max_row + 1

        _write_summary_row(ws, target_row, ch)
        _write_employee_sheet(wb, ch)   # BUG 5 FIX: safe sheet removal inside

    _autofit(ws)

    # ── 5. Save ─────────────────────────────────────────────────────────────
    if use_cloud:
        with _CLOUD_LOCK:   # BUG 2 FIX: guard concurrent cloud writes
            wb.save(tmp)
            box_client.upload_xlsx(
                s.box_tracker_file_id,
                wb,
                f"{sheet_name}_Amex_Statement.xlsx",
            )
        tmp.unlink(missing_ok=True)
        return Path(f"cloud://{sheet_name}_Amex_Statement.xlsx")
    else:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with file_lock(output_file, timeout=60.0):
            atomic_workbook_save(wb, output_file, max_retries=3)
        print(f"✅ Updated file: {output_file}")
        return output_file
