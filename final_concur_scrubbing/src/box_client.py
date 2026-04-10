"""
src/box_client.py
──────────────────
Box API wrapper for the cloud architecture.

Handles:
  - JWT authentication (server-side, no user interaction needed)
  - Downloading a PDF by Box file ID → bytes
  - Searching for files in a Box folder by name/path
  - Reading the tracker XLSX from Box → openpyxl Workbook
  - Writing an updated tracker XLSX back to Box

All methods are stateless — a fresh token is fetched (and cached for the
token lifetime) on each BoxClient instantiation.

Required .env / environment variables:
  BOX_CLIENT_ID         OAuth2 client ID
  BOX_CLIENT_SECRET     OAuth2 client secret
  BOX_ENTERPRISE_ID     Enterprise ID (for JWT app auth)
  BOX_JWT_KEY_ID        JWT key ID from the Box app config
  BOX_JWT_PRIVATE_KEY   PEM private key (newlines as \\n in env var)
  BOX_TRACKER_FILE_ID   Box file ID of the tracker XLSX
  BOX_AMEX_FOLDER_ID    Box folder ID for AMEX PDFs
  BOX_CONCUR_FOLDER_ID  Box folder ID for Concur PDFs
"""
from __future__ import annotations

import io
import json
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import requests

from utils.logging_config import get_logger

logger = get_logger(__name__)

_TOKEN_CACHE: dict[str, object] = {}   # simple in-memory cache


@dataclass
class BoxFile:
    file_id:   str
    name:      str
    folder_id: str
    size:      int = 0


class BoxClient:
    """
    Thin wrapper around the Box REST API.
    Uses JWT (server authentication) — no browser redirect needed.
    """

    AUTH_URL   = "https://api.box.com/oauth2/token"
    API_BASE   = "https://api.box.com/2.0"
    UPLOAD_URL = "https://upload.box.com/api/2.0"

    def __init__(
        self,
        client_id: str,
        client_secret: str,
        enterprise_id: str,
        jwt_key_id: str,
        jwt_private_key: str,
    ) -> None:
        self._client_id      = client_id
        self._client_secret  = client_secret
        self._enterprise_id  = enterprise_id
        self._jwt_key_id     = jwt_key_id
        self._jwt_private_key = jwt_private_key
        self._access_token: Optional[str] = None
        self._token_expiry: float = 0.0

    # ── Authentication ────────────────────────────────────────────────────────

    def _get_token(self) -> str:
        """Return a valid Bearer token, refreshing if expired."""
        if self._access_token and time.time() < self._token_expiry - 60:
            return self._access_token

        logger.debug("box_token_refresh")

        # Build the JWT assertion
        try:
            import jwt as pyjwt
        except ImportError:
            raise ImportError(
                "PyJWT is required for Box JWT auth. "
                "Add 'PyJWT>=2.8.0 cryptography>=41.0.0' to requirements.txt"
            )

        now = int(time.time())
        claim = {
            "iss": self._client_id,
            "sub": self._enterprise_id,
            "box_sub_type": "enterprise",
            "aud": self.AUTH_URL,
            "jti": f"box-jwt-{now}",
            "exp": now + 45,
        }
        private_key = self._jwt_private_key.replace("\\n", "\n")
        assertion = pyjwt.encode(
            claim,
            private_key,
            algorithm="RS256",
            headers={"kid": self._jwt_key_id},
        )

        resp = requests.post(
            self.AUTH_URL,
            data={
                "grant_type":            "urn:ietf:params:oauth:grant-type:jwt-bearer",
                "assertion":             assertion,
                "client_id":             self._client_id,
                "client_secret":         self._client_secret,
            },
            timeout=15,
        )
        resp.raise_for_status()
        body = resp.json()

        self._access_token = body["access_token"]
        self._token_expiry = time.time() + body.get("expires_in", 3600)
        logger.debug("box_token_ok", expires_in=body.get("expires_in"))
        return self._access_token

    def _headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self._get_token()}",
            "Content-Type":  "application/json",
        }

    # ── File operations ───────────────────────────────────────────────────────

    def list_folder(self, folder_id: str, limit: int = 500) -> list[BoxFile]:
        """Return all files in a Box folder (non-recursive)."""
        url = f"{self.API_BASE}/folders/{folder_id}/items"
        params = {"limit": limit, "fields": "id,name,type,size,parent"}
        resp = requests.get(url, headers=self._headers(), params=params, timeout=15)
        resp.raise_for_status()
        entries = resp.json().get("entries", [])
        return [
            BoxFile(
                file_id=e["id"],
                name=e["name"],
                folder_id=folder_id,
                size=e.get("size", 0),
            )
            for e in entries
            if e["type"] == "file"
        ]

    def download_file(self, file_id: str) -> bytes:
        """Download a Box file by ID and return its raw bytes."""
        url = f"{self.API_BASE}/files/{file_id}/content"
        resp = requests.get(
            url,
            headers={"Authorization": f"Bearer {self._get_token()}"},
            timeout=60,
            allow_redirects=True,
        )
        resp.raise_for_status()
        logger.debug("box_download_ok", file_id=file_id, bytes=len(resp.content))
        return resp.content

    def download_pdf_to_temp(self, file_id: str, filename: str) -> Path:
        """
        Download a Box PDF to a temp file and return its Path.
        Caller is responsible for deleting the temp file when done.
        """
        import tempfile, os
        data = self.download_file(file_id)
        fd, tmp_path = tempfile.mkstemp(suffix=".pdf", prefix=filename + "_")
        os.close(fd)
        Path(tmp_path).write_bytes(data)
        logger.debug("box_pdf_to_temp", path=tmp_path, bytes=len(data))
        return Path(tmp_path)

    def read_xlsx(self, file_id: str):
        """
        Download a Box XLSX file and return an openpyxl Workbook.
        The workbook is in-memory — call upload_xlsx() to push changes back.
        """
        from openpyxl import load_workbook
        data = self.download_file(file_id)
        wb = load_workbook(io.BytesIO(data))
        logger.debug("box_xlsx_loaded", file_id=file_id, sheets=wb.sheetnames)
        return wb

    def upload_xlsx(self, file_id: str, workbook, filename: str = "tracker.xlsx") -> None:
        """
        Save an openpyxl Workbook and upload it as a new version of an
        existing Box file (file_id).  Does NOT create a new file.
        """
        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)

        url = f"{self.UPLOAD_URL}/files/{file_id}/content"
        resp = requests.post(
            url,
            headers={"Authorization": f"Bearer {self._get_token()}"},
            files={"file": (filename, buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=60,
        )
        resp.raise_for_status()
        logger.info("box_xlsx_uploaded", file_id=file_id, filename=filename)

    def get_file_info(self, file_id: str) -> dict:
        """Return metadata for a Box file."""
        url = f"{self.API_BASE}/files/{file_id}"
        resp = requests.get(url, headers=self._headers(), timeout=15)
        resp.raise_for_status()
        return resp.json()

    # ── Webhook verification ──────────────────────────────────────────────────

    @staticmethod
    def verify_webhook_signature(
        payload: bytes,
        headers: dict[str, str],
        primary_key: str,
        secondary_key: str,
    ) -> bool:
        """
        Verify a Box webhook delivery signature.
        Box signs with HMAC-SHA256 using both primary and secondary keys.

        Args:
            payload:       Raw request body bytes.
            headers:       HTTP headers dict (case-insensitive lookup done here).
            primary_key:   Box webhook primary signature key.
            secondary_key: Box webhook secondary signature key.

        Returns:
            True if the signature is valid, False otherwise.
        """
        import hmac, hashlib, base64

        def _header(name: str) -> str:
            return next((v for k, v in headers.items() if k.lower() == name.lower()), "")

        delivery_ts  = _header("box-delivery-timestamp")
        sig_primary  = _header("box-signature-primary")
        sig_secondary= _header("box-signature-secondary")

        message = payload + delivery_ts.encode()

        def _check(key: str, sig: str) -> bool:
            digest = hmac.new(key.encode(), message, hashlib.sha256).digest()
            expected = base64.b64encode(digest).decode()
            return hmac.compare_digest(expected, sig)

        return _check(primary_key, sig_primary) or _check(secondary_key, sig_secondary)


def box_client_from_settings() -> BoxClient:
    """Construct a BoxClient from environment settings."""
    from config.settings import get_settings
    s = get_settings()
    return BoxClient(
        client_id=s.box_client_id,
        client_secret=s.box_client_secret,
        enterprise_id=s.box_enterprise_id,
        jwt_key_id=s.box_jwt_key_id,
        jwt_private_key=s.box_jwt_private_key,
    )
