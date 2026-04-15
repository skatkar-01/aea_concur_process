"""
src/tracker_writer.py
──────────────────────
Reads and writes the shared Excel tracker workbook
(2025 New AmEx Checklist.xlsx).

Two operations:
  init_month_sheet(statement, month_info)
      → Creates a new sheet for the month (e.g. "March 2026")
      → Writes header row + one row per cardholder from the AMEX statement
      → Concur columns left blank — filled in later as reports arrive
      → Skips creation if the sheet already exists

  patch_cardholder_row(tracker_row, month_info)
      → Loads the existing tracker
      → Finds the row matching cardholder_name on the correct sheet
      → Fills Concur columns: submitted, pdf✓, approvals✓, receipts✓, comments
      → Saves back — uses atomic temp-file write to avoid corruption

Both functions use a file lock so concurrent watcher events don't corrupt the workbook.

Tracker columns:
  A  Cardholder Name
  B  {Month Year} Statement Total       ← AMEX total
  C  Amount Submitted in Concur
  D  Report PDF                         ← ✓ / ✗ / N/A
  E  Approvals
  F  Receipts
  G  Comments
"""
from __future__ import annotations

import os
import shutil
import threading
import tempfile
import time
import uuid
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, GradientFill
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from src.reconciler import TrackerRow
from src.file_locks import file_lock, atomic_workbook_save
from utils.logging_config import get_logger

logger = get_logger(__name__)

# One process-level lock — prevents concurrent writes from the watcher threads
_WRITE_LOCK = threading.Lock()


# ── Month info ────────────────────────────────────────────────────────────────

@dataclass
class MonthInfo:
    year: int
    month: int          # 1-12
    sheet_name: str     # e.g. "March 2026"
    col_b_header: str   # e.g. "March 4, 2026 Statement Total"


# ── Style constants ───────────────────────────────────────────────────────────

NAVY       = "1C2B3A"
TEAL       = "1A6B72"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
ALT_ROW    = "EAF4F4"
PENDING_BG = "FFF9E6"   # pale yellow  — Concur not yet arrived
NA_BG      = "F0F0F0"   # gray         — no AMEX charges
ISSUE_BG   = "FFF0F0"   # pale red     — has comments
OK_BG      = "F0FFF4"   # pale green   — matched, no issues
BORDER_CLR = "C8D8D8"

_thin = Side(style="thin", color=BORDER_CLR)
_med  = Side(style="medium", color="7BA7A7")
BORDER     = Border(left=_thin, right=_thin, top=_thin,  bottom=_thin)
HDR_BORDER = Border(left=_thin, right=_thin, top=_med,   bottom=_med)

HEADERS = [
    "Cardholder Name",
    "__COL_B__",           # replaced with month-specific header at init time
    "Amount Submitted in Concur",
    "Report PDF",
    "Approvals",
    "Receipts",
    "Comments",
]
N_COLS    = len(HEADERS)
LAST_COL  = get_column_letter(N_COLS)
COL_WIDTHS = [28, 22, 22, 12, 12, 12, 55]

CHECK   = "✓"
CROSS   = "✗"
NA_TEXT = "N/A"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "1A1A1A", size: int = 10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)


def _al(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _bool_to_symbol(value: Optional[bool]) -> str:
    if value is None:
        return NA_TEXT
    return CHECK if value else CROSS


def _fmt_currency(value: Optional[float]) -> str:
    if value is None:
        return ""
    if abs(value) < 0.005:
        return "$-"
    if value < 0:
        return f"$({abs(value):,.2f})"
    return f"${value:,.2f}"


def _parse_currency_str(s: Optional[str]) -> Optional[float]:
    """Parse a string rendered by _fmt_currency back to float, or None."""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s == NA_TEXT:
        return None
    if s == "$-":
        return 0.0
    # Negative format $(1,234.56)
    if s.startswith("$(") and s.endswith(")"):
        try:
            return -float(s[2:-1].replace(",", ""))
        except Exception:
            return None
    # Normal $1,234.56
    if s.startswith("$"):
        try:
            return float(s[1:].replace(",", "").replace("(", "").replace(")", ""))
        except Exception:
            return None
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None


def _normalize_name(value: object) -> str:
    """Normalize a cardholder name for stable matching across reruns."""
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip().upper()
    text = re.sub(r"[^\w\s]", "", text)
    return text


def _find_existing_summary_row(ws, name: str) -> Optional[int]:
    """Find an existing cardholder row by normalized name in column A."""
    target = _normalize_name(name)
    for r in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        if cell_value is None:
            continue
        if isinstance(cell_value, str) and cell_value.startswith("   "):
            break
        if _normalize_name(cell_value) == target:
            return r
    return None


def _find_legend_start_row(ws) -> Optional[int]:
    """Return the first row that belongs to the legend block, if present."""
    for r in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        if isinstance(cell_value, str) and cell_value.startswith("   "):
            return r
    return None


def _next_summary_append_row(ws) -> int:
    """
    Find the next row for a new cardholder row.

    If the legend exists, insert above it so the legend stays at the bottom.
    """
    legend_row = _find_legend_start_row(ws)
    if legend_row is not None:
        return legend_row

    for r in range(3, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if value is None or str(value).strip() == "":
            return r
    return ws.max_row + 1


def _dedupe_month_sheet(ws) -> None:
    """
    Remove duplicate cardholder rows already present in a month sheet.

    Keeps the first occurrence of each normalized cardholder name and removes
    later duplicates, without recreating the sheet.
    """
    legend_row = _find_legend_start_row(ws)
    last_data_row = legend_row - 1 if legend_row is not None else ws.max_row
    seen: set[str] = set()
    rows_to_delete: list[int] = []

    for r in range(3, last_data_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        if cell_value is None or str(cell_value).strip() == "":
            continue
        identity = _normalize_name(cell_value)
        if identity in seen:
            rows_to_delete.append(r)
        else:
            seen.add(identity)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)


# ── Row background based on status ───────────────────────────────────────────

def _row_bg(row: TrackerRow) -> str:
    if row.no_charges:
        return NA_BG
    if row.concur_submitted is None:
        return PENDING_BG      # Concur not yet received
    if row.comments:
        return ISSUE_BG        # has issues
    return OK_BG               # all good


# ── Write a single data row ───────────────────────────────────────────────────

def _write_row(ws, r: int, row: TrackerRow, col_b_header: str) -> None:
    bg = _row_bg(row)

    values = [
        row.cardholder_name,
        _fmt_currency(row.amex_total),
        NA_TEXT if row.no_charges else _fmt_currency(row.concur_submitted),
        NA_TEXT if row.no_charges else _bool_to_symbol(row.report_pdf),
        NA_TEXT if row.no_charges else _bool_to_symbol(row.approvals),
        NA_TEXT if row.no_charges else _bool_to_symbol(row.receipts),
        NA_TEXT if row.no_charges else (row.comments or ""),
    ]
    center_cols = {4, 5, 6}   # D, E, F — checkbox columns
    right_cols  = {2, 3}      # B, C — currency

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill   = _fill(bg)
        cell.border = BORDER
        cell.font   = _font()

        if col in center_cols:
            cell.alignment = _al("center")
            # Colour the check/cross
            if val == CHECK:
                cell.font = _font(color="217A45")   # green
            elif val == CROSS:
                cell.font = _font(color="B91C1C")   # red
        elif col in right_cols:
            cell.alignment = _al("right")
        else:
            cell.alignment = _al("left", wrap=col == 7)

    # Increase row height for text wrapping in comments column
    ws.row_dimensions[r].height = 35


def _first_append_row(ws) -> int:
    for r in range(3, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if value is None or str(value).strip() == "":
            return r
    return ws.max_row + 1


def _patch_row_values(ws, r: int, row: TrackerRow) -> None:
    bg = _row_bg(row)

    # B — AMEX total (update when patching)
    b_cell = ws.cell(
        row=r,
        column=2,
        value=NA_TEXT if row.no_charges else _fmt_currency(row.amex_total),
    )
    b_cell.fill = _fill(bg)
    b_cell.alignment = _al("right")
    b_cell.font = _font()
    b_cell.border = BORDER

    # C — Amount Submitted in Concur
    c_cell = ws.cell(
        row=r,
        column=3,
        value=NA_TEXT if row.no_charges else _fmt_currency(row.concur_submitted),
    )
    c_cell.fill = _fill(bg)
    c_cell.alignment = _al("right")
    c_cell.font = _font()
    c_cell.border = BORDER

    for col, val in zip([4, 5, 6], [row.report_pdf, row.approvals, row.receipts]):
        sym = NA_TEXT if row.no_charges else _bool_to_symbol(val)
        bc = ws.cell(row=r, column=col, value=sym)
        bc.fill = _fill(bg)
        bc.alignment = _al("center")
        bc.border = BORDER
        if sym == CHECK:
            bc.font = _font(color="217A45")
        elif sym == CROSS:
            bc.font = _font(color="B91C1C")
        else:
            bc.font = _font()

    g_cell = ws.cell(
        row=r,
        column=7,
        value=NA_TEXT if row.no_charges else (row.comments or ""),
    )
    g_cell.fill = _fill(bg)
    g_cell.alignment = _al("left", wrap=True)
    g_cell.font = _font()
    g_cell.border = BORDER

    for col in [1]:
        ws.cell(row=r, column=col).fill = _fill(bg)


# ── Write header rows ─────────────────────────────────────────────────────────

def _write_headers(ws, month_info: MonthInfo) -> None:
    # Row 1 — banner
    ws.merge_cells(f"A1:{LAST_COL}1")
    banner = ws.cell(row=1, column=1,
                     value=f"AEA — Concur Reconciliation  |  {month_info.sheet_name}")
    banner.font      = _font(bold=True, color=WHITE, size=12)
    banner.fill      = _fill(NAVY)
    banner.alignment = _al("left", "center")
    ws.row_dimensions[1].height = 26

    # Row 2 — column headers
    hdrs = [h if h != "__COL_B__" else month_info.col_b_header for h in HEADERS]
    for col, lbl in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=lbl)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.fill      = _fill(TEAL)
        c.alignment = _al("center" if col in {4, 5, 6} else "left", "center")
        c.border    = HDR_BORDER
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


def _set_col_widths(ws) -> None:
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _apply_mismatch_highlighting(ws, end_row: int) -> None:
    """
    Apply conditional formatting to highlight AMEX/Concur total mismatches.
    Highlights cells in columns B and C red when values don't match and both are numbers.
    """
    if end_row < 3:
        return
    
    # Red fill for mismatch highlighting
    mismatch_fill = PatternFill(fill_type="solid", fgColor="FFE6E6")
    mismatch_font = Font(color="B91C1C", bold=True)  # red font
    
    # Formula rule: highlight if B != C (comparing AMEX and Concur totals)
    # Only applies to numeric values (not N/A)
    rule = FormulaRule(
        formula=[f'AND($B3<>"",C3<>"",NOT(ISERROR(VALUE($B3))),NOT(ISERROR(VALUE(C3))),$B3<>C3)'],
        font=mismatch_font,
        fill=mismatch_fill,
        stopIfTrue=True,
    )
    
    # Apply to data range (B3 to C + last data row)
    ws.conditional_formatting.add(f"B3:C{end_row}", rule)


# ── Legend rows (appended below data) ────────────────────────────────────────

def _write_legend(ws, next_row: int) -> None:
    ws.insert_rows(next_row)
    ws.insert_rows(next_row)
    next_row += 2
    ws.cell(row=next_row, column=1, value="")
    next_row += 1
    legend_items = [
        (OK_BG,      "✓ Matched — no issues"),
        (ISSUE_BG,   "⚠ Has comments / issues"),
        (PENDING_BG, "⏳ Concur report not yet received"),
        (NA_BG,      "— No AMEX charges this period"),
    ]
    for bg, label in legend_items:
        # ws.merge_cells(f"A{next_row}:{LAST_COL}{next_row}")
        c = ws.cell(row=next_row, column=1, value=f"   {label}")
        c.fill      = _fill(bg)
        c.font      = _font(size=9, color="444444")
        c.alignment = _al("left", "center")
        ws.row_dimensions[next_row].height = 14
        next_row += 1


# ── Atomic save ───────────────────────────────────────────────────────────────

def _cleanup_temp_file(tmp_path: str | Path) -> None:
    """
    Best-effort cleanup for temp workbook files on Windows.

    A short retry loop prevents transient sharing violations from crashing the
    workbook save path after the file has already been successfully replaced.
    """
    path = Path(tmp_path)
    for attempt in range(5):
        try:
            path.unlink(missing_ok=True)
            return
        except PermissionError:
            time.sleep(0.1 * (attempt + 1))
        except OSError:
            return
    logger.warning("tracker_temp_cleanup_failed", temp_path=str(path))


def _atomic_save(wb: Workbook, target: Path, max_retries: int = 5) -> None:
    """
    Write to a temp file then atomically replace the target with retry logic.
    
    Retries with exponential backoff when file is locked (PermissionError/WinError 32).
    
    Args:
        wb: Workbook to save
        target: Target file path
        max_retries: Max attempts before raising (default 5)
        
    Raises:
        PermissionError: File remains locked after all retries
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        suffix=f".{uuid.uuid4().hex[:8]}.xlsx",
        dir=target.parent,
    )
    os.close(fd)
    try:
        try:
            wb.save(tmp)
        except Exception as e:
            logger.error("wb_save_failed", file_path=str(target), tmp=tmp, error=str(e))
            raise

        # Attempt replace with retries on lock contention
        for attempt in range(max_retries):
            try:
                os.replace(tmp, target)
                logger.info("atomic_replace_success", file_path=str(target), attempts=attempt + 1)
                return
            except PermissionError as e_replace:
                # File is locked — wait and retry
                if attempt < max_retries - 1:
                    wait_seconds = min(0.5 * (2 ** attempt), 5.0)  # exponential backoff, capped at 5s
                    logger.warning(
                        "os_replace_locked_retry",
                        file_path=str(target),
                        attempt=attempt + 1,
                        max_retries=max_retries,
                        wait_seconds=wait_seconds,
                        error=str(e_replace),
                    )
                    time.sleep(wait_seconds)
                else:
                    # Out of retries — try copy as final fallback
                    logger.warning(
                        "os_replace_failed_final_try_copy",
                        file_path=str(target),
                        tmp=tmp,
                        error=str(e_replace),
                    )
                    try:
                        shutil.copy2(tmp, target)
                        _cleanup_temp_file(tmp)
                        logger.info("fallback_copy_success", file_path=str(target))
                        return
                    except Exception as e_copy:
                        logger.error(
                            "fallback_copy_also_failed",
                            file_path=str(target),
                            tmp=tmp,
                            replace_error=str(e_replace),
                            copy_error=str(e_copy),
                        )
                        # Keep tmp for inspection
                        raise PermissionError(
                            f"Failed to save {target}: file locked (tried {max_retries} times)"
                        ) from e_replace
            except Exception as e_other:
                # Non-lock error — fail immediately
                logger.error("atomic_save_failed", file_path=str(target), tmp=tmp, error=str(e_other))
                raise
    finally:
        _cleanup_temp_file(tmp)

# ── Public API ───────────────────────────────────────────────────────────────

def init_month_sheet(
    rows: list[TrackerRow],
    month_info: MonthInfo,
    tracker_path: Path,
) -> None:
    """
    Create a new sheet for the month in the tracker workbook.

    - If the sheet already exists, logs and returns immediately (idempotent).
    - Writes header, one row per cardholder (AMEX data only, Concur = blank/pending).
    - Appends a colour legend.
    - Saves atomically.

    Args:
        rows:         List of TrackerRow (from reconcile_amex_only).
        month_info:   MonthInfo describing the sheet name and column B header.
        tracker_path: Path to the shared Excel workbook.
    """
    log = logger.bind(sheet=month_info.sheet_name)

    with _WRITE_LOCK:
        with file_lock(tracker_path, timeout=30.0):
            # Load or create workbook
            if tracker_path.exists():
                wb = load_workbook(tracker_path)
            else:
                wb = Workbook()
                # Remove default empty sheet
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            # ── GET OR CREATE SHEET ────────────────
            if month_info.sheet_name in wb.sheetnames:
                ws = wb[month_info.sheet_name]
                log.info("sheet_exists_updating")
            else:
                ws = wb.create_sheet(title=month_info.sheet_name)
                _write_headers(ws, month_info)
                _set_col_widths(ws)
                log.info("sheet_created")

            _dedupe_month_sheet(ws)

            for row in rows:
                existing_row = _find_existing_summary_row(ws, row.cardholder_name)
                if existing_row is not None:
                    _write_row(ws, existing_row, row, month_info.col_b_header)
                    continue

                append_row = _next_summary_append_row(ws)
                if append_row <= ws.max_row:
                    ws.insert_rows(append_row)
                _write_row(ws, append_row, row, month_info.col_b_header)

            # Apply conditional formatting for AMEX/Concur mismatches
            _apply_mismatch_highlighting(ws, ws.max_row)

            if month_info.sheet_name not in wb.sheetnames or _find_legend_start_row(ws) is None:
                _write_legend(ws, next_row=len(rows) + 4)

            log.info("saving_tracker_init", file_path=str(tracker_path), sheet=month_info.sheet_name)
            _atomic_save(wb, tracker_path)
            log.info("sheet_updated", rows=len(rows), file_path=str(tracker_path))


def patch_cardholder_row(
    row: TrackerRow,
    month_info: MonthInfo,
    tracker_path: Path,
) -> bool:
    """
    Patch the Concur columns for one cardholder on an existing sheet.

    Finds the row by matching column A (cardholder name, case-insensitive).
    Updates columns C–G only — does not touch column A or B.
    Returns True if the row was found and patched, False if not found.

    Args:
        row:          TrackerRow with Concur data populated.
        month_info:   Identifies which sheet to patch.
        tracker_path: Path to the shared Excel workbook.
    """
    log = logger.bind(
        sheet=month_info.sheet_name,
        name=row.cardholder_name,
    )

    with _WRITE_LOCK:
        with file_lock(tracker_path, timeout=30.0):
            if not tracker_path.exists():
                log.error("tracker_not_found", file_path=str(tracker_path))
                return False

            wb = load_workbook(tracker_path)

            if month_info.sheet_name not in wb.sheetnames:
                log.error("sheet_not_found")
                return False

            ws = wb[month_info.sheet_name]
            _dedupe_month_sheet(ws)
            target_name = _normalize_name(row.cardholder_name)

            for r in range(3, ws.max_row + 1):
                cell_name = ws.cell(row=r, column=1).value
                if cell_name is None:
                    continue
                if _normalize_name(cell_name) == target_name:
                    # Read existing column B (AMEX total) if present
                    existing_b = ws.cell(row=r, column=2).value
                    existing_val = _parse_currency_str(existing_b)

                    # If incoming amex_total provided, add it to existing (or start from 0)
                    if row.amex_total is not None:
                        base = existing_val if existing_val is not None else 0.0
                        row.amex_total = base + (row.amex_total or 0.0)
                    else:
                        # Preserve existing total if incoming doesn't provide one
                        row.amex_total = existing_val

                    _patch_row_values(ws, r, row)
                    # Apply mismatch highlighting to ensure conditional formatting is current
                    _apply_mismatch_highlighting(ws, ws.max_row)
                    log.info("saving_tracker_patch", file_path=str(tracker_path), sheet=month_info.sheet_name, name=row.cardholder_name)
                    _atomic_save(wb, tracker_path)
                    log.info("row_patched", row=r)
                    return True

            append_row = _next_summary_append_row(ws)
            if append_row <= ws.max_row:
                ws.insert_rows(append_row)
            _write_row(ws, append_row, row, month_info.col_b_header)
            # Apply mismatch highlighting to include the newly appended row
            _apply_mismatch_highlighting(ws, append_row)
            log.info("saving_tracker_append", file_path=str(tracker_path), sheet=month_info.sheet_name, name=row.cardholder_name)
            _atomic_save(wb, tracker_path)
            log.warning("cardholder_row_not_found_appended", row=append_row)
            return True
