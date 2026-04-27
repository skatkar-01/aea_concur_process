"""
tests/test_writer.py
─────────────────────
Integration tests for the XLSX writer.
Verifies that the workbook is generated and has expected content.
"""
from __future__ import annotations

import pytest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from src.models import (
    Cardholder,
    ConcurEmployeeReport,
    ConcurRecord,
    ConcurReceipt,
    Statement,
    Transaction,
)
from src.concur_writer import write_concur_excel
from src.writer import write_xlsx


def _make_statement() -> Statement:
    txns = [
        Transaction(
            last_name="DOE", first_name="JANE",
            card_number="3782-123456-12345",
            process_date="01/15/2026",
            merchant_name="AMAZON",
            transaction_desc="PRIME MEMBERSHIP",
            charges=14.99,
        ),
        Transaction(
            last_name="DOE", first_name="JANE",
            card_number="3782-123456-12345",
            process_date="01/20/2026",
            merchant_name="DELTA AIR LINES",
            transaction_desc="FLIGHT NYC-LAX",
            charges=350.00,
        ),
    ]
    ch = Cardholder(
        last_name="DOE", first_name="JANE",
        card_number="3782-123456-12345",
        transactions=txns,
        total_row=Transaction(
            charges=364.99,
            current_closing=364.99,
            is_total_row=True,
        ),
    )
    return Statement(
        company_name="TEST CORP",
        statement_type="CORPORATE CARD",
        period="JAN_042026",
        cardholders=[ch],
    )


class TestWriteXlsx:
    def test_file_created(self, tmp_path: Path):
        stmt = _make_statement()
        out  = write_xlsx(stmt, tmp_path / "test.xlsx")
        assert out.exists()

    def test_worksheet_name(self, tmp_path: Path):
        stmt = _make_statement()
        out  = write_xlsx(stmt, tmp_path / "test.xlsx")
        wb   = load_workbook(out)
        assert "Statement" in wb.sheetnames

    def test_row_count_reasonable(self, tmp_path: Path):
        """
        Row count = banner(2) + header(1) + ch_header(1) + txns(2) + total(1)
        = 7 rows minimum.
        """
        stmt = _make_statement()
        out  = write_xlsx(stmt, tmp_path / "test.xlsx")
        wb   = load_workbook(out)
        ws   = wb.active
        assert ws.max_row >= 7

    def test_banner_content(self, tmp_path: Path):
        stmt = _make_statement()
        out  = write_xlsx(stmt, tmp_path / "test.xlsx")
        wb   = load_workbook(out)
        ws   = wb.active
        # Row 1 should contain company name somewhere
        banner_cell = ws.cell(row=1, column=1).value or ""
        assert "TEST CORP" in banner_cell

    def test_output_dir_created(self, tmp_path: Path):
        stmt   = _make_statement()
        nested = tmp_path / "deep" / "nested" / "dir"
        out    = write_xlsx(stmt, nested / "test.xlsx")
        assert out.exists()


class TestConcurWriter:
    def test_receipt_details_written_to_excel(self, tmp_path: Path):
        record = ConcurRecord(
            employee_report=ConcurEmployeeReport(employee_name="Jane Doe", report_id="R-1"),
            receipts=[
                ConcurReceipt(
                    receipt_id="rcp_1",
                    order_id="ORD-1",
                    date="01/15/2026",
                    vendor="Demo Vendor",
                    amount=12.34,
                    line_items=[{"name": "Item", "amount": "$12.34"}],
                    details="Header line\nItem line",
                    summary="Header line; Item line",
                )
            ],
        )
        month_info = SimpleNamespace(month=1, year=2026)
        out = write_concur_excel(record, tmp_path, month_info, "demo.pdf")

        wb = load_workbook(out)
        ws = wb["Receipts"]

        assert ws.cell(row=3, column=6).value == "Line Items"
        assert ws.cell(row=4, column=6).value == '[{"name": "Item", "amount": "$12.34"}]'
        assert ws.cell(row=3, column=7).value == "Details"
        assert ws.cell(row=4, column=7).value == "Header line\nItem line"
