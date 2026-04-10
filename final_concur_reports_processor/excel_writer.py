"""
excel_writer.py
Builds a formatted multi-sheet Excel report from extracted expense data.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# Theme constants
# ─────────────────────────────────────────────────────────────

NAVY         = "1C2B3A"
COL_HDR_BG   = "2C3E50"
WHITE        = "FFFFFF"
ALT_ROW      = "F5F6F7"
LABEL_BG     = "ECEEF0"
BORDER_CLR   = "D0D5DC"
BODY_COLOR   = "2C2C2C"
MATCHED_BG   = "EBF5EB"
UNMATCHED_BG = "FDECEA"
WARN_BG      = "FDFAE8"

_bdr    = Side(style="thin", color=BORDER_CLR)
BORDER  = Border(left=_bdr, right=_bdr, top=_bdr, bottom=_bdr)

STATUS_MAP = {
    ("matched", "high"):   MATCHED_BG,
    ("matched", "medium"): WARN_BG,
    ("matched", "low"):    WARN_BG,
}

# ─────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(size: int = 10) -> Font:
    return Font(name="Calibri", bold=True, color=WHITE, size=size)


def _body_font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=BODY_COLOR, size=size)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _autofit(ws, mn: int = 10, mx: int = 50) -> None:
    """Fit column widths to content, skipping merged cells and the banner row."""
    for col in ws.columns:
        best = mn
        col_letter = None
        for cell in col:
            try:
                ltr = cell.column_letter
                if col_letter is None:
                    col_letter = ltr
                if cell.row != 1 and cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, mx))
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = best


# ─────────────────────────────────────────────────────────────
# Layout helpers
# ─────────────────────────────────────────────────────────────

def _sheet_banner(ws, title: str) -> None:
    last = get_column_letter(20)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 28


def _section_header(ws, row: int, col_start: int, col_end: int, label: str) -> None:
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.font      = _hdr_font(10)
    c.fill      = _fill(COL_HDR_BG)
    c.alignment = _align("left", "center")
    c.border    = BORDER


def _col_headers(ws, row: int, labels: list[str], col_start: int = 1) -> None:
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i,
                    value=lbl.replace("_", " ").title())
        c.font      = _hdr_font()
        c.fill      = _fill(COL_HDR_BG)
        c.alignment = _align("center", "center")
        c.border    = BORDER


def _data_row(ws, row: int, values: list, alt: bool = False,
              row_fills: dict | None = None, wrap: bool = False,
              max_height: int | None = None, col_start: int = 1) -> None:
    bg = ALT_ROW if alt else WHITE
    for i, val in enumerate(values):
        col = col_start + i
        v   = val if val not in (None, "null", "NULL", "") else "—"
        c   = ws.cell(row=row, column=col, value=v)
        c.fill      = _fill(row_fills.get(col, bg) if row_fills else bg)
        c.font      = _body_font()
        c.alignment = _align("left", "top" if wrap else "center", wrap=wrap)
        c.border    = BORDER
    if max_height:
        ws.row_dimensions[row].height = max_height


def _kv(ws, start_row: int, label: str, value, alt: bool = False) -> int:
    """
    Write a key-value pair. Multi-line values each get their own row.
    Returns number of rows consumed.
    """
    bg  = ALT_ROW if alt else WHITE
    raw = value if value not in (None, "null", "NULL", "") else "—"
    lines = [l.strip() for l in str(raw).replace("\\n", "\n").split("\n") if l.strip()] or ["—"]

    for offset, line in enumerate(lines):
        row = start_row + offset
        lc = ws.cell(row=row, column=2, value=label if offset == 0 else "")
        lc.font      = _body_font(bold=(offset == 0))
        lc.fill      = _fill(LABEL_BG)
        lc.alignment = _align("left", "center")
        lc.border    = BORDER

        vc = ws.cell(row=row, column=3, value=line)
        vc.font      = _body_font()
        vc.fill      = _fill(bg)
        vc.alignment = _align("left", "center")
        vc.border    = BORDER

    return len(lines)


# ─────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────

def _build_employee_report(wb: Workbook, result: dict,
                           txns: list, rcps: list, recon: list,
                           matched_ct: int, unmatched_ct: int, high_ct: int) -> None:
    er      = result.get("employee_report", {})
    apr_log = result.get("approval_log", [])

    ws = wb.create_sheet("Employee Report", 0)
    _sheet_banner(ws, "Employee Report")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 44

    r = 3

    # Report Information
    _section_header(ws, r, 2, 3, "Report Information"); r += 1
    for i, (lbl, val) in enumerate([
        ("Employee Name",   er.get("employee_name")),
        ("Employee ID",     er.get("employee_id")),
        ("Report ID",       er.get("report_id")),
        ("Report Date",     er.get("report_date")),
        ("Approval Status", er.get("approval_status")),
        ("Payment Status",  er.get("payment_status")),
        ("Currency",        er.get("currency")),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # Financial Summary
    _section_header(ws, r, 2, 3, "Financial Summary"); r += 1
    for i, (lbl, val) in enumerate([
        ("Report Total",             er.get("report_total")),
        ("Total Amount Claimed",     er.get("total_amount_claimed")),
        ("Amount Approved",          er.get("amount_approved")),
        ("Personal Expenses",        er.get("personal_expenses")),
        ("Amount Due Employee",      er.get("amount_due_employee")),
        ("Amount Due Company Card",  er.get("amount_due_company_card")),
        ("Total Paid By Company",    er.get("total_paid_by_company")),
        ("Amount Due From Employee", er.get("amount_due_from_employee")),
        ("Total Paid By Employee",   er.get("total_paid_by_employee")),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # Reconciliation Overview
    _section_header(ws, r, 2, 3, "Reconciliation Overview"); r += 1
    for i, (lbl, val) in enumerate([
        ("Total Transactions",      len(txns)),
        ("Total Receipts",          len(rcps)),
        ("Matched",                 matched_ct),
        ("Unmatched",               unmatched_ct),
        ("High-Confidence Matches", high_ct),
    ]):
        r += _kv(ws, r, lbl, val, alt=(i % 2 == 0))

    r += 1

    # Approval Log
    APR_COLS     = ["date", "approver_name", "status", "note"]
    APR_COL_LBLS = ["Date", "Approver Name", "Status", "Note"]
    APR_START    = 2

    _section_header(ws, r, APR_START, APR_START + len(APR_COLS) - 1, "Approval Log")
    r += 1
    _col_headers(ws, r, APR_COL_LBLS, col_start=APR_START)
    r += 1

    for i, entry in enumerate(apr_log):
        _data_row(ws, r, [entry.get(k) for k in APR_COLS],
                  alt=(i % 2 == 0), col_start=APR_START)
        r += 1

    _autofit(ws)
    ws.column_dimensions["A"].width = 2


def _build_transactions(wb: Workbook, txns: list) -> None:
    ws = wb.create_sheet("Transactions")
    _sheet_banner(ws, f"Transactions  ({len(txns)} records)")
    cols = ["transaction_id", "transaction_date", "expense_type", "business_purpose",
            "vendor_description", "payment_type", "amount", "cost_center",
            "project", "attendees", "comments"]
    _col_headers(ws, 3, cols)
    for i, txn in enumerate(txns, 1):
        _data_row(ws, 3 + i, [txn.get(k) for k in cols], alt=(i % 2 == 0))
    _autofit(ws)
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_receipts(wb: Workbook, rcps: list) -> None:
    ws = wb.create_sheet("Receipts")
    _sheet_banner(ws, f"Receipts  ({len(rcps)} records)")
    cols = ["receipt_id", "order_id", "date", "vendor", "amount", "summary"]
    _col_headers(ws, 3, cols)
    for i, rcp in enumerate(rcps, 1):
        _data_row(ws, 3 + i, [rcp.get(k) for k in cols],
                  alt=(i % 2 == 0), wrap=True, max_height=80)
    _autofit(ws, mx=35)
    ws.column_dimensions[get_column_letter(6)].width = 58
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_reconciliation(wb: Workbook, recon: list,
                           matched_ct: int, unmatched_ct: int) -> None:
    ws = wb.create_sheet("Reconciliation")
    _sheet_banner(ws, f"Reconciliation  —  Matched: {matched_ct}   Unmatched: {unmatched_ct}")
    cols = ["transaction_id", "receipt_id", "match_status", "confidence"]
    _col_headers(ws, 3, cols)
    for i, rec in enumerate(recon, 1):
        ms = (rec.get("match_status") or "").lower()
        cf = (rec.get("confidence")   or "").lower()
        bg = STATUS_MAP.get((ms, cf), UNMATCHED_BG if ms == "unmatched" else MATCHED_BG)
        _data_row(ws, 3 + i, [rec.get(k) for k in cols],
                  row_fills={1: bg, 2: bg, 3: bg, 4: bg})
    _autofit(ws)
    ws.freeze_panes = ws.cell(row=4, column=1)


# ─────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────

def write_excel(result: dict, output_path: Path) -> None:
    """
    Build and save the reconciliation Excel workbook.

    Args:
        result:      Parsed LLM extraction dict.
        output_path: Destination .xlsx path.
    """
    txns  = result.get("transactions",   [])
    rcps  = result.get("receipts",       [])
    recon = result.get("reconciliation", [])

    matched_ct   = sum(1 for x in recon if (x.get("match_status") or "").lower() == "matched")
    unmatched_ct = len(recon) - matched_ct
    high_ct      = sum(1 for x in recon if (x.get("confidence")   or "").lower() == "high")

    wb = Workbook()
    wb.remove(wb.active)

    _build_employee_report(wb, result, txns, rcps, recon, matched_ct, unmatched_ct, high_ct)
    _build_transactions(wb, txns)
    _build_receipts(wb, rcps)
    _build_reconciliation(wb, recon, matched_ct, unmatched_ct)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel saved → '%s'", output_path)
