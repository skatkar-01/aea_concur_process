import base64
import json
import os
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell, MergedCell

client = OpenAI(
      
)

pdf_path = "../inputs/amex/BAKER_C_Feb_042026.pdf"  # Update path as needed

XLSX_PATH = "amex_output.xlsx"
# ──────────────────────────────────────────────
# 1.  Call GPT-4o
# ──────────────────────────────────────────────
with open(pdf_path, "rb") as f:
    base64_string = base64.b64encode(f.read()).decode("utf-8")

SYSTEM_PROMPT = """
Extract ALL data from this AMEX statement.
Use the IMAGE for layout/column positions.
Use the RAW TEXT below for character-accurate values.

Watch for fused first_name+card_number and fused date+merchant.
Multi-line descriptions: merge all parts into transaction_desc.
Blank amount cells → null (not "0.00").

Extract ALL data from this AMEX statement image.
Columns L→R: Last Name | First Name | Card No | Process Date |
Merchant Name | Transaction Description | Current Opening | Charges | Credits | Closing

Watch for fused first_name+card_number and fused date+merchant.
Blank cells → null. Negative = (x.xx) → -x.xx.
"""

EXTRACTION_PROMPT  = """\
You are a precise financial document parser for American Express corporate statements.

DOCUMENT STRUCTURE:
Header: company name, statement type, period (e.g. JAN_042026)
Columns (L→R): Last Name | First Name | Card No | Process Date |
               Merchant Name | Transaction Description |
               Current Opening | Current Period Charges |
               Current Period Credits | Current Closing

KNOWN LAYOUT EXCEPTIONS:
1. First Name + Card Number fused: split at card pattern ####-######-#####
2. Date + Merchant fused: split at date boundary (M/D/YYYY or MM/DD/YYYY)
3. Multi-line descriptions: merge continuation lines into transaction_desc
4. Sparse amounts: blank columns = null (never "0.00" for blank)
5. Negative amounts: "(90.73)" → "-90.73"

Return ONLY valid JSON — no markdown, no explanation:
{
  "company_name": "",
  "statement_type": "",
  "period": "",
  "cardholders": [
    {
      "last_name": "",
      "first_name": "",
      "card_number": "",
      "transactions": [
        {
          "last_name": "", "first_name": "", "card_number": "",
          "process_date": "MM/DD/YYYY or null",
          "merchant_name": null,
          "transaction_desc": "",
          "current_opening": null, "charges": null,
          "credits": null, "current_closing": null,
          "is_total_row": false
        }
      ],
      "total_row": {
        "last_name": "", "first_name": "", "card_number": "",
        "process_date": null, "merchant_name": null, "transaction_desc": null,
        "current_opening": null, "charges": null,
        "credits": null, "current_closing": null,
        "is_total_row": true
      }
    }
  ]
}
"""






# print("Calling GPT-5-mini...")
# response = client.responses.create(
#     model="gpt-5-mini",
#     input=[
#         {"role": "system", "content": SYSTEM_PROMPT},
#         {
#             "role": "user",
#             "content": [
#                 {
#                     "type": "input_file",
#                     "filename": "Baker $906.50.pdf",
#                     "file_data": f"data:application/pdf;base64,{base64_string}",
#                 },
#                 {"type": "input_text", "text": EXTRACTION_PROMPT},
#             ],
#         },
#     ],
# )

# raw = response.output_text.strip()
# if raw.startswith("```"):
#     raw = raw.split("```")[1]
#     if raw.startswith("json"):
#         raw = raw[4:]
#     raw = raw.strip()
# if raw.endswith("```"):
#     raw = raw[:-3].strip()

# result = json.loads(raw)
# print("Extraction complete")


## Save json Load Json from cache
CACHE_PATH="amex_extraction_cache.json"
# with open(CACHE_PATH, "w") as f:
#         json.dump(result, f, indent=2)
# print(f"Extraction complete — cached to '{CACHE_PATH}' ✓")

print(f"Cache found — loading from '{CACHE_PATH}'")
print("  (delete this file to force a fresh API call)")
with open(CACHE_PATH) as f:
    result = json.load(f)
print("Cache loaded ✓")

# ─── Step 2: Style ───────────────────────────────────────────
NAVY        = "1C2B3A"
COL_HDR_BG  = "2C3E50"
WHITE       = "FFFFFF"
ALT_ROW     = "F5F6F7"
LABEL_BG    = "ECEEF0"
TOTAL_BG    = "DDE3EA"
SECTION_BG  = "E8ECF0"    # cardholder section divider
BORDER_CLR  = "D0D5DC"
BODY_COLOR  = "2C2C2C"
 
bdr     = Side(style="thin",   color=BORDER_CLR)
med     = Side(style="medium", color="8A9BB0")
BORDER      = Border(left=bdr, right=bdr, top=bdr,  bottom=bdr)
TOT_BORDER  = Border(left=bdr, right=bdr, top=med,  bottom=med)
SEC_BORDER  = Border(left=bdr, right=bdr, top=med,  bottom=bdr)
 
COLS = ["last_name", "first_name", "card_number", "process_date",
        "merchant_name", "transaction_desc",
        "current_opening", "charges", "credits", "current_closing"]
NUM_COLS = {7, 8, 9, 10}   # 1-based column indices that are amounts
 
def fill(c):                      return PatternFill("solid", fgColor=c)
def hdr_font(sz=10):              return Font(name="Calibri", bold=True, color=WHITE,      size=sz)
def body_font(bold=False, sz=10): return Font(name="Calibri", bold=bold, color=BODY_COLOR, size=sz)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
 
def fmt(val):
    if val is None or val == "":
        return None
    try:
        return f"${float(val):,.2f}"
    except (ValueError, TypeError):
        return str(val)
 
def autofit(ws, mn=10, mx=50):
    for col in ws.columns:
        best, col_letter = mn, None
        for cell in col:
            try:
                ltr = cell.column_letter
                if col_letter is None:
                    col_letter = ltr
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, mx))
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = best
 
def put(ws, row, col, value, bg=WHITE, bold=False, border=BORDER, h="left"):
    c = ws.cell(row=row, column=col,
                value=value if value not in (None, "null", "") else "—")
    c.fill      = fill(bg)
    c.font      = body_font(bold=bold)
    c.alignment = al(h, "center")
    c.border    = border
 
# ─── Step 3: Single sheet ────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Statement"
 
cardholders = result.get("cardholders", [])
company     = result.get("company_name",   "") or ""
period      = result.get("period",         "") or ""
stmt_type   = result.get("statement_type", "") or ""
total_txns  = sum(len(ch.get("transactions", [])) for ch in cardholders)
 
n_cols = len(COLS)
last_col_letter = get_column_letter(n_cols)
 
r = 1
 
# ── Document banner ──────────────────────────────────────────
ws.merge_cells(f"A{r}:{last_col_letter}{r}")
c = ws.cell(row=r, column=1,
            value=f"{company}  —  {stmt_type}  |  Period: {period}".strip(" —|"))
c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
c.fill      = fill(NAVY)
c.alignment = al("left", "center")
ws.row_dimensions[r].height = 28
r += 1
 
ws.merge_cells(f"A{r}:{last_col_letter}{r}")
c2 = ws.cell(row=r, column=1,
             value=f"{len(cardholders)} cardholder(s)   |   {total_txns} transaction(s) total")
c2.font      = Font(name="Calibri", color="6B7280", size=9)
c2.fill      = fill("F9FAFB")
c2.alignment = al("left", "center")
r += 1
 
# ── Global column header ─────────────────────────────────────
for col, lbl in enumerate(COLS, 1):
    c = ws.cell(row=r, column=col, value=lbl.replace("_", " ").title())
    c.font      = hdr_font()
    c.fill      = fill(COL_HDR_BG)
    c.alignment = al("center" if col in NUM_COLS else "left", "center")
    c.border    = BORDER
ws.row_dimensions[r].height = 22
ws.freeze_panes = ws.cell(row=r + 1, column=1)
r += 1
 
# ── One block per cardholder ─────────────────────────────────
for ch_idx, ch in enumerate(cardholders):
    first = (ch.get("first_name") or "").strip()
    last  = (ch.get("last_name")  or "").strip()
    card  = (ch.get("card_number") or "").strip()
    txns  = ch.get("transactions", [])
    tot   = ch.get("total_row") or {}
 
    # Cardholder divider row
    ws.merge_cells(f"A{r}:{last_col_letter}{r}")
    dc = ws.cell(row=r, column=1,
                 value=f"  {first} {last}".strip()
                       + (f"   |   Card: {card}" if card else "")
                       + f"   |   {len(txns)} transaction(s)")
    dc.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    dc.fill      = fill(COL_HDR_BG)
    dc.alignment = al("left", "center")
    dc.border    = SEC_BORDER
    ws.row_dimensions[r].height = 20
    r += 1
 
    # Transaction rows
    for i, txn in enumerate(txns):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        vals = [
            txn.get("last_name")       or ch.get("last_name"),
            txn.get("first_name")      or ch.get("first_name"),
            txn.get("card_number")     or ch.get("card_number"),
            txn.get("process_date"),
            txn.get("merchant_name"),
            txn.get("transaction_desc"),
            fmt(txn.get("current_opening")),
            fmt(txn.get("charges")),
            fmt(txn.get("credits")),
            fmt(txn.get("current_closing")),
        ]
        for col, val in enumerate(vals, 1):
            put(ws, r, col, val, bg=bg,
                h="right" if col in NUM_COLS else "left")
        r += 1
 
    # Total row
    tot_vals = [
        ch.get("last_name") + " Total",
        None, #ch.get("first_name"),
        None, #ch.get("card_number"),
        None,
        None,
        None, # "TOTAL",
        fmt(tot.get("current_opening")),
        fmt(tot.get("charges")),
        fmt(tot.get("credits")),
        fmt(tot.get("current_closing")),
    ]
    for col, val in enumerate(tot_vals, 1):
        put(ws, r, col, val, bg=TOTAL_BG, bold=True,
            border=TOT_BORDER,
            h="right" if col in NUM_COLS else "left")
    r += 1
 
    # Blank spacer between cardholders (skip after last)
    if ch_idx < len(cardholders) - 1:
        for col in range(1, n_cols + 1):
            ws.cell(row=r, column=col).fill = fill(WHITE)
        ws.row_dimensions[r].height = 8
        r += 1
 
# ─── Step 4: Finish ──────────────────────────────────────────
autofit(ws)
wb.save(XLSX_PATH)
print(f"Excel saved -> '{XLSX_PATH}'")
print(f"  Single sheet with {len(cardholders)} cardholder block(s), {total_txns} transactions")