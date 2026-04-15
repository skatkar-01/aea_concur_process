import base64
import json
import os
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell, MergedCell

client = OpenAI(
      )

pdf_path = "Baker $906.50.pdf"  # Update path as needed

XLSX_PATH = "reconciliation_output.xlsx"
# ──────────────────────────────────────────────
# 1.  Call GPT-4o
# ──────────────────────────────────────────────
with open(pdf_path, "rb") as f:
    base64_string = base64.b64encode(f.read()).decode("utf-8")

SYSTEM_PROMPT = """
You are a data extraction assistant.
CRITICAL RULES:
- Return ALL names, emails, IDs, and personal data exactly as they appear in the document.
- Do NOT redact, mask, anonymize, or replace any value with [REDACTED] or similar placeholders.
- Do NOT add disclaimers, warnings, or notes about personal data.
- Return ONLY raw JSON. No markdown, no code fences, no explanation text before or after.
"""

EXTRACTION_PROMPT = """
You are an expert in financial document extraction and reconciliation.
Extract structured data from the document and return FOUR tables in STRICT JSON format.

IMPORTANT: Return ALL names, emails, employee IDs, approver names exactly as written. Never redact anything.

========================
TABLE 1: transactions
========================
Columns (MANDATORY):
- transaction_id (create unique id like txn_1, txn_2)
- transaction_date
- expense_type
- business_purpose
- vendor_description
- payment_type
- amount
- cost_center
- project
- attendees
- comments

========================
TABLE 2: employee_report
========================
Columns (MANDATORY):
- employee_name
- employee_id
- report_id
- report_date
- approval_status
- payment_status
- currency
- report_total
- personal_expenses
- total_amount_claimed
- amount_approved
- amount_due_employee
- amount_due_company_card
- total_paid_by_company
- amount_due_from_employee
- total_paid_by_employee

========================
TABLE 4: approval_log
========================
Columns (MANDATORY):
- date
- approver_name
- status
- note

========================
TABLE 4: receipts
========================
Columns (MANDATORY):
- receipt_id (create unique id like rcp_1, rcp_2)
- order_id (if available, else null)
- date
- vendor
- amount (final amount including tax and tips etc)
- summary (detailed summary of each item in invoice and all extra information)

========================
TABLE 5: reconciliation
========================
Columns (MANDATORY):
- transaction_id
- receipt_id
- match_status (matched / unmatched)
- confidence (high / medium / low)

========================
RULES
========================
- Extract ALL transactions from transaction table section
- Extract employee_report as a SINGLE object
- Extract ALL receipts from receipt text blocks
- Do NOT mix transactions and receipts
- Normalize vendor names (e.g., SWEETGREEN MIDTOWN → Sweetgreen)
- Some fields may be split across multiple lines; reconstruct full values
- If any field is missing → return null
- Strictly add all transactions; do not miss any; avoid duplicates

RECONCILIATION LOGIC:
- Match transactions to receipts using amount (primary), date (exact or near), vendor similarity
- All match → matched (high confidence)
- Partial match → matched (medium/low)
- No match → unmatched

========================
OUTPUT FORMAT (STRICT JSON ONLY)
========================
{
  "transactions": [...],
  "employee_report": {...},
  "approval_log": [...],
  "receipts": [...],
  "reconciliation": [...]
}
"""

# print("Calling GPT-5-mini...")
# response = client.responses.create(
#     model="gpt-5-mini",
#     input=[
#         {"role": "system", "content": SYSTEM_PROMPT},
#         {
#             "role": "user",
#             "content": [
#                 {
#                     "type": "input_file",
#                     "filename": "Baker $906.50.pdf",
#                     "file_data": f"data:application/pdf;base64,{base64_string}",
#                 },
#                 {"type": "input_text", "text": EXTRACTION_PROMPT},
#             ],
#         },
#     ],
# )

# raw = response.output_text.strip()
# if raw.startswith("```"):
#     raw = raw.split("```")[1]
#     if raw.startswith("json"):
#         raw = raw[4:]
#     raw = raw.strip()
# if raw.endswith("```"):
#     raw = raw[:-3].strip()

# result = json.loads(raw)
# print("Extraction complete")


## Save json Load Json from cache
CACHE_PATH="extraction_cache.json"
# with open(CACHE_PATH, "w") as f:
#         json.dump(result, f, indent=2)
# print(f"Extraction complete — cached to '{CACHE_PATH}' ✓")

print(f"Cache found — loading from '{CACHE_PATH}'")
print("  (delete this file to force a fresh API call)")
with open(CACHE_PATH) as f:
    result = json.load(f)
print("Cache loaded ✓")

# ─────────────────────────────────────────────────────────────
# STEP 2 — Style
# ─────────────────────────────────────────────────────────────
NAVY         = "1C2B3A"
COL_HDR_BG   = "2C3E50"
WHITE        = "FFFFFF"
ALT_ROW      = "F5F6F7"
LABEL_BG     = "ECEEF0"
BORDER_CLR   = "D0D5DC"
BODY_COLOR   = "2C2C2C"
MATCHED_BG   = "EBF5EB"
UNMATCHED_BG = "FDECEA"
WARN_BG      = "FDFAE8"
 
bdr = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=bdr, right=bdr, top=bdr, bottom=bdr)
 
def fill(c):
    return PatternFill("solid", fgColor=c)
 
def hdr_font(sz=10):
    return Font(name="Calibri", bold=True, color=WHITE, size=sz)
 
def body_font(bold=False, sz=10):
    return Font(name="Calibri", bold=bold, color=BODY_COLOR, size=sz)
 
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
 
def autofit(ws, mn=10, mx=50):
    """Column autofit — skips merged cells safely via try/except."""
    for col in ws.columns:
        best = mn
        col_letter = None
        for cell in col:
            try:
                ltr = cell.column_letter
                if col_letter is None:
                    col_letter = ltr
                if cell.row != 1 and cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, mx))
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = best
 
def section_header(ws, row, col_start, col_end, label):
    """Dark sub-header spanning cols col_start:col_end."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    c = ws.cell(row=row, column=col_start, value=label)
    c.font      = hdr_font(10)
    c.fill      = fill(COL_HDR_BG)
    c.alignment = al("left", "center")
    c.border    = BORDER
 
def kv(ws, start_row, label, value, alt=False):
    """
    Key-value row(s). If value contains newlines each part gets its own
    Excel row. Label shows only on first row; continuation rows have a
    blank label cell. Returns number of rows consumed.
    """
    bg = ALT_ROW if alt else WHITE
    raw = value if value not in (None, "null", "NULL", "") else "—"
    parts = str(raw).replace("\\n", "\n").split("\n")
    lines = [l.strip() for l in parts if l.strip()]
    if not lines:
        lines = ["—"]
 
    for offset, line in enumerate(lines):
        row = start_row + offset
        lc = ws.cell(row=row, column=2,
                     value=label if offset == 0 else "")
        lc.font      = body_font(bold=(offset == 0))
        lc.fill      = fill(LABEL_BG)
        lc.alignment = al("left", "center")
        lc.border    = BORDER
 
        vc = ws.cell(row=row, column=3, value=line)
        vc.font      = body_font()
        vc.fill      = fill(bg)
        vc.alignment = al("left", "center")
        vc.border    = BORDER
 
    return len(lines)
 
def sheet_banner(ws, title):
    """Single full-width navy title row."""
    last = get_column_letter(20)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.fill      = fill(NAVY)
    c.alignment = al("left", "center")
    ws.row_dimensions[1].height = 28
 
def col_headers(ws, row, labels, col_start=1):
    """Standard column header row starting at col_start."""
    for i, lbl in enumerate(labels):
        col = col_start + i
        c = ws.cell(row=row, column=col,
                    value=lbl.replace("_", " ").title())
        c.font      = hdr_font()
        c.fill      = fill(COL_HDR_BG)
        c.alignment = al("center", "center")
        c.border    = BORDER
 
def data_row(ws, row, values, alt=False, row_fills=None, wrap=False,
             max_height=None, col_start=1):
    """Write a data row starting at col_start; optionally cap row height."""
    bg = ALT_ROW if alt else WHITE
    for i, val in enumerate(values):
        col = col_start + i
        v = val if val not in (None, "null", "NULL", "") else "—"
        c = ws.cell(row=row, column=col, value=v)
        c.fill      = fill(row_fills.get(col, bg) if row_fills else bg)
        c.font      = body_font()
        c.alignment = al("left", "top" if wrap else "center", wrap=wrap)
        c.border    = BORDER
    if max_height:
        ws.row_dimensions[row].height = max_height
 
# ─────────────────────────────────────────────────────────────
# STEP 3 — Data
# ─────────────────────────────────────────────────────────────
txns    = result.get("transactions",   [])
rcps    = result.get("receipts",       [])
recon   = result.get("reconciliation", [])
er      = result.get("employee_report", {})
apr_log = result.get("approval_log", [])
 
matched_ct   = sum(1 for x in recon if (x.get("match_status") or "").lower() == "matched")
unmatched_ct = len(recon) - matched_ct
high_ct      = sum(1 for x in recon if (x.get("confidence")   or "").lower() == "high")
 
wb = Workbook()
wb.remove(wb.active)
 
# ─────────────────────────────────────────────────────────────
# TAB 1 — Employee Report
# ─────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Employee Report", 0)
sheet_banner(ws1, "Employee Report")
 
# Column widths:
#  A = gutter (narrow)
#  B = label  (wide enough for KV labels AND approval log labels)
#  C = value  (for KV rows)
#  D,E = extra columns used by the 4-column approval table
ws1.column_dimensions["A"].width = 2
ws1.column_dimensions["B"].width = 26
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 44
 
r = 3
 
# ── Report Information ──
section_header(ws1, r, 2, 3, "Report Information"); r += 1
report_info = [
    ("Employee Name",   er.get("employee_name")),
    ("Employee ID",     er.get("employee_id")),
    ("Report ID",       er.get("report_id")),
    ("Report Date",     er.get("report_date")),
    ("Approval Status", er.get("approval_status")),
    ("Payment Status",  er.get("payment_status")),
    ("Currency",        er.get("currency")),
]
for i, (lbl, val) in enumerate(report_info):
    r += kv(ws1, r, lbl, val, alt=(i % 2 == 0))
 
r += 1  # spacer
 
# ── Financial Summary ──
section_header(ws1, r, 2, 3, "Financial Summary"); r += 1
fin_info = [
    ("Report Total",            er.get("report_total")),
    ("Total Amount Claimed",    er.get("total_amount_claimed")),
    ("Amount Approved",         er.get("amount_approved")),
    ("Personal Expenses",       er.get("personal_expenses")),
    ("Amount Due Employee",     er.get("amount_due_employee")),
    ("Amount Due Company Card", er.get("amount_due_company_card")),
    ("Total Paid By Company",   er.get("total_paid_by_company")),
    ("Amount Due From Employee",er.get("amount_due_from_employee")),
    ("Total Paid By Employee",  er.get("total_paid_by_employee")),
]
for i, (lbl, val) in enumerate(fin_info):
    r += kv(ws1, r, lbl, val, alt=(i % 2 == 0))
 
r += 1  # spacer
 
# ── Reconciliation Overview ──
section_header(ws1, r, 2, 3, "Reconciliation Overview"); r += 1
recon_info = [
    ("Total Transactions",      len(txns)),
    ("Total Receipts",          len(rcps)),
    ("Matched",                 matched_ct),
    ("Unmatched",               unmatched_ct),
    ("High-Confidence Matches", high_ct),
]
for i, (lbl, val) in enumerate(recon_info):
    r += kv(ws1, r, lbl, val, alt=(i % 2 == 0))
 
r += 1  # spacer
 
# ── Approval Log ──
# FIX 1: section header now spans cols 2–5 to cover all 4 data columns
# FIX 2: no blank row between section header and column headers
# FIX 3: correct snake_case keys matching the LLM JSON output
APR_COLS     = ["date", "approver_name", "status", "note"]
APR_COL_LBLS = ["Date", "Approver Name", "Status", "Note"]
APR_START    = 2   # columns B–E
 
section_header(ws1, r, APR_START, APR_START + len(APR_COLS) - 1, "Approval Log")
r += 1
 
# Column headers — placed in cols B, C, D, E
col_headers(ws1, r, APR_COL_LBLS, col_start=APR_START)
r += 1
 
# Data rows — correct snake_case keys from LLM output
for i, entry in enumerate(apr_log):
    row_vals = [entry.get(k) for k in APR_COLS]
    data_row(ws1, r, row_vals, alt=(i % 2 == 0), col_start=APR_START)
    r += 1
 
autofit(ws1)
ws1.column_dimensions["A"].width = 2

# ─────────────────────────────────────────────────────────────
# TAB 2 — Transactions
# ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Transactions")
sheet_banner(ws2, f"Transactions  ({len(txns)} records)")
cols_t = ["transaction_id", "transaction_date", "expense_type", "business_purpose",
          "vendor_description", "payment_type", "amount", "cost_center",
          "project", "attendees", "comments"]
col_headers(ws2, 3, cols_t)
for i, txn in enumerate(txns, 1):
    data_row(ws2, 3 + i, [txn.get(k) for k in cols_t], alt=(i % 2 == 0))
autofit(ws2)
ws2.freeze_panes = ws2.cell(row=3, column=1)
 
# ─────────────────────────────────────────────────────────────
# TAB 3 — Receipts
# ─────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Receipts")
sheet_banner(ws3, f"Receipts  ({len(rcps)} records)")
cols_r = ["receipt_id", "order_id", "date", "vendor", "amount", "summary"]
col_headers(ws3, 3, cols_r)
for i, rcp in enumerate(rcps, 1):
    data_row(ws3, 3 + i, [rcp.get(k) for k in cols_r],
             alt=(i % 2 == 0), wrap=True, max_height=80)
autofit(ws3, mx=35)
ws3.column_dimensions[get_column_letter(6)].width = 58
ws3.freeze_panes = ws3.cell(row=3, column=1)
 
# ─────────────────────────────────────────────────────────────
# TAB 4 — Reconciliation
# ─────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Reconciliation")
sheet_banner(ws4, f"Reconciliation  —  Matched: {matched_ct}   Unmatched: {unmatched_ct}")
cols_rc = ["transaction_id", "receipt_id", "match_status", "confidence"]
col_headers(ws4, 3, cols_rc)
 
STATUS_MAP = {
    ("matched",   "high"):   MATCHED_BG,
    ("matched",   "medium"): WARN_BG,
    ("matched",   "low"):    WARN_BG,
}
for i, rec in enumerate(recon, 1):
    ms = (rec.get("match_status") or "").lower()
    cf = (rec.get("confidence")   or "").lower()
    bg = STATUS_MAP.get((ms, cf), UNMATCHED_BG if ms == "unmatched" else MATCHED_BG)
    data_row(ws4, 3 + i, [rec.get(k) for k in cols_rc],
             row_fills={1: bg, 2: bg, 3: bg, 4: bg})
autofit(ws4)
ws4.freeze_panes = ws4.cell(row=3, column=1)
 
# ─────────────────────────────────────────────────────────────
# STEP 4 — Save
# ─────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"Excel saved → '{XLSX_PATH}' ✓")