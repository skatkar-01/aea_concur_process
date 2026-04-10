"""
storage/tracker_store.py
Read and write the master tracker Excel/CSV file.
This is the only file in the project that touches the tracker.
No business logic — pure I/O.
"""
from __future__ import annotations
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

from models.tracker import TrackerRecord, TRACKER_COLUMNS
from shared.exceptions import TrackerError
from shared.logger import get_logger

log = get_logger(__name__)


class TrackerStore:
    """
    Manages the master tracker file.
    Supports CSV (always) and Excel (if openpyxl installed).

    One row per cardholder per period.
    Rows are keyed by (employee_name, period) — upsert semantics.
    """

    def __init__(self, tracker_dir: Path):
        self.tracker_dir = Path(tracker_dir)
        self.tracker_dir.mkdir(parents=True, exist_ok=True)
        self._csv_path   = self.tracker_dir / "tracker.csv"
        self._excel_path = self.tracker_dir / "tracker.xlsx"

    # ── Read ──────────────────────────────────────────────────────────────────

    def load_all(self) -> list[dict]:
        """Load all existing tracker rows. Returns [] if tracker doesn't exist yet."""
        if not self._csv_path.exists():
            return []
        try:
            with open(self._csv_path, newline="", encoding="utf-8") as f:
                return list(csv.DictReader(f))
        except Exception as exc:
            raise TrackerError(f"Failed to read tracker: {exc}") from exc

    def find(self, employee_name: str, period: str) -> Optional[dict]:
        """Find an existing tracker row by employee + period."""
        rows = self.load_all()
        target_name   = employee_name.strip().lower()
        target_period = period.strip().lower()
        for row in rows:
            if (
                row.get("employee_name", "").strip().lower() == target_name
                and row.get("period", "").strip().lower() == target_period
            ):
                return row
        return None

    # ── Write ─────────────────────────────────────────────────────────────────

    def upsert(self, record: TrackerRecord) -> None:
        """
        Insert or update a tracker row for this employee+period.
        Preserves all other rows. Writes CSV and optionally Excel.
        """
        new_row = record.to_tracker_row()
        rows    = self.load_all()

        # Find existing row index
        target_name   = record.employee_name.strip().lower()
        target_period = record.period.strip().lower()
        updated = False
        for i, row in enumerate(rows):
            if (
                row.get("employee_name", "").strip().lower() == target_name
                and row.get("period", "").strip().lower() == target_period
            ):
                rows[i] = new_row
                updated = True
                break

        if not updated:
            rows.append(new_row)

        self._write_csv(rows)
        self._write_excel(rows)

        log.info(
            "Tracker %s: %s | %s | status=%s",
            "updated" if updated else "inserted",
            record.employee_name,
            record.period,
            record.status.value,
        )

    def upsert_many(self, records: list[TrackerRecord]) -> None:
        """Batch upsert — more efficient than calling upsert() in a loop."""
        rows = self.load_all()

        # Build lookup: (name_lower, period_lower) → row index
        index_map: dict[tuple, int] = {}
        for i, row in enumerate(rows):
            key = (
                row.get("employee_name", "").strip().lower(),
                row.get("period", "").strip().lower(),
            )
            index_map[key] = i

        for record in records:
            new_row = record.to_tracker_row()
            key     = (
                record.employee_name.strip().lower(),
                record.period.strip().lower(),
            )
            if key in index_map:
                rows[index_map[key]] = new_row
            else:
                rows.append(new_row)
                index_map[key] = len(rows) - 1

        self._write_csv(rows)
        self._write_excel(rows)
        log.info("Tracker batch upsert: %d record(s)", len(records))

    # ── Private writers ───────────────────────────────────────────────────────

    def _write_csv(self, rows: list[dict]) -> None:
        try:
            with open(self._csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=TRACKER_COLUMNS,
                    extrasaction="ignore",
                )
                writer.writeheader()
                writer.writerows(rows)
            log.debug("Tracker CSV written: %s (%d rows)", self._csv_path.name, len(rows))
        except Exception as exc:
            raise TrackerError(f"Failed to write tracker CSV: {exc}") from exc

    def _write_excel(self, rows: list[dict]) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return  # Excel output is optional

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tracker"

            # Header row
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(color="FFFFFF", bold=True)
            for col_idx, col_name in enumerate(TRACKER_COLUMNS, 1):
                cell              = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
                cell.fill         = header_fill
                cell.font         = header_font
                cell.alignment    = Alignment(horizontal="center")

            # Data rows with status-based colour coding
            status_colours = {
                "APPROVED":       "C6EFCE",  # green
                "FLAGGED":        "FFC7CE",  # red
                "PENDING_REVIEW": "FFEB9C",  # yellow
                "ERROR":          "F4CCCC",  # dark red
            }
            for row_idx, row in enumerate(rows, 2):
                status = row.get("status", "").upper()
                fill   = PatternFill("solid", fgColor=status_colours.get(status, "FFFFFF"))
                for col_idx, col_name in enumerate(TRACKER_COLUMNS, 1):
                    cell       = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
                    cell.fill  = fill

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col), default=10
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            wb.save(self._excel_path)
            log.debug("Tracker Excel written: %s", self._excel_path.name)
        except Exception as exc:
            log.warning("Excel write failed (non-fatal): %s", exc)
