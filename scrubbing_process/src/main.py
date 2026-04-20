"""
main.py - Main Entry Point for AmEx Scrubber
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime
from copy import copy
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
from dotenv import load_dotenv

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from scrubber import AmExScrubber


# Load environment variables
load_dotenv()

EXCLUDED_SHEETS = {
    'Employee List',
    'Vendor List',
    'AmEx Load Raw',
    'AmEx All',
    'Summary',
    'Flagged Items',
}

ENTITY_FIELD_MAP = {
    'Employee First Name': 'employee_first_name',
    'Employee Middle Name': 'employee_middle_name',
    'Employee Last Name': 'employee_last_name',
    'Blank/Placeholder': 'blank_placeholder',
    'Report Entry Transaction Date': 'transaction_date',
    'Report Entry Description': 'description',
    'Journal Amount': 'amount',
    'Report Entry Payment Type Name': 'pay_type',
    'Report Entry Expense Type Name': 'expense_code',
    'Report Entry Vendor Name': 'vendor',
    'Project': 'project',
    'Cost Center': 'cost_center',
    'Report Purpose': 'report_purpose',
    'Employee ID': 'employee_id',
}

HIGHLIGHT_FILL = PatternFill(fill_type='solid', fgColor='FFFFC000')

# Processing metadata columns (shown when data is available)
PROCESSING_HEADERS = [
    'Changes',
    'Reasoning',
    'Flags',
    'Confidence',
]

DEBUG_MATCH_HEADERS = [
    'Memory File',
    'Memory Txn ID',
    'Memory Receipt ID',
]

LLM_DEBUG_HEADERS = [
    'LLM Transaction Type',
    'LLM Formatted Description',
    'LLM Description Changed',
    'LLM Expense Code',
    'LLM Expense Code Changed',
    'LLM Confidence',
    'LLM Reasoning',
    'LLM Flags',
    'LLM Is Refund',
    'LLM Error',
]

AMEX_ALL_HEADERS = [
    'Employee First Name',
    'Employee Middle Name',
    'Employee Last Name',
    'Blank/Placeholder',
    'Report Entry Transaction Date',
    'Report Entry Description',
    'Journal Amount',
    'Report Entry Payment Type Name',
    'Report Entry Expense Type Name',
    'Report Entry Vendor Description',
    'Report Entry Vendor Name',
    'Project',
    'Cost Center',
    'Report Purpose',
    'Employee ID',
]


def _is_transaction_sheet(sheet_name: str) -> bool:
    """Return True for actual transaction sheets in the batch workbook."""
    return sheet_name not in EXCLUDED_SHEETS


def _sheet_key(value) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def _compare_values(old_value, new_value) -> bool:
    """Return True when the cell value actually changed."""
    if pd.isna(old_value) and pd.isna(new_value):
        return False
    if old_value is None and new_value is None:
        return False
    if isinstance(old_value, pd.Timestamp):
        old_value = old_value.to_pydatetime()
    if isinstance(new_value, pd.Timestamp):
        new_value = new_value.to_pydatetime()
    if hasattr(old_value, 'strftime'):
        old_value = old_value.strftime('%Y-%m-%d %H:%M:%S')
    if hasattr(new_value, 'strftime'):
        new_value = new_value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(old_value, (int, float)) and isinstance(new_value, (int, float)):
        return abs(float(old_value) - float(new_value)) > 1e-9
    return str(old_value).strip() != str(new_value).strip()


def _cell_value_for_output(value):
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _header_map(ws) -> dict[str, int]:
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if value is not None:
            headers[str(value).strip()] = col_idx
    return headers


def _build_note(result: dict) -> str:
    """Build the short workbook note shown in column O (now unused - use split columns instead)."""
    note = result.get('note') or ''
    if not note:
        return ''
    return str(note).strip()


def _processing_metadata_values(result: dict) -> list:
    """Extract processing metadata for separate columns: Changes, Reasoning, Flags, Confidence."""
    changes = result.get('changes', {})
    
    # Build changes string
    changes_list = []
    if 'description' in changes:
        before, after = changes['description']
        changes_list.append(f"Desc: {before} => {after}")
    if 'expense_code' in changes:
        before, after = changes['expense_code']
        changes_list.append(f"ExpCode: {before} => {after}")
    if 'pay_type' in changes:
        before, after = changes['pay_type']
        changes_list.append(f"PayType: {before} => {after}")
    changes_str = " | ".join(changes_list) if changes_list else ""
    
    # Get reasoning and flags
    reasoning = result.get('reasoning', '')
    flags = result.get('flags', [])
    flags_str = " | ".join(str(f) for f in flags if f) if flags else ""
    confidence = result.get('confidence', 0)
    
    return [
        changes_str,
        reasoning,
        flags_str,
        f"{confidence:.2f}" if confidence else ""
    ]


def _memory_match_values(result: dict) -> list:
    match = result.get('memory_match') or {}
    return [
        match.get('source_file', ''),
        match.get('transaction_id', ''),
        match.get('receipt_id', ''),
    ]


def _llm_debug_values(result: dict) -> list:
    llm = result.get('llm_result') or {}
    flags = llm.get('flags') or []
    if isinstance(flags, list):
        flags_value = ' | '.join(str(flag) for flag in flags if flag is not None)
    else:
        flags_value = str(flags)

    return [
        llm.get('transaction_type', ''),
        llm.get('formatted_description', ''),
        llm.get('description_changed', ''),
        llm.get('expense_code', ''),
        llm.get('expense_code_changed', ''),
        llm.get('confidence', ''),
        llm.get('reasoning', ''),
        flags_value,
        llm.get('is_refund', ''),
        llm.get('error', ''),
    ]


def _copy_cell_style(src, dst) -> None:
    dst._style = copy(src._style)
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


def _result_to_row(result: dict) -> dict:
    """Convert a processed transaction into the AmEx All row format."""
    scrubbed = result['scrubbed']
    row = {
        'Employee First Name': scrubbed.get('employee_first_name', ''),
        'Employee Middle Name': scrubbed.get('employee_middle_name', ''),
        'Employee Last Name': scrubbed.get('employee_last_name', ''),
        'Blank/Placeholder': '',
        'Report Entry Transaction Date': _cell_value_for_output(scrubbed.get('transaction_date', '')),
        'Report Entry Description': scrubbed.get('description', ''),
        'Journal Amount': scrubbed.get('amount', 0),
        'Report Entry Payment Type Name': scrubbed.get('pay_type', ''),
        'Report Entry Expense Type Name': scrubbed.get('expense_code', ''),
        'Report Entry Vendor Description': scrubbed.get('vendor_desc', ''),
        'Report Entry Vendor Name': scrubbed.get('vendor', ''),
        'Project': scrubbed.get('project', ''),
        'Cost Center': scrubbed.get('cost_center', ''),
        'Report Purpose': scrubbed.get('report_purpose', ''),
        'Employee ID': scrubbed.get('employee_id', ''),
    }
    return row


def _delete_sheet_if_exists(wb, sheet_name: str) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]


def _write_dataframe_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    """Write a simple sheet from row dicts."""
    ws.delete_rows(1, ws.max_row)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    for row_idx, record in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(header, ''))


def _write_amex_all_sheet(
    wb,
    all_rows: list[dict],
    results: list[dict] | None = None,
    debug_memory: bool = False,
) -> None:
    """Update the AmEx All sheet while preserving workbook styling."""
    if 'AmEx All' in wb.sheetnames:
        ws = wb['AmEx All']
    else:
        if 'AmEx Load Raw' in wb.sheetnames:
            insert_at = wb.sheetnames.index('AmEx Load Raw') + 1
        else:
            insert_at = len(wb.sheetnames)
        ws = wb.create_sheet('AmEx All', insert_at)

    for col_idx, header in enumerate(AMEX_ALL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx > 1:
            _copy_cell_style(ws.cell(row=1, column=col_idx - 1), cell)
        cell.font = Font(bold=True)

    # Processing metadata columns (Changes, Reasoning, Flags, Confidence)
    processing_start_col = 16
    for offset, header in enumerate(PROCESSING_HEADERS, start=0):
        cell = ws.cell(row=1, column=processing_start_col + offset, value=header)
        cell.font = Font(bold=True)
    
    len_col = processing_start_col + len(PROCESSING_HEADERS)  # Column 20
    len_header_cell = ws.cell(row=1, column=len_col, value='LEN')
    len_header_cell.font = Font(bold=True)

    debug_start_col = None
    if debug_memory:
        debug_start_col = len_col + 1
        for offset, header in enumerate(DEBUG_MATCH_HEADERS, start=0):
            cell = ws.cell(row=1, column=debug_start_col + offset, value=header)
            cell.font = Font(bold=True)
        llm_debug_start_col = debug_start_col + len(DEBUG_MATCH_HEADERS)
        for offset, header in enumerate(LLM_DEBUG_HEADERS, start=0):
            cell = ws.cell(row=1, column=llm_debug_start_col + offset, value=header)
            cell.font = Font(bold=True)
    else:
        llm_debug_start_col = None

    for row_idx, record in enumerate(all_rows, start=2):
        for col_idx, header in enumerate(AMEX_ALL_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ''))
            if col_idx > 1:
                _copy_cell_style(ws.cell(row=row_idx, column=col_idx - 1), cell)
        
        # Populate processing metadata columns (Changes, Reasoning, Flags, Confidence)
        if results and row_idx - 2 < len(results):
            result = results[row_idx - 2]
            processing_vals = _processing_metadata_values(result)
            for offset, value in enumerate(processing_vals, start=0):
                cell = ws.cell(row=row_idx, column=processing_start_col + offset, value=value)
                if value and str(value).strip():
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        ws.cell(row=row_idx, column=len_col, value=f'=LEN(F{row_idx}&K{row_idx})+12')
        if debug_memory and debug_start_col and results and row_idx - 2 < len(results):
            result = results[row_idx - 2]
            # Populate memory match columns
            for offset, value in enumerate(_memory_match_values(result), start=0):
                ws.cell(row=row_idx, column=debug_start_col + offset, value=value)
            # Populate LLM debug columns - ensure data is present
            llm_debug_vals = _llm_debug_values(result)
            for offset, value in enumerate(llm_debug_vals, start=0):
                cell = ws.cell(row=row_idx, column=llm_debug_start_col + offset, value=value)
                # Add borders/formatting to indicate data presence
                if value and str(value).strip():
                    cell.alignment = Alignment(wrap_text=True)

    for row_idx in range(len(all_rows) + 2, ws.max_row + 1):
        limit = (len_col + len(DEBUG_MATCH_HEADERS) + len(LLM_DEBUG_HEADERS)) if debug_memory else len_col
        for col_idx in range(1, limit + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)


def _apply_entity_updates(ws, results: list[dict], debug_memory: bool = False) -> None:
    """Update a transaction sheet in place and highlight changed cells."""
    header_map = _header_map(ws)
    
    # Column positions: 1-15 data, 16-19 processing metadata, 20 LEN formula
    processing_start_col = 16
    len_col = processing_start_col + len(PROCESSING_HEADERS)  # Column 20
    debug_start_col = len_col + 1 if debug_memory else None
    llm_debug_start_col = len_col + 1 + len(DEBUG_MATCH_HEADERS) if debug_memory else None

    # Add headers for processing metadata and len columns with bold formatting
    for offset, header in enumerate(PROCESSING_HEADERS, start=0):
        cell = ws.cell(row=1, column=processing_start_col + offset, value=header)
        cell.font = Font(bold=True)
    
    len_header_cell = ws.cell(row=1, column=len_col, value='LEN')
    len_header_cell.font = Font(bold=True)
    
    if debug_memory:
        for offset, header in enumerate(DEBUG_MATCH_HEADERS, start=0):
            cell = ws.cell(row=1, column=debug_start_col + offset, value=header)
            cell.font = Font(bold=True)
        for offset, header in enumerate(LLM_DEBUG_HEADERS, start=0):
            cell = ws.cell(row=1, column=llm_debug_start_col + offset, value=header)
            cell.font = Font(bold=True)

    for result in results:
        original = result.get('original', {})
        scrubbed = result.get('scrubbed', {})
        row_num = int(original.get('row_index', 0)) + 2

        for header, field_name in ENTITY_FIELD_MAP.items():
            if header not in header_map:
                continue
            if field_name == 'blank_placeholder':
                continue

            if field_name == 'amount':
                new_value = scrubbed.get('amount', original.get('amount', ''))
            else:
                new_value = scrubbed.get(field_name, original.get(field_name, ''))

            cell = ws.cell(row=row_num, column=header_map[header])
            old_value = cell.value
            if _compare_values(old_value, new_value):
                cell.value = _cell_value_for_output(new_value)
                cell.fill = HIGHLIGHT_FILL

        # Populate processing metadata columns
        processing_vals = _processing_metadata_values(result)
        for offset, value in enumerate(processing_vals, start=0):
            cell = ws.cell(row=row_num, column=processing_start_col + offset, value=value)
            if value and str(value).strip():
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        len_value = f'=LEN(F{row_num}&J{row_num})+12'
        ws.cell(row=row_num, column=len_col, value=len_value)
        if debug_memory and debug_start_col:
            for offset, value in enumerate(_memory_match_values(result), start=0):
                ws.cell(row=row_num, column=debug_start_col + offset, value=value)
            for offset, value in enumerate(_llm_debug_values(result), start=0):
                ws.cell(row=row_num, column=llm_debug_start_col + offset, value=value)

def _style_sheet_headers(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)


def _auto_fit_columns(ws, max_width: int = 50) -> None:
    """Auto-fit column widths to content without increasing row height."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                # Calculate max content length
                if cell.value:
                    cell_length = len(str(cell.value))
                    # Account for bold headers (slightly wider)
                    if cell.row == 1:  # Header row
                        cell_length = min(cell_length + 2, max_width)
                    max_length = max(max_length, min(cell_length, max_width))
            except:
                pass
        
        # Set column width with reasonable min/max
        adjusted_width = min(max(max_length + 2, 12), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width



def load_batch_file(filepath: Path) -> dict:
    """
    Load batch Excel file
    
    Expected structure:
    - Employee List sheet
    - Vendor List sheet
    - Transaction sheets (AEA_Posted, SBF_Posted, etc.)
    
    Returns:
        Dict with employee_list, vendor_list, and transactions
    """
    print(f"\nLoading batch file: {filepath.name}")
    
    # Load Excel file
    xl_file = pd.ExcelFile(filepath)
    print(f"  Found sheets: {', '.join(xl_file.sheet_names)}")
    
    # Load Employee List
    employee_list = {}
    if 'Employee List' in xl_file.sheet_names:
        df_emp = pd.read_excel(filepath, sheet_name='Employee List')
        print(f"  Loaded {len(df_emp)} employees")
        employee_list = df_emp.to_dict('records')
    
    # Load Vendor List
    vendor_list = {}
    if 'Vendor List' in xl_file.sheet_names:
        df_vendor = pd.read_excel(filepath, sheet_name='Vendor List')
        print(f"  Loaded {len(df_vendor)} vendors")
        # Create lookup dictionary (uppercase keys)
        for _, row in df_vendor.iterrows():
            if 'VENDOR' in df_vendor.columns and 'CANONICAL_NAME' in df_vendor.columns:
                vendor_list[str(row['VENDOR']).upper()] = row['CANONICAL_NAME']
    
    # Load transactions from entity sheets
    transactions = []
    transaction_sheets = [s for s in xl_file.sheet_names if _is_transaction_sheet(s)]
    
    for sheet_name in transaction_sheets:
        df_txn = pd.read_excel(filepath, sheet_name=sheet_name)
        
        # Convert to transaction dictionaries
        for idx, row in df_txn.iterrows():
            txn = {
                'sheet_source': sheet_name,
                '_source_sheet': sheet_name,
                'row_index': idx,
                'employee_first_name': row.get('Employee First Name', ''),
                'employee_middle_name': row.get('Employee Middle Name', ''),
                'employee_last_name': row.get('Employee Last Name', ''),
                'transaction_date': row.get('Report Entry Transaction Date', ''),
                'description': row.get('Report Entry Description', ''),
                'amount': float(row.get('Journal Amount', 0)),
                'pay_type': row.get('Report Entry Payment Type Name', ''),
                'expense_code': row.get('Report Entry Expense Type Name', ''),
                'vendor_desc': row.get('Report Entry Vendor Description', ''),
                'vendor': row.get('Report Entry Vendor Name', ''),
                'project': row.get('Project', ''),
                'cost_center': row.get('Cost Center', ''),
                'report_purpose': row.get('Report Purpose', ''),
                'employee_id': row.get('Employee ID', ''),
            }
            transactions.append(txn)
        
        print(f"  Loaded {len(df_txn)} transactions from {sheet_name}")
    
    print(f"\nTotal transactions loaded: {len(transactions)}\n")
    
    return {
        'employee_list': employee_list,
        'vendor_list': vendor_list,
        'transactions': transactions
    }


def save_results(
    results: list,
    output_file: Path,
    original_file: Path = None,
    debug_memory: bool = False,
):
    """
    Save scrubbed results to Excel file
    
    Creates tabs:
    - AmEx All: All transactions
    - AEA_Posted: AEA entity
    - SBF_Posted: SBF entity  
    - DEBT_Reviewed: DEBT entity
    - GROWTH_Reviewed: GROWTH entity
    - Summary: Statistics and flagged items
    """
    print(f"\nSaving results to: {output_file.name}")
    
    if original_file and original_file.exists():
        keep_vba = original_file.suffix.lower() == '.xlsm'
        wb = load_workbook(original_file, keep_vba=keep_vba)
    else:
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

    all_rows = [_result_to_row(result) for result in results]

    # Update the master sheet with all scrubbed rows.
    # Validate that results are being passed correctly
    if results:
        sample_result = results[0]
        llm_data_present = 'llm_result' in sample_result and sample_result.get('llm_result')
        if debug_memory and not llm_data_present:
            import sys
            print(f"[WARNING] Debug columns enabled but first result has no llm_result data", file=sys.stderr)
            print(f"  Keys in result: {list(sample_result.keys())}", file=sys.stderr)
            if 'llm_result' in sample_result:
                print(f"  llm_result value: {sample_result['llm_result']}", file=sys.stderr)
    
    _write_amex_all_sheet(wb, all_rows, results=results, debug_memory=debug_memory)

    # Update entity sheets in place and highlight changed values.
    sheet_groups: dict[str, list[dict]] = {}
    for result in results:
        original = result.get('original', {})
        sheet_name = original.get('sheet_source') or original.get('_source_sheet')
        if not sheet_name:
            continue
        sheet_groups.setdefault(sheet_name, []).append(result)

    for sheet_name, sheet_results in sheet_groups.items():
        if sheet_name not in wb.sheetnames:
            continue
        _apply_entity_updates(wb[sheet_name], sheet_results, debug_memory=debug_memory)

    # Ensure all sheet headers are bold
    for sheet_name in wb.sheetnames:
        _style_sheet_headers(wb[sheet_name])
        # Auto-fit columns for readability WITHOUT increasing row height
        _auto_fit_columns(wb[sheet_name])

    _delete_sheet_if_exists(wb, 'Summary')
    _delete_sheet_if_exists(wb, 'Flagged Items')
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"\nResults saved successfully!\n")
    return

    # Rebuild summary and flagged tabs.
    _delete_sheet_if_exists(wb, 'Summary')
    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = 'Metric'
    summary_ws['B1'] = 'Value'
    summary_ws['A1'].font = Font(bold=True)
    summary_ws['B1'].font = Font(bold=True)

    summary_rows = [
        ('Total Transactions', len(results)),
        ('Auto-Approved (≥95%)', sum(1 for r in results if r['confidence'] >= 0.95 and not r.get('flags'))),
        ('Needs Review (80-95%)', sum(1 for r in results if 0.80 <= r['confidence'] < 0.95)),
        ('Flagged (<80%)', sum(1 for r in results if r['confidence'] < 0.80)),
        ('Description Changes', sum(1 for r in results if 'description' in r.get('changes', {}))),
        ('Expense Code Changes', sum(1 for r in results if 'expense_code' in r.get('changes', {}))),
        ('Vendor Changes', sum(1 for r in results if 'vendor' in r.get('changes', {}))),
        ('Processing Time (s)', 0),
    ]
    for idx, (metric, value) in enumerate(summary_rows, start=2):
        summary_ws.cell(row=idx, column=1, value=metric)
        summary_ws.cell(row=idx, column=2, value=value)

    flagged = [r for r in results if r.get('flags')]
    _delete_sheet_if_exists(wb, 'Flagged Items')
    if flagged:
        flagged_ws = wb.create_sheet('Flagged Items')
        headers = ['Employee', 'Date', 'Description', 'Amount', 'Flags', 'Confidence']
        for col_idx, header in enumerate(headers, start=1):
            cell = flagged_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        for row_idx, result in enumerate(flagged, start=2):
            scrubbed = result['scrubbed']
            flagged_ws.cell(row=row_idx, column=1, value=f"{scrubbed.get('employee_first_name', '')} {scrubbed.get('employee_last_name', '')}".strip())
            flagged_ws.cell(row=row_idx, column=2, value=_cell_value_for_output(scrubbed.get('transaction_date', '')))
            flagged_ws.cell(row=row_idx, column=3, value=scrubbed.get('description', ''))
            flagged_ws.cell(row=row_idx, column=4, value=scrubbed.get('amount', 0))
            flagged_ws.cell(row=row_idx, column=5, value=' | '.join(result.get('flags', [])))
            flagged_ws.cell(row=row_idx, column=6, value=result.get('confidence', 0))

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"\nResults saved successfully!\n")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='AmEx Expense Scrubber - Enhanced with LLM'
    )
    parser.add_argument(
        '--input',
        type=Path,
        required=True,
        help='Input batch Excel file'
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=None,
        help='Output Excel file (default: input_scrubbed.xlsx)'
    )
    parser.add_argument(
        '--config',
        type=Path,
        default=Path('src/config'),
        help='Configuration directory (default: ./config)'
    )
    parser.add_argument(
        '--memory-folder',
        type=Path,
        required=True,
        help='Folder with historical transaction Excel files'
    )
    parser.add_argument(
        '--no-cache',
        action='store_true',
        help='Disable LLM result caching'
    )
    parser.add_argument(
        '--no-checkpoint',
        action='store_true',
        help='Disable checkpointing'
    )
    parser.add_argument(
        '--checkpoint-interval',
        type=int,
        default=50,
        help='Save checkpoint every N transactions (default: 50)'
    )
    parser.add_argument(
        '--llm-batch-size',
        type=int,
        default=5,
        help='Number of transactions to send to the LLM in one batch (default: 5)'
    )
    parser.add_argument(
        '--debug-memory',
        action='store_true',
        help='Write matched memory file, transaction ID, and receipt ID into the output workbook'
    )
    
    args = parser.parse_args()
    
    # Validate input file
    if not args.input.exists():
        print(f"Error: Input file not found: {args.input}")
        sys.exit(1)
    
    # Set output file
    if not args.output:
        suffix = args.input.suffix if args.input.suffix.lower() in {'.xlsx', '.xlsm'} else '.xlsx'
        args.output = args.input.parent / f"{args.input.stem}_scrubbed{suffix}"
    
    # Validate config directory
    # if not args.config.exists():
    #     print(f"Error: Config directory not found: {args.config}")
    #     sys.exit(1)

    # Validate memory folder
    if not args.memory_folder.exists():
        print(f"Error: Memory folder not found: {args.memory_folder}")
        sys.exit(1)
    
    # Check Azure OpenAI credentials
    if not os.getenv('AZURE_OPENAI_ENDPOINT') or not os.getenv('AZURE_OPENAI_API_KEY'):
        print("Error: Azure OpenAI credentials not found")
        print("   Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY environment variables")
        sys.exit(1)
    
    try:
        # Load batch file
        batch_data = load_batch_file(args.input)
        
        # Initialize scrubber
        scrubber = AmExScrubber(
            config_dir=args.config,
            memory_folder=args.memory_folder,
            use_cache=not args.no_cache,
            use_checkpoints=not args.no_checkpoint,
            checkpoint_interval=args.checkpoint_interval,
            llm_batch_size=args.llm_batch_size
        )
        
        # Process batch
        batch_id = f"batch_{args.input.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        results = scrubber.process_batch(
            batch_data['transactions'],
            batch_id=batch_id,
            vendor_list=batch_data['vendor_list']
        )
        
        # Print statistics
        scrubber.print_stats()
        
        # Save results
        save_results(
            results,
            output_file=args.output,
            original_file=args.input,
            debug_memory=args.debug_memory,
        )
        
        print("Processing complete!")
        print(f"\nOutput file: {args.output}")
        
    except KeyboardInterrupt:
        print("\n\nProcessing interrupted by user")
        print("Run again with same batch to resume from checkpoint")
        sys.exit(1)
        
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
