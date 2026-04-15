"""
reference_loader.py — AEA Reference Data Loader
================================================
Loads and manages Vendor List and Employee List from a reference workbook.

Usage:
    # Validate reference file
    python reference_loader.py \\
        --reference  path/to/reference.xlsm \\
        --command    validate

    # Prepare CSV for scrubbing (create 3-sheet XLSX)
    python reference_loader.py \\
        --reference  path/to/reference.xlsm \\
        --csv        input.csv \\
        --command    prepare \\
        --output     prepared.xlsx

Commands:
    validate      — Check if reference file has required sheets and columns
    stats         — Print statistics about vendors and employees
    create-template — Generate a blank template reference workbook
    prepare       — Combine CSV + reference into ready-to-scrub XLSX

Dependencies: pip install pandas openpyxl

python reference_loader.py --csv "Batch # 1 - $119,802.46.csv"  --reference "Batch # 1 - $119,802.46 - AEA - Completed.xlsm" --command prepare
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from amex_rules import (
    VENDOR_TITLE_FIXES, VENDOR_ABBREVIATIONS, VENDOR_KEYWORD_MAP,
    ENTITY_DEFAULT,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("reference_loader")


# ══════════════════════════════════════════════════════════════════════════════
#  REFERENCE DATA
# ══════════════════════════════════════════════════════════════════════════════
class ReferenceData:
    """Loads Vendor List and Employee List from a workbook."""

    def __init__(self, wb_path: Optional[str] = None):
        self.vendor_map:      dict[str, str] = {}
        self.employee_entity: dict[str, str] = {}
        if wb_path and Path(wb_path).exists():
            self._load(wb_path)

    def _load(self, path: str):
        log.info(f"Loading reference data: {path}")
        try:
            wb = load_workbook(path, read_only=True, data_only=True)

            if "Vendor List" in wb.sheetnames:
                for row in wb["Vendor List"].iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        raw   = str(row[0]).strip()
                        clean = str(row[1]).strip()
                        self.vendor_map[raw]         = clean
                        self.vendor_map[raw.upper()]  = clean
                n = len([k for k in self.vendor_map if k == k.upper()])
                log.info(f"  Vendor List: {n} entries")

            if "Employee List" in wb.sheetnames:
                for row in wb["Employee List"].iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        self.employee_entity[str(row[1]).strip()] = (
                            str(row[3] or "").strip() if len(row) > 3 else ""
                        )
                log.info(f"  Employee List: {len(self.employee_entity)} entries")

        except Exception as exc:
            log.warning(f"Reference data load failed: {exc}")

    def lookup_vendor(self, vendor_desc: str) -> str:
        """
        Vendor name resolution with 5-step fallback:
        1. Exact match in Vendor List
        2. Uppercase match in Vendor List
        3. Keyword match in VENDOR_KEYWORD_MAP (for chains not in list)
        4. Strip leading numbers and trailing codes, then keyword match
        5. title() with VENDOR_TITLE_FIXES + VENDOR_ABBREVIATIONS
        """
        vd = str(vendor_desc or "").strip()
        if not vd:
            return ""

        # Step 1 & 2: Vendor List lookup (exact + uppercase)
        result = self.vendor_map.get(vd) or self.vendor_map.get(vd.upper())
        if result:
            # Apply any capitalisation fix on top of vendor list result
            return VENDOR_TITLE_FIXES.get(result, result)

        # Step 3: Keyword match on the raw vendor description
        vd_lower = vd.lower()
        for keyword, canonical in VENDOR_KEYWORD_MAP.items():
            if keyword in vd_lower:
                return canonical

        # Step 4: Strip leading digits/codes and retry keyword match
        # e.g. "4254 STARBUCKS" → "STARBUCKS", "JFK2 SHAKE SHACK B37 1051" → ...
        stripped = re.sub(r'^[\d\s]+', '', vd).strip()
        stripped_lower = stripped.lower()
        for keyword, canonical in VENDOR_KEYWORD_MAP.items():
            if keyword in stripped_lower:
                return canonical

        # Step 5: title() with fixes
        titled = vd.title()
        fixed  = VENDOR_TITLE_FIXES.get(titled, titled)
        for long_form, short_form in VENDOR_ABBREVIATIONS.items():
            if fixed.lower() == long_form.lower():
                return short_form

        return fixed

    def lookup_entity(self, employee_id: str) -> str:
        if not employee_id:
            return ENTITY_DEFAULT
        return self.employee_entity.get(str(employee_id).strip(), ENTITY_DEFAULT)


# ══════════════════════════════════════════════════════════════════════════════
#  VALIDATION & DIAGNOSTICS
# ══════════════════════════════════════════════════════════════════════════════
class ReferenceValidator:
    """Validate and inspect reference workbooks."""

    @staticmethod
    def validate(ref_path: str) -> bool:
        """Check if reference file has required structure."""
        p = Path(ref_path)
        if not p.exists():
            log.error(f"File not found: {ref_path}")
            return False

        try:
            wb = load_workbook(ref_path, read_only=True, data_only=True)
        except Exception as e:
            log.error(f"Cannot open workbook: {e}")
            return False

        errors = []

        # Check sheets
        if "Vendor List" not in wb.sheetnames:
            errors.append("Missing sheet: 'Vendor List'")
        if "Employee List" not in wb.sheetnames:
            errors.append("Missing sheet: 'Employee List'")

        # Check Vendor List columns
        if "Vendor List" in wb.sheetnames:
            ws = wb["Vendor List"]
            header = [str(c or "").strip() for c in next(ws.iter_rows(max_row=1, values_only=True), [])]
            if len(header) < 2:
                errors.append("Vendor List: expected at least 2 columns")
            row_count = ws.max_row - 1
            if row_count < 1:
                errors.append("Vendor List: no data rows (header only)")

        # Check Employee List columns
        if "Employee List" in wb.sheetnames:
            ws = wb["Employee List"]
            header = [str(c or "").strip() for c in next(ws.iter_rows(max_row=1, values_only=True), [])]
            if len(header) < 4:
                errors.append("Employee List: expected at least 4 columns")
            row_count = ws.max_row - 1
            if row_count < 1:
                errors.append("Employee List: no data rows (header only)")

        if errors:
            log.error("Validation FAILED:")
            for err in errors:
                log.error(f"  ✗ {err}")
            return False

        log.info("Validation PASSED ✓")
        return True

    @staticmethod
    def print_stats(ref_path: str):
        """Print statistics about reference data."""
        try:
            ref = ReferenceData(ref_path)
        except Exception as e:
            log.error(f"Cannot load reference: {e}")
            return

        print("\n" + "=" * 70)
        print("  REFERENCE DATA STATISTICS")
        print("=" * 70)
        print(f"  File: {ref_path}")
        print()
        print(f"  Vendor List:")
        print(f"    Total entries:    {len(ref.vendor_map)}")
        unique_vendors = len(set(ref.vendor_map.values()))
        print(f"    Unique vendors:   {unique_vendors}")
        print()
        print(f"  Employee List:")
        print(f"    Total employees:  {len(ref.employee_entity)}")
        entities = set(ref.employee_entity.values())
        print(f"    Entities used:    {', '.join(sorted(e for e in entities if e))}")
        print()
        print("=" * 70)
        print()


# ══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE GENERATION
# ══════════════════════════════════════════════════════════════════════════════
class TemplateGenerator:
    """Generate template reference workbooks."""

    @staticmethod
    def create_template(output_path: str):
        """Create a blank template reference workbook."""
        wb = Workbook()
        wb.remove(wb.active)

        # Vendor List sheet
        ws_vendors = wb.create_sheet("Vendor List", 0)
        ws_vendors.append(["Raw Vendor Description", "Clean Vendor Name"])
        ws_vendors.append(["STARBUCKS COFFEE", "Starbucks"])
        ws_vendors.append(["UNITED AIRLINES", "United"])
        ws_vendors.append(["MARRIOTT HOTELS", "Marriott"])

        # Employee List sheet
        ws_employees = wb.create_sheet("Employee List", 1)
        ws_employees.append(["First Name", "Employee ID", "Last Name", "Entity"])
        ws_employees.append(["John", "EMP001", "Doe", "AEA_Posted"])
        ws_employees.append(["Jane", "EMP002", "Smith", "SBF_Posted"])
        ws_employees.append(["Bob", "EMP003", "Johnson", "DEBT_Reviewed"])

        wb.save(output_path)
        log.info(f"Template created: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  PREPARE FOR SCRUBBING
# ══════════════════════════════════════════════════════════════════════════════
class ScrubPreparator:
    """Prepare CSV + reference data into ready-to-scrub XLSX."""

    @staticmethod
    def prepare(csv_path: str, reference_path: str, output_path: str):
        """
        Prepare CSV for scrubbing:
        1. Copy Employee List + Vendor List from reference
        2. Classify CSV rows using Employee List lookup
        3. Populate entity tabs with properly typed data
        4. amex_scrubber.py will apply styling/highlighting
        """
        csv_path = Path(csv_path)
        ref_path = Path(reference_path)
        out_path = Path(output_path)

        if not csv_path.exists():
            log.error(f"CSV file not found: {csv_path}")
            return False

        if not ref_path.exists():
            log.error(f"Reference file not found: {ref_path}")
            return False

        try:
            # 1. Load CSV data
            log.info(f"Reading CSV: {csv_path}")
            df_csv = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
            csv_rows = len(df_csv)
            csv_cols = list(df_csv.columns)
            log.info(f"  Loaded {csv_rows} rows, {len(csv_cols)} columns")

            # Detect column data types
            col_types = {}
            for col in csv_cols:
                if "Date" in col or "date" in col:
                    col_types[col] = "DATE"
                elif col == "Journal Amount":
                    col_types[col] = "NUMERIC"
                else:
                    col_types[col] = "TEXT"
            
            date_format = "mm/dd/yyyy"
            number_format = "0.00"

            # 2. Load reference workbook and build Employee List lookup
            log.info(f"Reading reference: {ref_path}")
            ref_wb = load_workbook(ref_path)
            
            # Build Employee ID → Entity mapping from Employee List
            emp_to_entity = {}
            try:
                emp_list_ws = ref_wb["Employee List"]
                rows = list(emp_list_ws.iter_rows(values_only=True))
                if len(rows) > 1:  # Has header
                    header = rows[0]
                    
                    # Find Employee ID (Paycom EE ID#) and ENTITY columns
                    emp_id_idx = None
                    entity_idx = None
                    
                    # Look for the columns (handle spaces and variations)
                    for idx, col_name in enumerate(header):
                        if col_name and "EE ID" in str(col_name):
                            emp_id_idx = idx
                        if col_name and str(col_name).strip().upper() == "ENTITY":
                            entity_idx = idx
                    
                    if emp_id_idx is not None and entity_idx is not None:
                        for row in rows[1:]:
                            if len(row) > max(emp_id_idx, entity_idx):
                                emp_id = str(row[emp_id_idx]).strip() if row[emp_id_idx] else ""
                                entity = str(row[entity_idx]).strip() if row[entity_idx] else ""
                                if emp_id and entity:
                                    emp_to_entity[emp_id] = entity
                    else:
                        log.warning(f"  Employee List columns - EE ID idx: {emp_id_idx}, ENTITY idx: {entity_idx}")
            except Exception as e:
                log.warning(f"  Error building Employee List mapping: {e}")
            
            log.info(f"  Found {len(emp_to_entity)} employee -> entity mappings")

            # 3. Create output workbook
            out_wb = Workbook()
            out_wb.remove(out_wb.active)

            # Copy Employee List and Vendor List from reference
            for idx, sheet_name in enumerate(["Employee List", "Vendor List"]):
                if sheet_name not in ref_wb.sheetnames:
                    log.warning(f"  Sheet '{sheet_name}' not found in reference")
                    continue
                
                src_ws = ref_wb[sheet_name]
                dst_ws = out_wb.create_sheet(sheet_name, idx)
                
                for row_idx, row in enumerate(src_ws.iter_rows(values_only=True), 1):
                    for col_idx, value in enumerate(row, 1):
                        dst_ws.cell(row_idx, col_idx, value)
                
                row_count = src_ws.max_row - 1 if src_ws.max_row > 1 else 0
                log.info(f"  OK '{sheet_name}': {row_count} rows (from reference)")

            # 4. Classify CSV rows and populate entity tabs
            entity_tabs = ["AEA_Posted", "SBF_Posted", "DEBT_Reviewed"]
            entity_data = {tab: [] for tab in entity_tabs}
            
            # Find Employee ID column in CSV (column name, not index)
            emp_id_col_name = "Employee ID"
            if emp_id_col_name not in csv_cols:
                log.warning(f"  CSV missing '{emp_id_col_name}' column - defaulting all rows to AEA_Posted")
                emp_id_col_name = None
            
            # Classify each row
            for _, row in df_csv.iterrows():
                emp_id = ""
                if emp_id_col_name:
                    emp_id = str(row[emp_id_col_name]).strip() if pd.notna(row[emp_id_col_name]) else ""
                
                # Look up entity from Employee List mapping
                entity = emp_to_entity.get(emp_id)
                
                # Default to AEA_Posted if not found
                if not entity or entity not in entity_tabs:
                    if entity and entity == "AEA":
                        entity = "AEA_Posted"
                    elif entity and entity == "SBF":
                        entity = "SBF_Posted"
                    elif entity and entity == "DEBT":
                        entity = "DEBT_Reviewed"
                    else:
                        entity = "AEA_Posted"
                
                entity_data[entity].append([row[col] for col in csv_cols])
            
            # Write headers and data to entity tabs with proper formatting
            for idx, tab_name in enumerate(entity_tabs, 2):
                ws_entity = out_wb.create_sheet(tab_name, idx + 1)
                
                # Write header row
                ws_entity.append(csv_cols)
                
                # Write data rows with proper type formatting
                for row_num, data_row in enumerate(entity_data[tab_name], start=2):
                    for col_num, (col_name, value) in enumerate(zip(csv_cols, data_row), start=1):
                        cell = ws_entity.cell(row=row_num, column=col_num)
                        
                        # Apply data type conversion and formatting
                        col_type = col_types.get(col_name, "TEXT")
                        
                        if col_type == "DATE" and value and str(value).strip():
                            try:
                                # Parse date from MM/DD/YYYY format
                                date_obj = pd.to_datetime(value, format="%m/%d/%Y")
                                cell.value = date_obj
                                cell.number_format = date_format
                            except:
                                cell.value = value
                        elif col_type == "NUMERIC" and value and str(value).strip():
                            try:
                                cell.value = float(value)
                                cell.number_format = number_format
                            except:
                                cell.value = value
                        else:
                            cell.value = value
                
                row_count = len(entity_data[tab_name])
                log.info(f"  OK '{tab_name}': {row_count} rows (classified from CSV)")

            # 5. Save output file
            out_wb.save(str(out_path))
            log.info(f"OK Saved: {out_path} (5 sheets)")
            log.info("  -> Employee List, Vendor List, AEA_Posted, SBF_Posted, DEBT_Reviewed")
            log.info("  -> Dates formatted as mm/dd/yyyy, Numbers as 0.00")
            log.info("  -> amex_scrubber.py will apply styling/highlighting")
            return True

        except Exception as e:
            log.error(f"Preparation failed: {e}")
            import traceback
            traceback.print_exc()
            return False


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════
def main():
    p = argparse.ArgumentParser(description="AEA Reference Data Loader")
    p.add_argument("--reference", required=True, help="Path to reference workbook")
    p.add_argument(
        "--command",
        default="validate",
        choices=["validate", "stats", "create-template", "prepare"],
        help="Command to run"
    )
    p.add_argument("--csv", help="Input CSV file (required for 'prepare' command)")
    p.add_argument("--output", help="Output XLSX file (for 'prepare' command)")
    args = p.parse_args()

    if args.command == "validate":
        ReferenceValidator.validate(args.reference)

    elif args.command == "stats":
        ReferenceValidator.print_stats(args.reference)

    elif args.command == "create-template":
        TemplateGenerator.create_template(args.reference)

    elif args.command == "prepare":
        if not args.csv:
            log.error("--csv required for 'prepare' command")
            return
        
        output_path = args.output
        if not output_path:
            # Default: replace .csv with .xlsx
            csv_path = Path(args.csv)
            output_path = str(csv_path.with_suffix(".xlsx"))
        
        log.info(f"Preparing scrub file...")
        ScrubPreparator.prepare(args.csv, args.reference, output_path)


if __name__ == "__main__":
    main()
