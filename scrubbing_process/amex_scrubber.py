"""
amex_scrubber.py — AEA AmEx Scrubber (corrected after full audit)
=================================================================
Output structure (matches reference):
  AmEx Load Raw    — 15 cols, untouched original
  AmEx All         — 17 cols, same order as raw, col15 always empty
  AEA_Posted       — 16 cols, sorted A-Z last name, col14 empty
  SBF_Posted       — 16 cols, sorted A-Z last name, col14 empty
  DEBT_Reviewed    — 16 cols, sorted A-Z last name, col14 empty
  Each entity tab: data rows + TOTAL row + blank row

Usage:
    python amex_scrubber.py \\
        --input     batch.xlsx \\
        --output    scrubbed.xlsx \\
        --reference same.xlsm        # Vendor List + Employee List sheets only
        --statement-date "Feb2026"

Dependencies: pip install pandas openpyxl openai
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from amex_rules import (
    RAW_COLS, AMEX_ALL_COLS, ENTITY_COLS,
    LEN_OVERHEAD, LEN_LIMIT,
    COL_H_FIND, COL_H_REPLACE,
    DESCRIPTION_RULES, DESCRIPTION_ABBREVIATIONS, TICKET_PREFIX_ABBREV,
    EXPENSE_CODE_REMAP, EXPENSE_CODE_DISALLOWED,
    VENDOR_TITLE_FIXES, VENDOR_ABBREVIATIONS, VENDOR_KEYWORD_MAP,
    ENTITY_DEFAULT, ENTITY_TABS,
    CAR_POLICY, CAR_FORMATS,
    MEAL_LIMITS, PERSONAL_PROJECT,
    GL_OVERRIDE_NOTES,
    PROJECT_DEPT_RULES, PROJECT_NAME_MAP,
    LLM_SYSTEM_PROMPT, LLM_ROW_SCHEMA,
)

from reference_loader import ReferenceData

try:
    from openai import OpenAI
    _OPENAI_OK = True
except ImportError:
    _OPENAI_OK = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("amex_scrubber")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from dotenv import load_dotenv

load_dotenv()  # reads variables from a .env file and sets them in os.environ


AZURE_API_KEY  = os.environ.get("AZURE_OPENAI_API_KEY")
AZURE_ENDPOINT = os.environ.get("AZURE_OPENAI_ENDPOINT")
AZURE_MODEL    = os.environ.get("AZURE_OPENAI_MODEL", "gpt-5-mini")

ORANGE = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED    = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
BOLD   = Font(bold=True)
LLM_APPLY_THRESHOLD = 0.80


# ══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class Row:
    idx: int

    # Original values — never changed after load
    first_name:     str    = ""
    middle_name:    str    = ""
    last_name:      str    = ""
    blank:          str    = ""
    tran_dt:        object = None
    description:    str    = ""
    amount:         float  = 0.0
    pay_type:       str    = ""
    expense_code:   str    = ""
    vendor_desc:    str    = ""
    vendor_name:    str    = ""
    project:        object = None
    cost_center:    str    = ""
    report_purpose: str    = ""
    employee_id:    str    = ""

    # Scrubbed output values
    desc_out:     str = ""
    pay_type_out: str = ""
    expense_out:  str = ""
    vendor_out:   str = ""
    entity_desc_out: str = ""
    entity_expense_out: str = ""
    entity_project_out: object = None
    entity_cost_center_out: str = ""

    # Metadata
    len_value: int  = 0
    entity:    str  = ""
    # Flags feed the summary report; review_comment is written only on entity tabs.
    flags:     list = field(default_factory=list)
    review_comment: str = ""
    changed:   bool = False
    desc_changed: bool = False
    pay_type_changed: bool = False
    expense_changed: bool = False
    vendor_changed: bool = False
    llm_confidence: float = 0.0
    llm_rule_ids: list = field(default_factory=list)

    def flag(self, msg: str, comment: str = ""):
        """Record a review flag and optional concise reviewer comment."""
        if msg and msg not in self.flags:
            self.flags.append(msg)
        comment = str(comment or "").strip()
        if comment and comment not in self.review_comment:
            self.review_comment = (
                comment if not self.review_comment
                else f"{self.review_comment} | {comment}"
            )

    def project_str(self) -> str:
        return str(self.project or "").strip()



# ══════════════════════════════════════════════════════════════════════════════
#  DETERMINISTIC SCRUBBER
# ══════════════════════════════════════════════════════════════════════════════
class DeterministicScrubber:

    @staticmethod
    def _scope_ok(scope: str, expense_l: str, desc_l: str) -> bool:
        return {
            "all":           True,
            "info_services": "info" in expense_l,
            "personal":      "personal" in desc_l,
            "airline":       "airline" in expense_l,
            "car_service":   "car service" in expense_l,
            "lodging":       "lodging" in expense_l,
        }.get(scope, True)

    def scrub_description(self, desc: str, expense_lower: str) -> tuple[str, bool]:
        """
        Apply verified description rules ONLY.
        Verified rules: Inflight Wifi casing, Personal expense strip,
        Meeting→Mtg, Meetings→Mtgs, Ticketing Fee→Tkting Fee, Ticket/→Tkt/
        """
        orig   = str(desc or "")
        result = orig

        # Apply DESCRIPTION_RULES (find-and-replace, scoped)
        for find, replace, scope, _note in DESCRIPTION_RULES:
            if not self._scope_ok(scope, expense_lower, result.lower()):
                continue
            result = result.replace(find, replace)

        # Apply word-boundary abbreviations
        for long_form, short_form in DESCRIPTION_ABBREVIATIONS:
            result = re.sub(
                r'\b' + re.escape(long_form) + r'\b',
                short_form,
                result,
                flags=re.IGNORECASE,
            )

        # Handle 'Ticket/' prefix abbreviation separately
        find_t, replace_t = TICKET_PREFIX_ABBREV
        result = result.replace(find_t, replace_t)

        if "car service" in expense_lower:
            result = re.sub(r"\s*-\s*", "-", result)
            result = re.sub(r"\s*/\s*", "/", result)

        result = re.sub(r"  +", " ", result).strip()
        return result, result != orig

    @staticmethod
    def scrub_pay_type(pay_type: str) -> tuple[str, bool]:
        orig = str(pay_type or "")
        new  = orig.replace(COL_H_FIND, COL_H_REPLACE)
        return new, new != orig

    @staticmethod
    def scrub_expense_code(code: str, desc: str, vendor_desc: str = "") -> tuple[str, bool]:
        c = str(code or "").strip()
        d = str(desc or "").strip().lower()

        if "inflight wifi" in d:
            new = "Info Services"
        elif (
            "tkt fee" in d
            or "ticketing fee" in d
            or "seat upgrade" in d
            or "checked bag fee" in d
            or d.startswith("exch tkt/")
            or d.startswith("refund/exch tkt/")
            or d.startswith("refund/tkt exch/")
            or d.startswith("tkt refund/")
        ) and "train" not in d and "bus" not in d and "boat" not in d:
            new = "Airline"
        elif "booking fee" in d and ("lodging" in d or "hotel" in d):
            new = "Lodging"
        elif any(k in d for k in ("train/", "bus.parking", "bus.fuel", "bus.rental", "travel insurance")):
            new = "Other Travel"
        else:
            new = EXPENSE_CODE_REMAP.get(c, c)
        return new, new != c

    @staticmethod
    def compute_len(desc: str, vendor: str) -> int:
        return len(str(desc or "")) + len(str(vendor or "")) + LEN_OVERHEAD

    def check_row(self, r: Row):
        """Flag checks — results go to r.flags (summary report only, NOT written to Excel)."""
        proj   = r.project_str()
        dept   = r.cost_center.strip()
        desc_l = r.desc_out.lower()
        exp_l  = r.expense_out.lower()

        # Expense code issues
        if r.expense_out in EXPENSE_CODE_DISALLOWED:
            r.flag(f"Disallowed expense code '{r.expense_out}'")

        if "inflight wifi" in desc_l and exp_l != "info services":
            r.flag(f"Inflight Wifi must use 'Info Services', not '{r.expense_out}'")

        if ("ticket" in desc_l or "tkt" in desc_l) and "fee" in desc_l:
            if exp_l == "other travel":
                r.flag("Ticketing fees must use expense code 'Airline'")

        # Project + dept
        required_dept = PROJECT_DEPT_RULES.get(proj)
        if required_dept:
            valid = required_dept if isinstance(required_dept, list) else [required_dept]
            if dept not in valid:
                r.flag(f"Project {proj} requires dept {valid}, got '{dept}'")

        proj_meta = PROJECT_NAME_MAP.get(proj, {})
        if proj_meta.get("requires_emp_id") and not r.employee_id.strip():
            r.flag("Personal project 1008 requires Employee ID")
        if proj_meta.get("board_only") and "bod" not in desc_l and "board" not in desc_l:
            r.flag(f"Project {proj} should be used only for board-meeting travel")

        if proj in GL_OVERRIDE_NOTES["14000"]["trigger_projects"]:
            r.flag("Prepaid project - review G/L 14000 at posting", GL_OVERRIDE_NOTES["14000"]["comment"])
        if (
            r.expense_out == GL_OVERRIDE_NOTES["58120"]["trigger_expense"]
            and proj in GL_OVERRIDE_NOTES["58120"]["trigger_projects"]
        ):
            r.flag("Corporate event Other expense - review G/L 58120 at posting", GL_OVERRIDE_NOTES["58120"]["comment"])
        if proj in GL_OVERRIDE_NOTES["58230"]["trigger_projects"]:
            r.flag("Intern event project - review G/L 58230 at posting", GL_OVERRIDE_NOTES["58230"]["comment"])
        if "holiday party" in desc_l and r.expense_out == "Car Service":
            r.flag("Holiday Party car - review G/L 58140 and entity booking", GL_OVERRIDE_NOTES["58140"]["comment"])

        # Meal limits
        for meal_kw, limit in MEAL_LIMITS.items():
            if meal_kw.lower() in desc_l and r.amount > limit:
                r.flag(
                    f"{meal_kw} ${r.amount:.2f} exceeds ${limit:.2f} limit",
                    f"Itemize overage to project {PERSONAL_PROJECT} as {meal_kw}/Overage/Personal",
                )
                break

        # Car-service policy rows need receipt-time verification.
        if exp_l == "car service" and any(
            phrase in desc_l
            for phrase in (
                "early arrival/home-office",
                "work late/office-home",
                "weekend/home-office",
                "weekend/office-home",
            )
        ):
            r.flag("Verify car-service receipt time/date against Early Arrival / Work Late / Weekend policy")

        if exp_l == "car service" and not any(fmt.lower() in desc_l for fmt in CAR_FORMATS.values()) and "/" not in desc_l:
            r.flag("Car Service description should include route/purpose/deal or one approved home-office format")

        # Refund
        if r.amount < 0:
            r.flag("Negative amount/refund - verify original booking in Sage")

        # LEN
        if r.len_value > LEN_LIMIT:
            r.flag(f"LEN {r.len_value} exceeds {LEN_LIMIT}")


# ══════════════════════════════════════════════════════════════════════════════
#  LLM VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════
class LLMValidator:

    def __init__(self, api_key: str, endpoint: str, model: str, batch_size: int = 10):
        if not _OPENAI_OK:
            raise ImportError("pip install openai")
        self.model      = model
        self.batch_size = batch_size
        self.client     = OpenAI(api_key=api_key, base_url=endpoint)
        self.rulebook   = self._load_rulebook()

    @staticmethod
    def _load_rulebook() -> str:
        rules_path = Path(__file__).with_name("aea_scrubbing_rules_with_examples.json")
        try:
            return rules_path.read_text(encoding="utf-8")
        except OSError:
            return ""

    def validate_batch(self, rows: list[Row]) -> list[dict]:
        if not rows:
            return []
        payload = [
            {
                "_i":          r.idx,
                "name":        f"{r.first_name} {r.last_name}".strip(),
                "date":        str(r.tran_dt or ""),
                "original_description": r.description,
                "current_description": r.entity_desc_out or r.desc_out,
                "amount":      round(r.amount, 2),
                "original_expense": r.expense_code,
                "current_expense": r.entity_expense_out or r.expense_out,
                "vendor_description": r.vendor_desc,
                "vendor_name":      r.vendor_out,
                "project":     str(r.entity_project_out if r.entity_project_out not in (None, "") else r.project_str()),
                "dept":        r.entity_cost_center_out or r.cost_center,
                "entity":      r.entity,
                "employee_id": r.employee_id,
            }
            for r in rows
        ]
        prompt = (
            f"Scrub {len(rows)} expense rows for entity-tab output using written rules as source of truth.\n\n"
            f"Schema: {json.dumps(LLM_ROW_SCHEMA, indent=2)}\n\n"
            f"Written rulebook JSON:\n{self.rulebook[:25000]}\n\n"
            f"Return JSON: {{\"results\": [...]}}\n\n"
            f"Rows:\n{json.dumps(payload, indent=2)}\n\n"
            "If any prompt text conflicts, written-rule JSON and this instruction win. "
            "Follow policy abbreviations such as Bus.Lunch/Bus.Dinner when row context supports them. "
            "Flag ONLY real issues. Do not suggest 'Business'→'Bus.' or ' to '→'-'."
        )
        prompt += (
            "\n\nImportant override: if any earlier prompt sentence conflicts with written-rule JSON, "
            "follow written rules and return conflict_note instead of forcing sample-specific wording. "
            "That means Business -> Bus., Car Service 'to' -> '-', and Car Service Transportation removal are allowed and expected when policy applies."
        )
        try:
            resp = self.client.chat.completions.create(
                model=self.model, max_completion_tokens=4000,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": LLM_SYSTEM_PROMPT},
                    {"role": "user",   "content": prompt},
                ],
            )
            raw = resp.choices[0].message.content or ""
            return self._parse(raw, len(rows))
        except Exception as exc:
            log.warning(f"LLM batch failed: {exc}")
            return [{}] * len(rows)

    @staticmethod
    def _parse(raw: str, n: int) -> list[dict]:
        raw = re.sub(r"^```(?:json)?|```$", "", raw.strip(), flags=re.MULTILINE).strip()
        try:
            data = json.loads(raw)
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                for v in data.values():
                    if isinstance(v, list):
                        return v
        except json.JSONDecodeError:
            pass
        return [{}] * n


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
class AmExScrubber:

    def __init__(
        self,
        input_path:     str,
        output_path:    str,
        reference_path: Optional[str] = None,
        statement_date: str           = "",
        llm_batch_size: int           = 10,
        azure_api_key:  str           = AZURE_API_KEY,
        azure_endpoint: str           = AZURE_ENDPOINT,
        azure_model:    str           = AZURE_MODEL,
    ):
        self.input_path     = Path(input_path)
        self.output_path    = Path(output_path)
        self.statement_date = statement_date

        self.ref = ReferenceData(reference_path)
        self.det = DeterministicScrubber()

        self.llm: Optional[LLMValidator] = None
        if not _OPENAI_OK:
            log.warning("openai not installed — LLM disabled")
        elif not azure_api_key or azure_api_key in ("", "xx"):
            log.warning("Azure API key not set — LLM disabled")
        else:
            self.llm = LLMValidator(
                api_key=azure_api_key, endpoint=azure_endpoint,
                model=azure_model, batch_size=llm_batch_size,
            )
            log.info(f"LLM ready: {azure_model}")

        self.rows:     list[Row]   = []
        self.raw_rows: list[tuple] = []
        self.report_purpose_label = self.input_path.stem

    # ── STEP 1: LOAD ─────────────────────────────────────────────────────────
    def load(self):
        log.info(f"Loading: {self.input_path}")
        ext = self.input_path.suffix.lower()

        if ext == ".csv":
            df       = pd.read_csv(self.input_path, dtype=str, keep_default_na=False)
            header   = list(df.columns)
            raw_data = [tuple(r) for r in df.itertuples(index=False, name=None)]
        elif ext in (".xlsx", ".xls", ".xlsm"):
            wb       = load_workbook(str(self.input_path), read_only=True, data_only=True)
            ws       = self._find_data_sheet(wb)
            all_rows = list(ws.iter_rows(values_only=True))
            header   = [str(c or "").strip() for c in all_rows[0]]
            # Keep only true data rows: first column must be an employee name
            raw_data = [
                r for r in all_rows[1:]
                if r[0] is not None
                and str(r[0]).strip() not in ("", "Employee First Name")
            ]
        else:
            raise ValueError(f"Unsupported file: {ext}")

        self.raw_rows = raw_data
        log.info(f"  {len(raw_data)} data rows | {len(header)} columns")

        col_idx = {str(h).strip(): i for i, h in enumerate(header) if h}

        def get(row, name, default=""):
            i = col_idx.get(name)
            if i is None or i >= len(row):
                return default
            return default if row[i] is None else row[i]

        def ss(v, default="") -> str:
            s = str(v).strip() if v is not None else ""
            return default if s.lower() in ("none", "nan") else s

        def sf(v) -> float:
            try:
                return float(v or 0)
            except (ValueError, TypeError):
                return 0.0

        def parse_date(v):
            if isinstance(v, datetime) or v is None:
                return v
            s = ss(v)
            if not s:
                return None
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
            return v

        def parse_project(v):
            s = ss(v)
            if not s:
                return ""
            return int(s) if s.isdigit() else s

        self.rows = []
        for idx, raw in enumerate(raw_data):
            r = Row(
                idx            = idx,
                first_name     = ss(get(raw, "Employee First Name")),
                middle_name    = ss(get(raw, "Employee Middle Name")),
                last_name      = ss(get(raw, "Employee Last Name")),
                blank          = ss(get(raw, "Blank/Placeholder")),
                tran_dt        = parse_date(get(raw, "Report Entry Transaction Date")),
                description    = ss(get(raw, "Report Entry Description")),
                amount         = sf(get(raw, "Journal Amount")),
                pay_type       = ss(get(raw, "Report Entry Payment Type Name")),
                expense_code   = ss(get(raw, "Report Entry Expense Type Name")),
                vendor_desc    = ss(get(raw, "Report Entry Vendor Description")),
                vendor_name    = ss(get(raw, "Report Entry Vendor Name")),
                project        = parse_project(get(raw, "Project")),
                cost_center    = ss(get(raw, "Cost Center")),
                report_purpose = self.report_purpose_label,
                employee_id    = ss(get(raw, "Employee ID")),
            )
            r.desc_out     = r.description
            r.pay_type_out = r.pay_type
            r.expense_out  = r.expense_code
            r.vendor_out   = r.vendor_name
            r.entity_desc_out = r.desc_out
            r.entity_expense_out = r.expense_out
            r.entity_project_out = r.project
            r.entity_cost_center_out = r.cost_center
            self.rows.append(r)

        log.info(f"  Loaded {len(self.rows)} rows")

    @staticmethod
    def _find_data_sheet(wb) -> object:
        for sname in wb.sheetnames:
            ws  = wb[sname]
            row = next(ws.iter_rows(max_row=1, values_only=True), None)
            if row and any("Report Entry Description" in str(c or "") for c in row):
                log.info(f"  Data sheet: '{sname}'")
                return ws
        log.warning("No data sheet found — using first sheet")
        return wb[wb.sheetnames[0]]

    # ── STEP 2: DETERMINISTIC SCRUB ──────────────────────────────────────────
    def deterministic_scrub(self):
        log.info("Step 2 — deterministic scrub")

        for r in self.rows:
            exp_l = r.expense_code.lower()

            # Description (verified rules only)
            new_desc, dc = self.det.scrub_description(r.description, exp_l)
            if dc:
                r.desc_out = new_desc
                r.desc_changed = True

            # Pay type
            new_pay, pc = self.det.scrub_pay_type(r.pay_type)
            if pc:
                r.pay_type_out = new_pay
                r.pay_type_changed = True

            # Expense code
            new_exp, ec = self.det.scrub_expense_code(
                r.expense_code,
                r.desc_out,
                r.vendor_desc,
            )
            if ec:
                r.expense_out = new_exp
                r.expense_changed = True

            # Vendor name
            looked_up = self.ref.lookup_vendor(r.vendor_desc)
            if looked_up and looked_up != r.vendor_out:
                r.vendor_out = looked_up
                r.vendor_changed = True

            # Entity
            r.entity = self.ref.lookup_entity(r.employee_id)

            # LEN
            r.len_value = self.det.compute_len(r.desc_out, r.vendor_out)

            # Flag checks (summary report only)
            self.det.check_row(r)

            r.changed = (
                r.desc_changed
                or r.pay_type_changed
                or r.expense_changed
                or r.vendor_changed
            )

        self._mirror_refund_expense_codes()

        for r in self.rows:
            r.entity_desc_out = r.desc_out
            r.entity_expense_out = r.expense_out
            r.entity_project_out = r.project
            r.entity_cost_center_out = r.cost_center

        self._flag_duplicate_office_home_and_trip_projects()

        n_changed = sum(r.changed for r in self.rows)
        n_flagged = sum(bool(r.flags) for r in self.rows)

        log.info(f"  {n_changed} rows changed, {n_flagged} flagged")

    def _mirror_refund_expense_codes(self):
        """Mirror refund expense type from a unique matching positive charge."""
        originals: dict[tuple[str, str, str, float], list[Row]] = {}
        for r in self.rows:
            if r.amount <= 0:
                continue
            key = (
                r.employee_id,
                r.project_str(),
                r.cost_center,
                round(r.amount, 2),
            )
            originals.setdefault(key, []).append(r)

        for r in self.rows:
            if r.amount >= 0:
                continue
            key = (
                r.employee_id,
                r.project_str(),
                r.cost_center,
                round(abs(r.amount), 2),
            )
            matches = originals.get(key, [])
            if len(matches) != 1:
                continue
            source = matches[0]
            if r.expense_out != source.expense_out:
                r.expense_out = source.expense_out
                r.expense_changed = True
            if r.desc_out == "Refund/Personal" and source.desc_out != "Personal":
                r.desc_out = f"Refund/{source.desc_out}"
                r.desc_changed = True

            r.changed = (
                r.desc_changed
                or r.pay_type_changed
                or r.expense_changed
                or r.vendor_changed
            )

    def _flag_duplicate_office_home_and_trip_projects(self):
        """Flag cross-row policy issues that require reviewing nearby rows together."""
        office_home_rows: dict[tuple[str, object], list[Row]] = {}
        trip_rows: dict[tuple[str, object], list[Row]] = {}

        for r in self.rows:
            employee_key = r.employee_id or f"{r.first_name}|{r.last_name}"
            day_key = (employee_key, r.tran_dt)
            desc_l = str(r.desc_out or "").lower()
            exp_l = str(r.expense_out or "").lower()

            if desc_l == "work late/office-home":
                office_home_rows.setdefault(day_key, []).append(r)

            if exp_l in {"airline", "lodging", "car service", "meals", "other travel", "other"} and any(
                token in desc_l
                for token in (
                    "lodging",
                    "travel meal",
                    "tkt fee",
                    "rt:",
                    "site visit",
                    "bod mtg",
                    "mgmt mtg",
                    "strategy mtg",
                    "portfolio mtg",
                    "hotel",
                    "train/",
                    "bus.parking",
                    "bus.fuel",
                    "bus.rental",
                )
            ):
                trip_rows.setdefault(day_key, []).append(r)

        for grouped_rows in office_home_rows.values():
            if len(grouped_rows) > 1:
                for r in grouped_rows:
                    r.flag(
                        "More than one Work Late/Office-Home row exists for same employee/date",
                        "Check whether one row is a tip on same ride or service date should be prior day after midnight",
                    )

        for grouped_rows in trip_rows.values():
            projects = {
                row.project_str()
                for row in grouped_rows
                if row.project_str() and row.project_str() != PERSONAL_PROJECT
            }
            if len(projects) > 1:
                for r in grouped_rows:
                    r.flag(
                        "Same-day trip-like rows use multiple project codes",
                        f"Review whether lodging/air/car/meals should share one project code: {', '.join(sorted(projects))}",
                    )

    # ── STEP 3: LLM VALIDATION ───────────────────────────────────────────────
    def llm_validate(self):
        if not self.llm:
            log.info("Step 3 — LLM skipped")
            return

        log.info("Step 3 — LLM validation")
        total = len(self.rows)
        bs    = self.llm.batch_size

        for start in range(0, total, bs):
            end   = min(start + bs, total)
            batch = self.rows[start:end]
            log.info(f"  Batch {start+1}–{end}/{total}")
            results = self.llm.validate_batch(batch)

            for r, lr in zip(batch, results):
                if not lr:
                    continue
                try:
                    confidence = float(lr.get("confidence", 0) or 0)
                except (TypeError, ValueError):
                    confidence = 0.0
                r.llm_confidence = max(r.llm_confidence, confidence)

                rule_ids = lr.get("rule_ids_applied") or []
                if isinstance(rule_ids, list):
                    for rule_id in rule_ids:
                        if rule_id and rule_id not in r.llm_rule_ids:
                            r.llm_rule_ids.append(rule_id)

                if lr.get("conflict_note"):
                    r.flag("Rule/reference conflict noted by LLM", lr.get("conflict_note") or "")

                if lr.get("description_fixed") and confidence >= LLM_APPLY_THRESHOLD:
                    r.entity_desc_out = str(lr["description_fixed"]).strip()
                if lr.get("vendor_fixed"):
                    r.vendor_out = lr["vendor_fixed"]
                    r.vendor_changed = (r.vendor_out != r.vendor_name)
                if lr.get("expense_code_fixed") and confidence >= LLM_APPLY_THRESHOLD:
                    r.entity_expense_out = str(lr["expense_code_fixed"]).strip()
                if lr.get("project_fixed") and confidence >= LLM_APPLY_THRESHOLD:
                    raw_project = str(lr["project_fixed"]).strip()
                    r.entity_project_out = int(raw_project) if raw_project.isdigit() else raw_project
                if lr.get("cost_center_fixed") and confidence >= LLM_APPLY_THRESHOLD:
                    r.entity_cost_center_out = str(lr["cost_center_fixed"]).strip()
                if lr.get("flag"):
                    r.flag(
                        lr.get("flag_reason") or "LLM flagged",
                        lr.get("comments") or "",
                    )

                r.changed = (
                    r.desc_changed
                    or r.pay_type_changed
                    or r.expense_changed
                    or r.vendor_changed
                    or r.entity_desc_out != r.desc_out
                    or r.entity_expense_out != r.expense_out
                    or r.entity_project_out != r.project
                    or r.entity_cost_center_out != r.cost_center
                )

            time.sleep(0.3)

        log.info("  LLM done")

    # ── STEP 4: ENTITY SPLIT ─────────────────────────────────────────────────
    def _split_by_entity(self, rows: list[Row]) -> dict[str, list[Row]]:
        """Split into entity buckets, sort A-Z last name then date."""
        buckets: dict[str, list[Row]] = {}
        for r in rows:
            entity = r.entity if r.entity in ENTITY_TABS else ENTITY_DEFAULT
            buckets.setdefault(entity, []).append(r)

        for entity in buckets:
            buckets[entity].sort(
                key=lambda r: (str(r.last_name or "").upper(), r.tran_dt or "")
            )
            total = sum(r.amount for r in buckets[entity])
            log.info(f"  {entity}: {len(buckets[entity])} rows  ${total:,.2f}")

        return buckets

    # ── STEP 5: WRITE OUTPUT ─────────────────────────────────────────────────
    def write_output(self):
        log.info(f"Writing: {self.output_path}")
        entity_data = self._split_by_entity(self.rows)

        wb = Workbook()
        wb.remove(wb.active)

        self._write_raw_tab(wb)
        self._write_amex_all_tab(wb, self.rows)   # original batch order

        for entity, tab_name in ENTITY_TABS.items():
            rows = entity_data.get(entity, [])
            if not rows and entity == "GROWTH":
                continue
            self._write_entity_tab(wb, tab_name, entity, rows)

        wb.save(str(self.output_path))
        log.info(f"  Saved: {self.output_path}")

    def _bold_header(self, ws):
        for cell in ws[1]:
            cell.font = BOLD

    def _write_raw_tab(self, wb: Workbook):
        """AmEx Load Raw — 15 cols, completely untouched."""
        ws = wb.create_sheet("AmEx Load Raw")
        ws.append(list(RAW_COLS))
        self._bold_header(ws)
        for idx, raw in enumerate(self.raw_rows):
            ws.append([raw[i] if i < len(raw) else None for i in range(15)])
            xr = ws.max_row
            row = self.rows[idx]
            ws.cell(xr, 5, row.tran_dt)
            ws.cell(xr, 7, row.amount)
            ws.cell(xr, 12, row.project)
            ws.cell(xr, 14, row.report_purpose)

    def _write_amex_all_tab(self, wb: Workbook, rows: list[Row]):
        """
        AmEx All — 17 cols, same order as raw.
        col15 (None header) is ALWAYS EMPTY — matches reference exactly.
        """
        ws = wb.create_sheet("AmEx All")
        ws.append(list(AMEX_ALL_COLS))
        self._bold_header(ws)

        total_amt = 0.0
        for r in rows:
            old_expense_out = r.expense_out
            if r.amount < 0 and str(r.desc_out or "").strip().lower() == "personal":
                r.expense_out = "Other" if r.expense_code == "Miscellaneous" else r.expense_code
            xr = ws.max_row + 1
            ws.append([
                r.first_name,
                r.middle_name or None,
                r.last_name,
                r.blank or None,
                r.tran_dt,
                r.desc_out,       # col 6 — scrubbed description
                r.amount,
                r.pay_type_out,   # col 8 — CBCP replaced
                r.expense_out,    # col 9 — expense code remapped
                r.vendor_desc,    # col 10 — original VendorDesc KEPT
                r.vendor_out,     # col 11 — cleaned VendorName
                r.project,
                r.cost_center,
                r.report_purpose,
                r.employee_id,
                None,             # col 16 — ALWAYS EMPTY (human fills in entity tabs)
                f"=LEN(F{xr}&K{xr})+12",  # col 17 - LEN formula
            ])
            total_amt += r.amount
            r.expense_out = old_expense_out

            # Red: LEN exceeded
            if r.len_value > LEN_LIMIT:
                ws.cell(xr, 17).fill = RED
                ws.cell(xr, 17).font = BOLD

        # TOTAL row + blank row
        tr = ws.max_row + 1
        ws.cell(tr, 6, "TOTAL - AmEx All").font = BOLD
        ws.cell(tr, 7, f"=SUM(G2:G{tr-1})").font = BOLD
        ws.append([None] * 17)

    def _write_entity_tab(self, wb: Workbook, tab_name: str,
                          entity: str, rows: list[Row]):
        """
        Entity tab — 16 cols (VendorDesc dropped), sorted A-Z last name.
        col15 (None header) stores concise comments only for flagged rows.
        TOTAL row + blank row at end.
        """
        ws = wb.create_sheet(tab_name)
        ws.append(list(ENTITY_COLS))
        self._bold_header(ws)

        total_amt = 0.0
        for r in rows:
            desc_out = self._entity_tab_description(r)
            expense_out = self._entity_tab_expense(r, desc_out)
            len_value = self.det.compute_len(desc_out, r.vendor_out)
            xr = ws.max_row + 1
            ws.append([
                r.first_name,
                r.middle_name or None,
                r.last_name,
                r.blank or None,
                r.tran_dt,
                desc_out,         # col 6
                r.amount,         # col 7
                r.pay_type_out,   # col 8
                expense_out,      # col 9
                r.vendor_out,     # col 10 — VendorDesc dropped; VendorName here
                r.entity_project_out,        # col 11
                r.entity_cost_center_out,    # col 12
                r.report_purpose, # col 13
                r.employee_id,    # col 14
                r.review_comment or None,  # col 15 — reviewer comments for flagged rows
                f"=LEN(F{xr}&J{xr})+12",  # col 16 - LEN formula
            ])
            total_amt += r.amount

            if r.desc_changed or desc_out != r.desc_out:
                ws.cell(xr, 6).fill = ORANGE
            if expense_out != r.expense_code:
                ws.cell(xr, 9).fill = ORANGE
            if len_value > LEN_LIMIT:
                ws.cell(xr, 16).fill = RED
                ws.cell(xr, 16).font = BOLD

        # TOTAL row + blank row
        tr = ws.max_row + 1
        ws.cell(tr, 6, f"TOTAL - {entity}").font = BOLD
        ws.cell(tr, 7, f"=SUM(G2:G{tr-1})").font = BOLD
        ws.append([None] * 16)

        log.info(f"  '{tab_name}': {len(rows)} rows  ${total_amt:,.2f}")

    def _entity_tab_description(self, r: Row) -> str:
        """Apply only general written-rule formatting, not sample-specific wording."""
        desc = str(r.entity_desc_out or r.desc_out or "").strip()

        if r.amount < 0 and desc.lower() == "personal":
            desc = "Refund/Personal"

        desc = re.sub(r"\bRT:\s+", "RT:", desc)
        desc = desc.replace("InFlight Wifi", "Inflight Wifi")
        desc = desc.replace("Business Lunch", "Bus.Lunch")
        desc = desc.replace("Business Dinner", "Bus.Dinner")
        desc = desc.replace("Bus. Lunch", "Bus.Lunch")
        desc = desc.replace("Bus. Dinner", "Bus.Dinner")
        desc = desc.replace("Work Late/Office to Home", "Work Late/Office-Home")
        desc = desc.replace("Early Arrival/Home to Office", "Early Arrival/Home-Office")
        desc = desc.replace("Weekend/Home to Office", "Weekend/Home-Office")
        desc = desc.replace("Weekend/Office to Home", "Weekend/Office-Home")

        if desc.startswith("Parking/"):
            desc = f"Bus.{desc}"

        return re.sub(r"  +", " ", desc).strip()

    @staticmethod
    def _entity_tab_expense(r: Row, desc_out: str) -> str:
        """Apply general written-rule expense remaps on entity tabs only."""
        code = str(r.entity_expense_out or r.expense_out or "").strip()
        desc_l = str(desc_out or "").strip().lower()
        vendor_l = str(r.vendor_desc or "").strip().lower()

        if "inflight wifi" in desc_l:
            return "Info Services"
        if "subscription" in desc_l and code in {"PubSub", "Miscellaneous", "Other", "Software"}:
            return "Info Services"
        if code != "Car Service" and (
            "lodging" in desc_l
            or "booking fee" in desc_l
            or "room rental" in desc_l
            or "room block" in desc_l
        ):
            return "Lodging"
        if code != "Car Service" and (
            "train" in desc_l
            or "bus.parking" in desc_l
            or "bus.fuel" in desc_l
            or "travel insurance" in desc_l
        ):
            return "Other Travel"
        if (
            "tkt fee" in desc_l
            and "train" not in desc_l
        ):
            return "Airline"
        return code

    # ── STEP 6: SAGE LOAD CSVs ───────────────────────────────────────────────
    def write_load_csvs(self, output_dir: str = ".") -> list[Path]:
        """One CSV per entity for Sage import. Report Purpose cleared."""
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)

        entity_data = self._split_by_entity(self.rows)
        date_tag    = self.statement_date or "XX"
        paths       = []
        cols14      = ENTITY_COLS[:14]  # cols 0–13

        for entity, tab_name in ENTITY_TABS.items():
            rows = entity_data.get(entity, [])
            if not rows:
                continue

            records = []
            for r in rows:
                desc_out = self._entity_tab_description(r)
                records.append({
                    cols14[0]:  r.first_name,
                    cols14[1]:  r.middle_name or "",
                    cols14[2]:  r.last_name,
                    cols14[3]:  r.blank or "",
                    cols14[4]:  r.tran_dt,
                    cols14[5]:  desc_out,
                    cols14[6]:  r.amount,
                    cols14[7]:  r.pay_type_out,
                    cols14[8]:  self._entity_tab_expense(r, desc_out),
                    cols14[9]:  r.vendor_out,
                    cols14[10]: r.entity_project_out,
                    cols14[11]: r.entity_cost_center_out,
                    cols14[12]: "",  # Report Purpose cleared
                    cols14[13]: r.employee_id,
                })

            fpath = out / f"AmEx_{date_tag}_{entity}.csv"
            pd.DataFrame(records).to_csv(fpath, index=False)
            paths.append(fpath)
            log.info(f"  CSV: {fpath}  ({len(rows)} rows)")

        return paths

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    def print_summary(self):
        total    = len(self.rows)
        changed  = sum(r.changed for r in self.rows)
        flagged  = sum(bool(r.flags) for r in self.rows)
        len_err  = sum(1 for r in self.rows if r.len_value > LEN_LIMIT)
        total_usd = sum(r.amount for r in self.rows)

        print("\n" + "=" * 65)
        print("  SCRUB SUMMARY")
        print("=" * 65)
        print(f"  Batch total:         ${total_usd:>12,.2f}")
        print(f"  Total rows:          {total:>5}")
        print(f"  Changed rows:        {changed:>5}  (orange in Excel)")
        print(f"  Flagged for review:  {flagged:>5}  (summary report below)")
        print(f"  LEN > {LEN_LIMIT}:          {len_err:>5}  (red in Excel)")
        print("=" * 65)

        if flagged:
            print("\n  FLAGGED ROWS (for human reviewer):")
            for r in self.rows:
                if r.flags:
                    name = f"{r.first_name} {r.last_name}"
                    desc = r.desc_out[:42]
                    print(f"  Row {r.idx+2:>4}  [{r.entity}]  {name:<22}  {desc}")
                    for f in r.flags:
                        print(f"             => {f}")
                    if r.review_comment:
                        print(f"             note: {r.review_comment}")
        print()

    # ── FULL PIPELINE ─────────────────────────────────────────────────────────
    def run(self) -> tuple[list[Row], list[Path]]:
        self.load()
        self.deterministic_scrub()
        self.llm_validate()
        self.write_output()
        csv_paths = self.write_load_csvs(
            str(self.output_path.parent / "load_csvs")
        )
        self.print_summary()
        return self.rows, csv_paths


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════
def main():
    p = argparse.ArgumentParser(description="AEA AmEx Scrubber")
    p.add_argument("--input",           required=True)
    p.add_argument("--output",          required=True)
    p.add_argument("--llm-batch-size",  type=int, default=10)
    args = p.parse_args()

    AmExScrubber(
        input_path     = args.input,
        output_path    = args.output,
        llm_batch_size = args.llm_batch_size,
    ).run()


if __name__ == "__main__":
    main()
