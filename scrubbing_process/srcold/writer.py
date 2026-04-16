"""
writer.py
─────────
Excel output writer.

Produces:
  Tab "AmEx Load Raw"    — 15 cols, untouched original data
  Tab "AmEx All"         — 17 cols, deterministic scrub applied
  Tab "AEA_Posted"       — 16 cols, entity data + LLM fixes
  Tab "SBF_Posted"       — 16 cols
  Tab "DEBT_Reviewed"    — 16 cols
  Tab "GROWTH_Reviewed"  — 16 cols (omitted if empty)

Styling:
  Orange fill  → cells that were changed by scrubber
  Red fill     → LEN > 70 in the LEN column
  Bold         → header rows and TOTAL rows
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .config import PROJECT_CFG, EXPENSE_CFG
from .models import Row
from .config import RAW_COLS, AMEX_ALL_COLS, ENTITY_COLS
from .rules_engine import RulesEngine

log = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────
ORANGE = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
RED    = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
BOLD   = Font(bold=True)


class ExcelWriter:

    def __init__(self, rules: RulesEngine):
        self.rules = rules

    # ── Main entry point ──────────────────────────────────────────────────────
    def write(
        self,
        output_path: Path,
        rows:        List[Row],
        raw_rows:    list,
        batch_total_label: str = "",
    ) -> None:
        log.info("Writing output: %s", output_path)
        entity_data = self._split_by_entity(rows)

        wb = Workbook()
        wb.remove(wb.active)

        self._write_raw_tab(wb, rows, raw_rows)
        self._write_amex_all_tab(wb, rows, batch_total_label)

        for entity, tab_name in PROJECT_CFG.entity_tabs.items():
            entity_rows = entity_data.get(entity, [])
            if not entity_rows and entity == "GROWTH":
                continue
            self._write_entity_tab(wb, tab_name, entity, entity_rows)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        log.info("Saved: %s", output_path)

    # ── Entity split ──────────────────────────────────────────────────────────
    def _split_by_entity(self, rows: List[Row]) -> Dict[str, List[Row]]:
        buckets: Dict[str, List[Row]] = {}
        for r in rows:
            entity = (
                r.entity
                if r.entity in PROJECT_CFG.entity_tabs
                else PROJECT_CFG.entity_default
            )
            buckets.setdefault(entity, []).append(r)
        for entity, elist in buckets.items():
            elist.sort(key=lambda r: (
                str(r.last_name or "").upper(),
                str(r.tran_dt or ""),
            ))
            total = sum(r.amount for r in elist)
            log.info("  %s: %d rows  $%,.2f", entity, len(elist), total)
        return buckets

    # ── Raw tab ───────────────────────────────────────────────────────────────
    def _write_raw_tab(self, wb: Workbook, rows: List[Row], raw_rows: list) -> None:
        ws = wb.create_sheet("AmEx Load Raw")
        ws.append(list(RAW_COLS))
        self._bold_header(ws)
        for idx, raw in enumerate(raw_rows):
            if idx >= len(rows):
                break
            r = rows[idx]
            ws.append([
                raw[i] if i < len(raw) else None for i in range(15)
            ])
            xr = ws.max_row
            ws.cell(xr, 5).value  = r.tran_dt
            ws.cell(xr, 7).value  = r.amount
            ws.cell(xr, 12).value = r.project
            ws.cell(xr, 14).value = r.report_purpose

    # ── AmEx All tab ─────────────────────────────────────────────────────────
    def _write_amex_all_tab(
        self, wb: Workbook, rows: List[Row], batch_total_label: str
    ) -> None:
        ws = wb.create_sheet("AmEx All")
        ws.append(list(AMEX_ALL_COLS))
        self._bold_header(ws)

        for r in rows:
            xr = ws.max_row + 1
            ws.append([
                r.first_name,
                r.middle_name or None,
                r.last_name,
                r.blank or None,
                r.tran_dt,
                r.desc_out,
                r.amount,
                r.pay_type_out,
                r.expense_out,
                r.vendor_desc,   # original VendorDesc kept in AmEx All
                r.vendor_out,
                r.project,
                r.cost_center,
                r.report_purpose,
                r.employee_id,
                None,            # col 16 always empty
                f"=LEN(F{xr}&K{xr})+{PROJECT_CFG.len_overhead}",
            ])
            if r.desc_changed:
                ws.cell(xr, 6).fill = ORANGE
            if r.expense_changed:
                ws.cell(xr, 9).fill = ORANGE
            if r.pay_type_changed:
                ws.cell(xr, 8).fill = ORANGE
            if r.vendor_changed:
                ws.cell(xr, 11).fill = ORANGE
            if r.len_value > PROJECT_CFG.len_limit:
                ws.cell(xr, 17).fill = RED
                ws.cell(xr, 17).font = BOLD

        # TOTAL row
        last_data = ws.max_row
        tr = last_data + 1
        label = batch_total_label or "TOTAL - AmEx All"
        ws.cell(tr, 6, label).font = BOLD
        ws.cell(tr, 7, f"=SUM(G2:G{last_data})").font = BOLD
        ws.append([None] * 17)

    # ── Entity tab ───────────────────────────────────────────────────────────
    def _write_entity_tab(
        self, wb: Workbook, tab_name: str, entity: str, rows: List[Row]
    ) -> None:
        ws = wb.create_sheet(tab_name)
        ws.append(list(ENTITY_COLS))
        self._bold_header(ws)

        for r in rows:
            desc_out    = self._final_entity_description(r)
            expense_out = self._final_entity_expense(r, desc_out)
            len_val     = self.rules.compute_len(desc_out, r.vendor_out)
            xr          = ws.max_row + 1

            ws.append([
                r.first_name,
                r.middle_name or None,
                r.last_name,
                r.blank or None,
                r.tran_dt,
                desc_out,
                r.amount,
                r.pay_type_out,
                expense_out,
                r.vendor_out,
                r.entity_project_out,
                r.entity_cost_center_out,
                r.report_purpose,
                r.employee_id,
                r.review_comment or None,
                f"=LEN(F{xr}&J{xr})+{PROJECT_CFG.len_overhead}",
            ])

            # Orange: changed description or expense
            if r.desc_changed or desc_out != r.desc_out:
                ws.cell(xr, 6).fill = ORANGE
            if expense_out != r.expense_code:
                ws.cell(xr, 9).fill = ORANGE
            if len_val > PROJECT_CFG.len_limit:
                ws.cell(xr, 16).fill = RED
                ws.cell(xr, 16).font = BOLD

        # TOTAL row
        last_data = ws.max_row
        tr = last_data + 1
        ws.cell(tr, 6, f"TOTAL - {entity}").font = BOLD
        ws.cell(tr, 7, f"=SUM(G2:G{last_data})").font = BOLD
        ws.append([None] * 16)

        log.info("  Tab '%s': %d rows  $%,.2f",
                 tab_name, len(rows), sum(r.amount for r in rows))

    # ── Entity description final pass ─────────────────────────────────────────
    def _final_entity_description(self, r: Row) -> str:
        desc = str(r.entity_desc_out or r.desc_out or "").strip()
        desc = self.rules.apply_entity_tab_rules(desc)
        # Refund prefix for negative personal
        if r.amount < 0 and desc.lower() == "personal":
            desc = "Refund/Personal"
        return desc

    # ── Entity expense final pass ─────────────────────────────────────────────
    @staticmethod
    def _final_entity_expense(r: Row, desc_out: str) -> str:
        code  = str(r.entity_expense_out or r.expense_out or "").strip()
        d_low = str(desc_out or "").lower()

        if "inflight wifi" in d_low:
            return "Info Services"
        if "subscription" in d_low and code in {"PubSub", "Miscellaneous", "Other", "Software"}:
            return "Info Services"
        if code != "Car Service" and any(
            k in d_low for k in ("lodging", "booking fee", "room rental", "room block")
        ):
            return "Lodging"
        if code != "Car Service" and any(
            k in d_low for k in ("train/", "bus.parking", "bus.fuel", "travel insurance")
        ):
            return "Other Travel"
        if "tkt fee" in d_low and "train" not in d_low:
            return "Airline"
        return code

    # ── Helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _bold_header(ws) -> None:
        for cell in ws[1]:
            cell.font = BOLD

    # ── Sage load CSVs ────────────────────────────────────────────────────────
    def write_load_csvs(
        self, output_dir: Path, rows: List[Row], statement_date: str = "XX"
    ) -> List[Path]:
        """One CSV per entity, columns A–N, Report Purpose cleared."""
        import pandas as pd
        output_dir.mkdir(parents=True, exist_ok=True)

        entity_data = self._split_by_entity(rows)
        entity_cols = ENTITY_COLS[:14]
        paths: List[Path] = []

        for entity, tab_name in PROJECT_CFG.entity_tabs.items():
            elist = entity_data.get(entity, [])
            if not elist:
                continue
            records = []
            for r in elist:
                desc_out    = self._final_entity_description(r)
                expense_out = self._final_entity_expense(r, desc_out)
                records.append({
                    entity_cols[0]:  r.first_name,
                    entity_cols[1]:  r.middle_name or "",
                    entity_cols[2]:  r.last_name,
                    entity_cols[3]:  r.blank or "",
                    entity_cols[4]:  r.tran_dt,
                    entity_cols[5]:  desc_out,
                    entity_cols[6]:  r.amount,
                    entity_cols[7]:  r.pay_type_out,
                    entity_cols[8]:  expense_out,
                    entity_cols[9]:  r.vendor_out,
                    entity_cols[10]: r.entity_project_out,
                    entity_cols[11]: r.entity_cost_center_out,
                    entity_cols[12]: "",   # Report Purpose cleared
                    entity_cols[13]: r.employee_id,
                })
            fpath = output_dir / f"AmEx_{statement_date}_{entity}.csv"
            pd.DataFrame(records).to_csv(fpath, index=False)
            paths.append(fpath)
            log.info("  CSV: %s  (%d rows)", fpath.name, len(elist))

        return paths
