"""
reference_loader.py
────────────────────
Loads the Vendor List and Employee List from the reference / batch Excel file.
The same batch file that contains the data also carries these lookup sheets.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Optional

log = logging.getLogger(__name__)


class ReferenceData:
    """
    Vendor List:   Original (uppercase key) → Modified (canonical name)
    Employee List: Employee ID → (Project, Entity, Name)
    """

    def __init__(self, path: Optional[str] = None):
        self.vendor_list:   Dict[str, str]  = {}
        self.employee_list: Dict[str, dict] = {}
        if path:
            self.load(path)

    def load(self, path: str) -> None:
        from openpyxl import load_workbook
        p = Path(path)
        if not p.exists():
            log.warning("Reference file not found: %s", p)
            return
        try:
            wb = load_workbook(str(p), read_only=True, data_only=True)
            self._load_vendor_list(wb)
            self._load_employee_list(wb)
            log.info(
                "Reference loaded: %d vendors, %d employees",
                len(self.vendor_list), len(self.employee_list),
            )
        except Exception as exc:
            log.warning("Failed to load reference file %s: %s", p.name, exc)

    def _load_vendor_list(self, wb) -> None:
        if "Vendor List" not in wb.sheetnames:
            return
        ws = wb["Vendor List"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            if row[0] is None:
                continue
            original = str(row[0] or "").strip().upper()
            modified = str(row[1] or "").strip() if len(row) > 1 else ""
            if original and modified:
                self.vendor_list[original] = modified

    def _load_employee_list(self, wb) -> None:
        if "Employee List" not in wb.sheetnames:
            return
        ws   = wb["Employee List"]
        rows = list(ws.iter_rows(values_only=True))
        # Header: EMPLOYEE | Paycom EE ID# | PROJECT CODE | ENTITY
        for row in rows[1:]:
            if row[0] is None:
                continue
            name   = str(row[0] or "").strip()
            emp_id = str(row[1] or "").strip().upper()
            proj   = str(row[2] or "").strip() if len(row) > 2 else ""
            entity = str(row[3] or "").strip() if len(row) > 3 else ""
            if emp_id:
                self.employee_list[emp_id] = {
                    "name":    name,
                    "project": proj,
                    "entity":  entity,
                }

    # ── Lookups ───────────────────────────────────────────────────────────────
    def lookup_vendor(self, vendor_desc: str) -> Optional[str]:
        """Exact match (case-insensitive) against Vendor List."""
        return self.vendor_list.get(str(vendor_desc or "").strip().upper())

    def lookup_entity(self, employee_id: str) -> str:
        """Return entity code for an employee ID, default 'AEA'."""
        from .config import PROJECT_CFG
        info = self.employee_list.get(str(employee_id or "").strip().upper(), {})
        entity = str(info.get("entity", "") or "").strip()
        return entity if entity in PROJECT_CFG.entity_tabs else PROJECT_CFG.entity_default

    def lookup_employee(self, employee_id: str) -> dict:
        return self.employee_list.get(str(employee_id or "").strip().upper(), {})
