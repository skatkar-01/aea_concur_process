"""
transaction_memory.py
─────────────────────
Reads "transaction memory" Excel files (Alers - $2,802.00.xlsx pattern).
These files contain employee expense reports from Concur with:
  - Employee Report tab  : basic employee info
  - Transactions tab     : line-level transaction data
  - Receipts tab         : receipt details (ticket numbers, amounts)
  - Reconciliation tab   : reconciliation entries
  - Summary tab          : totals

The memory is used to:
  1. Enrich batch rows with receipt details (ticket numbers, flight routes)
  2. Verify refund amounts against original charges
  3. Provide context for LLM description formatting
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

log = logging.getLogger(__name__)


@dataclass
class Transaction:
    txn_id:           str    = ""
    date:             str    = ""
    expense_type:     str    = ""
    business_purpose: str    = ""
    vendor_desc:      str    = ""
    payment_type:     str    = ""
    amount:           float  = 0.0
    cost_center:      str    = ""
    project:          str    = ""
    attendees:        str    = ""
    comments:         str    = ""


@dataclass
class Receipt:
    receipt_id: str   = ""
    order_id:   str   = ""
    date:       str   = ""
    vendor:     str   = ""
    amount:     float = 0.0
    summary:    str   = ""
    # Parsed extras
    ticket_number: str = ""
    passenger:     str = ""
    route:         str = ""


@dataclass
class EmployeeMemory:
    """All transaction + receipt data for one employee in one report."""
    employee_name:  str = ""
    employee_id:    str = ""
    report_name:    str = ""
    transactions:   List[Transaction] = field(default_factory=list)
    receipts:       List[Receipt]     = field(default_factory=list)
    source_file:    str = ""

    def receipt_for_transaction(self, txn: Transaction) -> Optional[Receipt]:
        """Best-effort receipt match by amount and vendor."""
        amt = abs(txn.amount)
        vendor_l = txn.vendor_desc.lower()
        for r in self.receipts:
            if abs(abs(r.amount) - amt) < 0.02:
                if vendor_l and vendor_l[:6] in r.vendor.lower():
                    return r
                return r  # amount match is sufficient
        return None


class TransactionMemory:
    """
    Loads all *memory* Excel files found in a directory (or a single file)
    and indexes them by employee ID and employee name for fast lookup.
    """

    def __init__(self):
        self._by_emp_id:   Dict[str, List[EmployeeMemory]] = {}
        self._by_name:     Dict[str, List[EmployeeMemory]] = {}
        self._all:         List[EmployeeMemory]            = []

    # ── File loading ───────────────────────────────────────────────────────────
    def load_file(self, path: Path) -> bool:
        """Load one transaction memory Excel file. Returns True if successful."""
        path = Path(path)
        if not path.exists():
            log.warning("Transaction memory file not found: %s", path)
            return False
        try:
            em = self._parse_file(path)
            if em:
                self._index(em)
                log.info("Loaded memory: %s  (%s, %d txns, %d receipts)",
                         path.name, em.employee_id,
                         len(em.transactions), len(em.receipts))
                return True
        except Exception as exc:
            log.warning("Failed to parse memory file %s: %s", path.name, exc)
        return False

    def load_directory(self, directory: Path) -> int:
        """Load all .xlsx files in a directory. Returns count loaded."""
        directory = Path(directory)
        count = 0
        if not directory.is_dir():
            return 0
        for xlsx in sorted(directory.glob("*.xlsx")):
            if self.load_file(xlsx):
                count += 1
        log.info("TransactionMemory: loaded %d file(s) from %s", count, directory)
        return count

    # ── Lookup ─────────────────────────────────────────────────────────────────
    def get_by_employee_id(self, emp_id: str) -> List[EmployeeMemory]:
        return self._by_emp_id.get(str(emp_id or "").strip().upper(), [])

    def get_by_name(self, last_name: str, first_name: str = "") -> List[EmployeeMemory]:
        key = self._name_key(last_name, first_name)
        return self._by_name.get(key, [])

    def get_receipt_context(self, emp_id: str, amount: float,
                             vendor_desc: str = "") -> Optional[Receipt]:
        """Find the best matching receipt for a batch row."""
        memos = self.get_by_employee_id(emp_id)
        vendor_l = vendor_desc.lower()
        amt = abs(amount)
        for em in memos:
            for r in em.receipts:
                if abs(abs(r.amount) - amt) < 0.02:
                    if not vendor_l or vendor_l[:5] in r.vendor.lower():
                        return r
        return None

    def get_all_transactions(self, emp_id: str) -> List[Transaction]:
        txns = []
        for em in self.get_by_employee_id(emp_id):
            txns.extend(em.transactions)
        return txns

    # ── Internal ───────────────────────────────────────────────────────────────
    def _index(self, em: EmployeeMemory) -> None:
        self._all.append(em)
        eid = em.employee_id.upper()
        if eid:
            self._by_emp_id.setdefault(eid, []).append(em)
        if em.employee_name:
            parts = em.employee_name.replace(",", " ").split()
            if parts:
                key = self._name_key(parts[0], parts[-1] if len(parts) > 1 else "")
                self._by_name.setdefault(key, []).append(em)

    @staticmethod
    def _name_key(last: str, first: str = "") -> str:
        return f"{last.strip().upper()}|{first.strip().upper()}"

    @staticmethod
    def _ss(v, default: str = "") -> str:
        s = str(v).strip() if v is not None else ""
        return default if s.lower() in ("none", "nan", "") else s

    @staticmethod
    def _sf(v) -> float:
        try:
            return float(v or 0)
        except (ValueError, TypeError):
            return 0.0

    def _parse_file(self, path: Path) -> Optional[EmployeeMemory]:
        wb  = load_workbook(str(path), read_only=True, data_only=True)
        em  = EmployeeMemory(source_file=path.name)

        # ── Employee Report tab ───────────────────────────────────────────────
        if "Employee Report" in wb.sheetnames:
            ws = wb["Employee Report"]
            for row in ws.iter_rows(max_row=20, values_only=True):
                if row[1] == "Employee Name" and row[2]:
                    em.employee_name = self._ss(row[2])
                elif row[1] == "Employee ID" and row[2]:
                    em.employee_id = self._ss(row[2])
                elif row[1] == "Report Name" and row[2]:
                    em.report_name = self._ss(row[2])

        # ── Transactions tab ─────────────────────────────────────────────────
        if "Transactions" in wb.sheetnames:
            ws   = wb["Transactions"]
            rows = list(ws.iter_rows(values_only=True))
            # Find header row
            header_row = None
            for i, r in enumerate(rows):
                if r and "Transaction Id" in str(r[0] or ""):
                    header_row = i
                    break
            if header_row is not None:
                cols = {str(v or "").strip(): idx
                        for idx, v in enumerate(rows[header_row])
                        if v is not None}
                for r in rows[header_row + 1:]:
                    if r[0] is None:
                        continue
                    txn = Transaction(
                        txn_id=self._ss(r[cols.get("Transaction Id", 0)]),
                        date=self._ss(r[cols.get("Transaction Date", 1)]),
                        expense_type=self._ss(r[cols.get("Expense Type", 2)]),
                        business_purpose=self._ss(r[cols.get("Business Purpose", 3)]),
                        vendor_desc=self._ss(r[cols.get("Vendor Description", 4)]),
                        payment_type=self._ss(r[cols.get("Payment Type", 5)]),
                        amount=self._sf(r[cols.get("Amount", 6)]),
                        cost_center=self._ss(r[cols.get("Cost Center", 7)]),
                        project=self._ss(r[cols.get("Project", 8)]),
                        attendees=self._ss(r[cols.get("Attendees", 9)]),
                        comments=self._ss(r[cols.get("Comments", 10)]),
                    )
                    em.transactions.append(txn)

        # ── Receipts tab ─────────────────────────────────────────────────────
        if "Receipts" in wb.sheetnames:
            ws   = wb["Receipts"]
            rows = list(ws.iter_rows(values_only=True))
            header_row = None
            for i, r in enumerate(rows):
                if r and "Receipt Id" in str(r[0] or ""):
                    header_row = i
                    break
            if header_row is not None:
                cols = {str(v or "").strip(): idx
                        for idx, v in enumerate(rows[header_row])
                        if v is not None}
                for r in rows[header_row + 1:]:
                    if r[0] is None:
                        continue
                    summary = self._ss(r[cols.get("Summary", 5)])
                    receipt = Receipt(
                        receipt_id=self._ss(r[cols.get("Receipt Id", 0)]),
                        order_id=self._ss(r[cols.get("Order Id", 1)]),
                        date=self._ss(r[cols.get("Date", 2)]),
                        vendor=self._ss(r[cols.get("Vendor", 3)]),
                        amount=self._sf(r[cols.get("Amount", 4)]),
                        summary=summary,
                        ticket_number=self._extract_ticket(summary),
                        passenger=self._extract_passenger(summary),
                        route=self._extract_route(summary),
                    )
                    em.receipts.append(receipt)

        return em if (em.employee_id or em.employee_name) else None

    # ── Receipt parsing helpers ───────────────────────────────────────────────
    @staticmethod
    def _extract_ticket(summary: str) -> str:
        m = re.search(r"Ticket\s+Number[:\s]+([A-Z0-9\-]+)", summary, re.I)
        if m:
            return m.group(1).strip()
        m = re.search(r"([A-Z]{2}-\d{10,})", summary)
        return m.group(1) if m else ""

    @staticmethod
    def _extract_passenger(summary: str) -> str:
        m = re.search(r"Passenger\s+Name[:\s]+([A-Z\s]+?)(?:;|$)", summary, re.I)
        if m:
            return m.group(1).strip()
        return ""

    @staticmethod
    def _extract_route(summary: str) -> str:
        """Try to extract airport code route from receipt summary."""
        # Looks for patterns like EWR-SLC or JFK-ORD-LHR
        m = re.search(r"\b([A-Z]{3}[-/][A-Z]{3}(?:[-/][A-Z]{3})*)\b", summary)
        return m.group(1) if m else ""
