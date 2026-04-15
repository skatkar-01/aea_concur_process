"""
enhanced_scrubber.py — Smart Data Scrubbing with Transaction Context
======================================================================

Takes prepared XLSX data from reference_loader.py and applies intelligent
scrubbing rules based on detailed transaction context (attendees, project,
cost center, expense type, etc.).

Usage:
    python enhanced_scrubber.py \\
        --prepared prepared.xlsx \\
        --transaction-source ../final_concur_scrubbing/outputs/concur/03-26/ \\
        --output scrubbed.xlsx
"""

from __future__ import annotations

import argparse
import logging
import json
from pathlib import Path
from typing import Optional, Dict, List, Any
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("enhanced_scrubber")


# ══════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class TransactionContext:
    """Rich context for a transaction."""
    transaction_id: str
    date: str
    vendor: str
    vendor_desc: str
    amount: float
    expense_type: str
    business_purpose: str
    cost_center: str
    project: str
    attendees: List[str]  # ["Name1", "Name2", ...]
    attendee_count: int
    comments: str
    
    def to_prompt_context(self) -> str:
        """Generate context string for AI-based scrubbing."""
        return f"""
Transaction Context:
- ID: {self.transaction_id}
- Date: {self.date}
- Vendor: {self.vendor_desc}
- Amount: ${self.amount:.2f}
- Expense Type: {self.expense_type}
- Purpose: {self.business_purpose}
- Cost Center: {self.cost_center}
- Project: {self.project}
- Attendees: {', '.join(self.attendees)} ({self.attendee_count} people)
- Comments: {self.comments}
"""


@dataclass
class ScrubResult:
    """Result of scrubbing a transaction."""
    transaction_id: str
    original_vendor: str
    cleaned_vendor: str
    rule_applied: str
    flag_type: Optional[str] = None  # "ANOMALY", "DUPLICATE", "MISSING", etc.
    flag_reason: Optional[str] = None
    audit_trail: Dict[str, Any] = None


# ══════════════════════════════════════════════════════════════════════════════
#  SCRUBBING RULES
# ══════════════════════════════════════════════════════════════════════════════

class ScrubbingRules:
    """Business rules for data scrubbing."""
    
    # Vendor name standardization (fuzzy mapping)
    VENDOR_MAPPING = {
        "STARBKS": "Starbucks",
        "STARBUCKS COFFEE": "Starbucks",
        "SBUX": "Starbucks",
        "UNITED AIRLINES": "United Airlines",
        "UNITED AIR": "United Airlines",
        "UAL": "United Airlines",
        "DELTA AIR LINES": "Delta Air Lines",
        "DELTA": "Delta Air Lines",
        "AMERICAN AIRLINES": "American Airlines",
        "UBER": "Uber",
        "UBER TRIP": "Uber",
        "LYFT": "Lyft",
        "MARRIOTT": "Marriott",
        "MARRIOTT HOTELS": "Marriott",
        "HYATT": "Hyatt",
        "HILTON": "Hilton",
    }
    
    # Cost center validation rules
    VALID_COST_CENTERS = {
        "VAIP-VAIP Team",
        "Sales-East",
        "Sales-West",
        "Engineering-Seattle",
        "Marketing-HQ",
        "Operations-Denver",
    }
    
    # Project code validation (pattern)
    PROJECT_PATTERNS = {
        "1035": "AEA Strategy Meeting",
        "2041": "Q1 Planning",
        "3022": "Board Meeting",
        "4015": "Client Presentation",
    }
    
    @staticmethod
    def normalize_vendor(vendor_name: str) -> tuple[str, str]:
        """
        Normalize vendor name using fuzzy mapping.
        
        Returns:
            (normalized_name, rule_applied)
        """
        vendor_upper = vendor_name.upper().strip()
        
        # Exact match in mapping
        if vendor_upper in ScrubbingRules.VENDOR_MAPPING:
            return ScrubbingRules.VENDOR_MAPPING[vendor_upper], "exact_vendor_map"
        
        # Fuzzy match (substring)
        for original, normalized in ScrubbingRules.VENDOR_MAPPING.items():
            if original in vendor_upper or vendor_upper in original:
                return normalized, f"fuzzy_vendor_map:{original}"
        
        return vendor_name, "no_mapping_applied"
    
    @staticmethod
    def validate_expense_rules(ctx: TransactionContext) -> Optional[ScrubResult]:
        """
        Validate expense based on type and context.
        
        Returns:
            ScrubResult if rule violation found, else None
        """
        # Uber/Lyft rules
        if ctx.expense_type.upper() in ["TRANSPORTATION", "GROUND TRANSPORTATION"]:
            if any(x in ctx.vendor.upper() for x in ["UBER", "LYFT"]):
                # Rule: Uber/Lyft must have attendee list
                if not ctx.attendees or ctx.attendee_count == 0:
                    return ScrubResult(
                        transaction_id=ctx.transaction_id,
                        original_vendor=ctx.vendor,
                        cleaned_vendor=ctx.vendor,
                        rule_applied="uber_attendee_check",
                        flag_type="MISSING_ATTENDEES",
                        flag_reason=f"Uber/Lyft trip without attendee list (${ctx.amount:.2f})",
                    )
                
                # Rule: High amount Uber/Lyft (>$100) needs multiple attendees
                if ctx.amount > 100 and ctx.attendee_count < 2:
                    return ScrubResult(
                        transaction_id=ctx.transaction_id,
                        original_vendor=ctx.vendor,
                        cleaned_vendor=ctx.vendor,
                        rule_applied="uber_amount_attendee_check",
                        flag_type="ANOMALY",
                        flag_reason=f"Expensive Uber trip (${ctx.amount:.2f}) but only {ctx.attendee_count} attendee",
                    )
        
        # Airline rules
        if ctx.expense_type.upper() in ["AIRLINE", "AIRFARE"]:
            # Rule: Airline expense must have project code
            if not ctx.project or "1035" not in ctx.project:
                return ScrubResult(
                    transaction_id=ctx.transaction_id,
                    original_vendor=ctx.vendor,
                    cleaned_vendor=ctx.vendor,
                    rule_applied="airline_project_check",
                    flag_type="MISSING_PROJECT",
                    flag_reason=f"Airline expense without Strategy Meeting project code",
                )
        
        # Meal rules
        if ctx.expense_type.upper() in ["MEAL", "MEALS"]:
            # Rule: Meals >$50 should have multiple attendees
            if ctx.amount > 50 and ctx.attendee_count < 2:
                return ScrubResult(
                    transaction_id=ctx.transaction_id,
                    original_vendor=ctx.vendor,
                    cleaned_vendor=ctx.vendor,
                    rule_applied="meal_amount_check",
                    flag_type="ANOMALY",
                    flag_reason=f"Meal expense ${ctx.amount:.2f} but only {ctx.attendee_count} attendee",
                )
        
        return None
    
    @staticmethod
    def check_cost_center(ctx: TransactionContext) -> Optional[ScrubResult]:
        """Validate cost center is in approved list."""
        if ctx.cost_center not in ScrubbingRules.VALID_COST_CENTERS:
            return ScrubResult(
                transaction_id=ctx.transaction_id,
                original_vendor=ctx.vendor,
                cleaned_vendor=ctx.vendor,
                rule_applied="cost_center_validation",
                flag_type="INVALID_COST_CENTER",
                flag_reason=f"Cost center '{ctx.cost_center}' not in approved list",
            )
        return None
    
    @staticmethod
    def check_project(ctx: TransactionContext) -> Optional[ScrubResult]:
        """Validate project code matches pattern."""
        if not ctx.project:
            return None
        
        project_code = ctx.project.split("-")[0] if "-" in ctx.project else ctx.project
        
        if project_code not in ScrubbingRules.PROJECT_PATTERNS:
            return ScrubResult(
                transaction_id=ctx.transaction_id,
                original_vendor=ctx.vendor,
                cleaned_vendor=ctx.vendor,
                rule_applied="project_validation",
                flag_type="INVALID_PROJECT",
                flag_reason=f"Project code '{project_code}' not recognized",
            )
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  ENHANCED SCRUBBER
# ══════════════════════════════════════════════════════════════════════════════

class EnhancedScrubber:
    """
    Smart scrubber that reads prepared data and enriches with transaction context.
    """
    
    def __init__(self, prepared_xlsx: str, transaction_source: str):
        """
        Initialize scrubber.
        
        Args:
            prepared_xlsx: Path to prepared.xlsx from reference_loader
            transaction_source: Path to transaction output directory
        """
        self.prepared_path = Path(prepared_xlsx)
        self.transaction_dir = Path(transaction_source)
        
        if not self.prepared_path.exists():
            log.error(f"Prepared file not found: {self.prepared_path}")
            raise FileNotFoundError(self.prepared_path)
        
        if not self.transaction_dir.exists():
            log.error(f"Transaction directory not found: {self.transaction_dir}")
            raise FileNotFoundError(self.transaction_dir)
        
        self.prepared_wb = load_workbook(self.prepared_path)
        self.transaction_contexts = {}
        self.scrub_results = []
        
        log.info(f"Initialized scrubber")
        log.info(f"  Prepared file: {self.prepared_path}")
        log.info(f"  Transaction dir: {self.transaction_dir}")
    
    def load_transaction_contexts(self):
        """Load all transaction details from source files."""
        log.info("Loading transaction contexts...")
        
        for xlsx_file in self.transaction_dir.glob("*.xlsx"):
            try:
                txn_wb = load_workbook(xlsx_file)
                if "Transactions" not in txn_wb.sheetnames:
                    continue
                
                ws = txn_wb["Transactions"]
                cardholder_name = xlsx_file.stem.split(" ")[0]  # e.g., "Bales"
                
                # Parse transactions (skip headers)
                rows = list(ws.iter_rows(values_only=True))
                # Header is in row 3
                header_row = rows[2]
                
                for row_idx, row in enumerate(rows[3:], start=4):
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    # Extract fields
                    txn_id = row[0]
                    attendees_str = row[9] if len(row) > 9 else ""
                    attendees = [a.strip() for a in str(attendees_str).split(",") if a] if attendees_str else []
                    
                    ctx = TransactionContext(
                        transaction_id=txn_id,
                        date=str(row[1]) if len(row) > 1 else "",
                        vendor=str(row[4]) if len(row) > 4 else "",
                        vendor_desc=str(row[4]) if len(row) > 4 else "",
                        amount=float(row[6]) if len(row) > 6 and row[6] else 0.0,
                        expense_type=str(row[2]) if len(row) > 2 else "",
                        business_purpose=str(row[3]) if len(row) > 3 else "",
                        cost_center=str(row[7]) if len(row) > 7 else "",
                        project=str(row[8]) if len(row) > 8 else "",
                        attendees=attendees,
                        attendee_count=len(attendees),
                        comments=str(row[10]) if len(row) > 10 else "",
                    )
                    
                    self.transaction_contexts[txn_id] = ctx
                
                log.info(f"  Loaded {len([k for k in self.transaction_contexts.keys() if k.startswith(cardholder_name)])} transactions from {xlsx_file.name}")
            
            except Exception as e:
                log.warning(f"  Failed to load {xlsx_file.name}: {e}")
    
    def scrub_row(self, row_data: Dict[str, Any], txn_id: str = None) -> ScrubResult:
        """
        Scrub a single transaction row.
        
        Returns:
            ScrubResult with cleaned data and flags
        """
        vendor = str(row_data.get("vendor", ""))
        
        # Get transaction context
        ctx = self.transaction_contexts.get(txn_id)
        if not ctx:
            ctx = TransactionContext(
                transaction_id=txn_id,
                date="",
                vendor=vendor,
                vendor_desc=vendor,
                amount=0.0,
                expense_type="",
                business_purpose="",
                cost_center="",
                project="",
                attendees=[],
                attendee_count=0,
                comments="",
            )
        
        # Apply rules
        cleaned_vendor, rule = ScrubbingRules.normalize_vendor(vendor)
        result = ScrubResult(
            transaction_id=txn_id,
            original_vendor=vendor,
            cleaned_vendor=cleaned_vendor,
            rule_applied=rule,
        )
        
        # Check expense rules
        expense_flag = ScrubbingRules.validate_expense_rules(ctx)
        if expense_flag:
            result.flag_type = expense_flag.flag_type
            result.flag_reason = expense_flag.flag_reason
            result.rule_applied = expense_flag.rule_applied
        
        # Check cost center
        cost_center_flag = ScrubbingRules.check_cost_center(ctx)
        if cost_center_flag and not result.flag_type:
            result.flag_type = cost_center_flag.flag_type
            result.flag_reason = cost_center_flag.flag_reason
        
        # Check project
        project_flag = ScrubbingRules.check_project(ctx)
        if project_flag and not result.flag_type:
            result.flag_type = project_flag.flag_type
            result.flag_reason = project_flag.flag_reason
        
        self.scrub_results.append(result)
        return result
    
    def scrub_sheet(self, sheet_name: str) -> int:
        """
        Scrub all rows in a sheet.
        
        Returns:
            Number of rows scrubbed
        """
        if sheet_name not in self.prepared_wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' not found in prepared file")
            return 0
        
        ws = self.prepared_wb[sheet_name]
        rows_scrubbed = 0
        
        # Get header
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Find key columns
        vendor_col = None
        emp_id_col = None
        for idx, col_name in enumerate(header_row):
            if col_name and "Vendor" in str(col_name):
                vendor_col = idx
            if col_name and "Employee ID" in str(col_name):
                emp_id_col = idx
        
        # Scrub each row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vendor = row[vendor_col] if vendor_col and vendor_col < len(row) else ""
            emp_id = row[emp_id_col] if emp_id_col and emp_id_col < len(row) else ""
            
            result = self.scrub_row(
                {"vendor": vendor},
                txn_id=str(emp_id)
            )
            
            rows_scrubbed += 1
        
        return rows_scrubbed
    
    def write_scrubbed_output(self, output_path: str):
        """Write scrubbed data to new Excel file."""
        log.info(f"Writing scrubbed output to {output_path}")
        
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        
        # Copy sheets with applied formatting
        for sheet_name in self.prepared_wb.sheetnames:
            src_ws = self.prepared_wb[sheet_name]
            dst_ws = out_wb.create_sheet(sheet_name)
            
            # Copy data
            for row in src_ws.iter_rows(values_only=True):
                dst_ws.append(row)
            
            rows_scrubbed += self.scrub_sheet(sheet_name)
        
        # Write Audit Trail sheet
        audit_ws = out_wb.create_sheet("Audit Trail", 0)
        audit_ws.append(["Transaction ID", "Original Vendor", "Cleaned Vendor", "Rule Applied", "Flag Type", "Flag Reason"])
        
        # Red fill for flagged
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        for idx, result in enumerate(self.scrub_results, start=2):
            audit_ws.append([
                result.transaction_id,
                result.original_vendor,
                result.cleaned_vendor,
                result.rule_applied,
                result.flag_type or "",
                result.flag_reason or "",
            ])
            
            # Highlight flagged rows
            if result.flag_type:
                for cell in audit_ws[idx]:
                    cell.fill = red_fill
        
        out_wb.save(output_path)
        log.info(f"✓ Saved: {output_path}")
        log.info(f"  Scrubbed {len(self.scrub_results)} transactions")
        log.info(f"  Flagged for review: {sum(1 for r in self.scrub_results if r.flag_type)}")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Apply intelligent scrubbing rules to prepared transaction data"
    )
    parser.add_argument("--prepared", required=True, help="Path to prepared.xlsx from reference_loader")
    parser.add_argument("--transaction-source", required=True, help="Path to transaction output directory")
    parser.add_argument("--output", default="scrubbed.xlsx", help="Output file path")
    
    args = parser.parse_args()
    
    try:
        scrubber = EnhancedScrubber(args.prepared, args.transaction_source)
        scrubber.load_transaction_contexts()
        
        # Scrub each entity sheet
        for sheet_name in ["AEA_Posted", "SBF_Posted", "DEBT_Reviewed"]:
            rows = scrubber.scrub_sheet(sheet_name)
            if rows > 0:
                log.info(f"Scrubbed '{sheet_name}': {rows} rows")
        
        scrubber.write_scrubbed_output(args.output)
        
    except Exception as e:
        log.error(f"Scrubbing failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    import sys
    sys.exit(0 if main() else 1)
